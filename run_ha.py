#!/usr/bin/env python3
"""B6-2 均K（Heikin-Ashi）｜o3333o — 獨立進程
規格：
  1. 每根 K 線收線後 +5 秒抓 K 線，算 HA，判燈號。
  2. 進場：PRE根反轉前色 + POST根反轉後色 + POST振幅累加達門檻 -> taker 市價進場。
  3. 出場：EXIT根反向色 + 振幅累加達門檻 -> taker 市價平倉。無 TP/SL/TE。
  4. clOrdId 前綴：進場 h / 出場 y（普K 用 n / x，互不干擾）。
  5. 心跳：超過 3 根 K 線未推進燈號判定即 TG 告警。
  6. OKX 為唯一真相來源；重啟接管既有持倉；Telegram 為旁路，發送失敗不影響交易。
  7. 週期 5m/10m/15m/30m/60m/120m/240m/480m/720m/1440m；非原生者由底層 bar 合成。
     邊界以 UTC+8 對齊（與 OKX App 的 12H/1D 日界一致）。
注意：本檔完全獨立於 run_bot.py（普K），使用自己的 bot token 與 state 檔。
"""
import sys, hmac, base64, hashlib, json, time, asyncio, uuid, os, sqlite3
from decimal import Decimal, ROUND_FLOOR, ROUND_CEILING, ROUND_DOWN
from datetime import datetime, timezone, timedelta
import httpx
from telegram import BotCommand, BotCommandScopeDefault, BotCommandScopeAllPrivateChats, BotCommandScopeChat
from telegram.ext import Application, CommandHandler, MessageHandler, filters
sys.path.insert(0, "/srv/1111bot")
from app.core import emoji as E
from app.strategy.ha import calc_ha

BASE = "https://www.okx.com"
ACCT = "o3333o"
TZ8 = timezone(timedelta(hours=8))
ACCOUNT_TF = "5m"
STATE_FILE = "/srv/1111bot/data/strategies_ha_o3333o.json"
HA_LAG = 0           # 收線後幾秒開始抓（0＝收線瞬間就開始輪詢）
POLL_MS = 100        # 密集輪詢間隔（毫秒），直到 OKX 標記該根已收線
POLL_MAX = 12.0      # 密集輪詢最長等幾秒，逾時放棄本輪
COLLECT_LAG = 3      # 背景收集器的延後秒數（非交易用途，不必搶快）
HA_HIST = 120        # 抓幾根歷史 K 線做 HA 遞迴
HB_BARS = 3          # 心跳超過幾根 K 線未推進就告警
MOVE_TICK = 1.0      # 框架移動任務心跳（秒）；各策略依自己的間隔對齊整秒觸發
ALGO_PFX = "a"       # 止盈止損 algo 單的 clOrdId 前綴（與 h/y/n/x 區隔）
HA_MAX  = 2000       # /ha 單次最多抓幾根
DB_FILE = "/srv/1111bot/data/ha_market.db"
DB_KEEP = 30         # 每個 (幣種,週期) 在 DB 保留幾根（連續均K 很少超過30）
DB_WARM = HA_HIST    # 開機補歷史抓幾根（僅供算 ATR14 暖身與連續段，仍只存 DB_KEEP 根）
DB_TICK = 15         # 收集器每幾秒巡一次
DB_GAP  = 0.15       # 每次 API 之間間隔秒數（避免打到 OKX 限流）
HA_MAX  = 2000       # /ha 單次最多抓幾根
DB_FILE = "/srv/1111bot/data/ha_market.db"
DB_KEEP = 30         # 每個 (幣種,週期) 在 DB 保留幾根（連續均K 很少超過30）
DB_WARM = HA_HIST    # 開機補歷史抓幾根（僅供算 ATR14 暖身與連續段，仍只存 DB_KEEP 根）
DB_TICK = 15         # 收集器每幾秒巡一次
DB_GAP  = 0.15       # 每次 API 之間間隔秒數（避免打到 OKX 限流）
HA_INLINE = 30       # /ha 根數 <= 此值直接顯示在 TG，超過則產 Excel 寄信

# 均K 專屬週期表（不共用原K 的 normal.TF_SEC，兩邊互不影響）
# tf -> (秒數, OKX bar, 合成倍數)  合成倍數 >1 表示該週期非交易所原生，需自行合成
HA_TF = {
    "5m":    (300,    "5m", 1),
    "10m":   (600,    "5m", 2),
    "15m":   (900,   "15m", 1),
    "30m":   (1800,  "30m", 1),
    "60m":   (3600,   "1H", 1),
    "120m":  (7200,   "2H", 1),
    "240m":  (14400,  "4H", 1),
    "480m":  (28800,  "4H", 2),
    "720m":  (43200, "12H", 1),
    "1440m": (86400,  "1D", 1),
}
TF_LIST = ["5m", "10m", "15m", "30m", "60m", "120m", "240m", "480m", "720m", "1440m"]
# OKX 的 12H/1D K 線以 UTC+8 為日界；短週期能整除 8 小時，套用位移不影響結果
ALIGN_OFF = 8 * 3600

def tf_sec(tf):
    return HA_TF[tf][0]

def next_open_epoch(now_epoch, tf):
    """下一根 K 線的開盤 epoch（秒），以 UTC+8 對齊。"""
    sec = tf_sec(tf)
    n = int(now_epoch) + ALIGN_OFF
    return ((n // sec) + 1) * sec - ALIGN_OFF

def load_env(p):
    d = {}
    for line in open(p):
        line = line.strip()
        if "=" in line and not line.startswith("#"):
            k, v = line.split("=", 1); d[k] = v
    return d

ACC = load_env("/srv/1111bot/config/accounts.env")
BOTS = load_env("/srv/1111bot/config/bots.env")
TOKEN = BOTS["BOT_o3333o_HA"]
SYMS = json.load(open("/srv/1111bot/config/symbols.json"))["symbols"]

PENDING = {}; STRATS = {}; TASKS = {}; STATS = {}
CHAT_ID = None
HTTP = None
SPEC_CACHE = {}

def skey(s, d): return f"{s}_{d}"
def inst_id(s): return s.replace("USDT", "") + "-USDT-SWAP"
def now8(): return datetime.now(TZ8)
def hhmmss(): return now8().strftime("%H:%M:%S")
def hhmm(): return now8().strftime("%H:%M")
def today8(): return now8().strftime("%Y-%m-%d")
def pct(v): return format(Decimal(str(v)).normalize(), "f")   # 用 f 格式，避免 10 變成 1E+1

# ---------- 狀態持久化（原子寫入） ----------
SAVE_FIELDS = ("sym","dir","tf","lev","margin","pre","post",
               "tp_pct","sl_pct","move_pct","interval","chat",
               "pos_open","pos_px","pos_ee","pos_sz","pnl_hist","last_bar",
               "algo_id","tp_px","sl_px","frame_base","move_n","last_move",
               "in_lag","in_slip","in_ref","in_poll_n","in_poll_s")

SHUTTING_DOWN = False

def save_state(force=False):
    """force=True：寫入當下真實狀態（使用者主動停止時用，避免幽靈策略殘留存檔）。
    force=False：保留原保護，避免競態把仍在跑的策略誤清空。"""
    try:
        data = {"chat": CHAT_ID, "tf": ACCOUNT_TF, "stats": STATS, "strats": []}
        for k, S in STRATS.items():
            if S.get("alive"):
                data["strats"].append({a: S[a] for a in SAVE_FIELDS if a in S})
        if not force and STRATS and not data["strats"]:
            return
        def enc(o): return str(o) if isinstance(o, Decimal) else o
        tmp = STATE_FILE + ".tmp"
        with open(tmp, "w") as f:
            json.dump(data, f, default=enc); f.flush(); os.fsync(f.fileno())
        os.replace(tmp, STATE_FILE)
    except Exception as e:
        print("save_state fail", e)

def bump(k, field):
    t = today8()
    if k not in STATS or STATS[k]["date"] != t:
        STATS[k] = {"date": t, "placed": 0, "entered": 0}
    STATS[k][field] += 1
    save_state()

def get_stat(k):
    t = today8()
    if k not in STATS or STATS[k]["date"] != t: return (0, 0)
    return (STATS[k]["placed"], STATS[k]["entered"])

# ---------- 交易紀錄 ----------
def trade_file(t):
    return f"/srv/1111bot/data/trades_h_{ACCT}_" + str(t).replace("-", "") + ".json"

def load_trades(t):
    try: return json.load(open(trade_file(t)))
    except Exception: return []

def log_trade(rec):
    try:
        fp = trade_file(rec.get("date"))
        try: arr = json.load(open(fp))
        except Exception: arr = []
        arr.append(rec)
        tmp = fp + ".tmp"
        with open(tmp, "w") as f:
            json.dump(arr, f, default=str); f.flush(); os.fsync(f.fileno())
        os.replace(tmp, fp)
    except Exception as e:
        print("log_trade fail", e)

# ---------- OKX API ----------
def ts_now():
    n = datetime.now(timezone.utc)
    return n.strftime("%Y-%m-%dT%H:%M:%S.") + f"{n.microsecond//1000:03d}Z"

def sign(sec, ts, m, p, b=""):
    return base64.b64encode(hmac.new(sec.encode(), f"{ts}{m}{p}{b}".encode(), hashlib.sha256).digest()).decode()

async def api(method, path, body=None):
    b = json.dumps(body) if body else ""
    ts = ts_now()
    h = {"OK-ACCESS-KEY": ACC[f"OKX_{ACCT}_API_KEY"],
         "OK-ACCESS-SIGN": sign(ACC[f"OKX_{ACCT}_SECRET"], ts, method, path, b),
         "OK-ACCESS-TIMESTAMP": ts,
         "OK-ACCESS-PASSPHRASE": ACC[f"OKX_{ACCT}_PASSPHRASE"],
         "Content-Type": "application/json"}
    try:
        r = await HTTP.request(method, BASE + path, headers=h, content=b)
        return r.json()
    except Exception as e:
        print("api fail", path, type(e).__name__, e)
        return {"code": "-1", "msg": str(e), "data": []}

async def pub(path):
    try:
        r = await HTTP.get(BASE + path)
        return r.json()
    except Exception as e:
        print("pub fail", path, type(e).__name__, e)
        return {"code": "-1", "msg": str(e), "data": []}

async def get_spec(s):
    if s in SPEC_CACHE: return SPEC_CACHE[s]
    iid = inst_id(s)
    r = await pub(f"/api/v5/public/instruments?instType=SWAP&instId={iid}")
    d = r["data"][0]
    spec = {"iid": iid, "tick": Decimal(d["tickSz"]), "lot": Decimal(d["lotSz"]),
            "minsz": Decimal(d["minSz"]), "ctval": Decimal(d["ctVal"]),
            "maxlev": Decimal(d["lever"]), "ctvalccy": d["ctValCcy"]}
    SPEC_CACHE[s] = spec
    return spec

async def get_last(iid):
    r = await pub(f"/api/v5/market/ticker?instId={iid}")
    return Decimal(r["data"][0]["last"])

def align(px, tick, d):
    """價格對齊 tickSz：d="L" 向下取、d="S" 向上取（保守方向）。"""
    return (Decimal(str(px)) / tick).to_integral_value(
        rounding=ROUND_FLOOR if d == "L" else ROUND_CEILING) * tick

def csize(m, lev, px, cv, lot):
    return ((m * lev / px) / cv / lot).to_integral_value(rounding=ROUND_DOWN) * lot

# ---------- K 線 / HA ----------
async def get_klines(iid, bar, limit=HA_HIST):
    """只取已收線（confirm=1）的 K 線，回傳舊->新。"""
    r = await pub(f"/api/v5/market/candles?instId={iid}&bar={bar}&limit={min(300, limit)}")
    if r.get("code") != "0": return []
    out = []
    for c in (r.get("data") or []):
        try:
            if len(c) >= 9 and str(c[8]) != "1": continue
            out.append({"ts": int(c[0]), "o": Decimal(c[1]), "h": Decimal(c[2]),
                        "l": Decimal(c[3]), "c": Decimal(c[4])})
        except Exception:
            continue
    out.reverse()
    return out

async def get_klines_paged(iid, bar, want):
    """翻頁抓已收線 K 線（舊->新）。
    /market/candles 只保留最近約 1440 根，翻不動時自動改用 /market/history-candles。"""
    out = {}
    after = None
    ep = "candles"
    for _ in range(40):                      # 上限保護，避免無限翻頁
        q = f"/api/v5/market/{ep}?instId={iid}&bar={bar}&limit=300"
        if after is not None:
            q += f"&after={after}"
        r = await pub(q)
        data = (r.get("data") or []) if r.get("code") == "0" else []
        if not data:
            if ep == "candles":
                ep = "history-candles"; continue
            break
        oldest = None
        for cd in data:
            try:
                ts = int(cd[0])
                oldest = ts if oldest is None else min(oldest, ts)
                if len(cd) >= 9 and str(cd[8]) != "1":
                    continue
                if ts in out:
                    continue
                out[ts] = {"ts": ts, "o": Decimal(cd[1]), "h": Decimal(cd[2]),
                           "l": Decimal(cd[3]), "c": Decimal(cd[4])}
            except Exception:
                continue
        if len(out) >= want:
            break
        if oldest is None or (after is not None and oldest >= after):
            if ep == "candles":
                ep = "history-candles"; after = oldest or after; continue
            break
        after = oldest
        await asyncio.sleep(0.15)
    kl = [out[t] for t in sorted(out)]
    return kl[-want:] if len(kl) > want else kl

async def klines_paged_for_tf(iid, tf, want):
    """依週期抓 want 根（非原生週期先抓底層 bar 再合成）。"""
    sec, bar, mul = HA_TF[tf]
    if mul == 1:
        return await get_klines_paged(iid, bar, want)
    kl = await get_klines_paged(iid, bar, want * mul + mul * 2)
    kl = align_head(kl, sec * 1000, mul)
    kl = merge_n(kl, mul)
    return kl[-want:] if len(kl) > want else kl

def align_head(kl, period_ms, n):
    """丟掉開頭沒對齊合成邊界的零星根。最多丟 n-1 根；丟完仍對不上就是資料異常，
    直接回空並記錄，避免無聲把整串資料清光。"""
    dropped = 0
    while kl and int(kl[0]["ts"]) % period_ms != 0 and dropped < n:
        kl.pop(0); dropped += 1
    if kl and int(kl[0]["ts"]) % period_ms != 0:
        print("align_head: 時間戳無法對齊合成邊界", kl[0]["ts"], period_ms)
        return []
    return kl

def merge_n(kl, n):
    """把 n 根合成 1 根（舊->新）。用於非原生週期：10m=2x5m、480m=2x4H。"""
    out = []
    for i in range(0, len(kl) - n + 1, n):
        g = kl[i:i+n]
        out.append({"ts": g[0]["ts"], "o": g[0]["o"],
                    "h": max(x["h"] for x in g), "l": min(x["l"] for x in g),
                    "c": g[-1]["c"]})
    return out

async def get_klines_paged(iid, bar, want):
    """往回翻頁抓已收線 K 線，回傳舊->新。
    candles 只保留最近約 1440 根，翻不動時自動改用 history-candles 續抓。"""
    got = {}
    after = None
    ep = "candles"
    for _ in range(40):
        if len(got) >= want: break
        q = f"/api/v5/market/{ep}?instId={iid}&bar={bar}&limit=300"
        if after: q += f"&after={after}"
        r = await pub(q)
        if r.get("code") != "0":
            if ep == "candles":
                ep = "history-candles"; continue
            break
        data = r.get("data") or []
        if not data:
            if ep == "candles":
                ep = "history-candles"; continue
            break
        new = 0
        for c in data:
            try:
                if len(c) >= 9 and str(c[8]) != "1": continue
                t = int(c[0])
                if t in got: continue
                got[t] = {"ts": t, "o": Decimal(c[1]), "h": Decimal(c[2]),
                          "l": Decimal(c[3]), "c": Decimal(c[4])}
                new += 1
            except Exception:
                continue
        after = data[-1][0]          # 該頁最舊的 ts，用來繼續往回翻
        if new == 0:
            if ep == "candles":
                ep = "history-candles"; continue
            break
        await asyncio.sleep(0.12)
    out = [got[t] for t in sorted(got)]
    return out[-want:] if len(out) > want else out

async def klines_paged_for_tf(iid, tf, want):
    """依週期翻頁抓 K 線；非原生週期先抓底層 bar 再對齊合成。"""
    sec, bar, mul = HA_TF[tf]
    if mul == 1:
        return await get_klines_paged(iid, bar, want)
    kl = await get_klines_paged(iid, bar, want * mul + mul * 4)
    pms = sec * 1000
    while kl and int(kl[0]["ts"]) % pms != 0:
        kl.pop(0)
    kl = merge_n(kl, mul)
    return kl[-want:] if len(kl) > want else kl

async def klines_and_ha(iid, tf, limit=HA_HIST):
    """回傳 (原始K線, HA序列)，兩者索引一一對應，皆為舊->新。
    非原生週期先抓底層 bar 再對齊邊界合成。"""
    sec, bar, mul = HA_TF[tf]
    if mul == 1:
        kl = await get_klines(iid, bar, limit)
    else:
        kl = await get_klines(iid, bar, min(300, limit * mul + mul * 2))
        kl = align_head(kl, sec * 1000, mul)
        kl = merge_n(kl, mul)
    if not kl: return [], []
    return kl, calc_ha(kl)

async def ha_series(iid, tf, limit=HA_HIST):
    """只要 HA 序列時用這個。"""
    kl, ha = await klines_and_ha(iid, tf, limit)
    return ha

def calc_atr(kl, period=14):
    """在【原始 K 線】上算 ATR14（Wilder 平滑）與 ATR14 ratio(%)。
    技術指標一律用正統 K 線，不使用 HA 平滑值。
    TR = max(h-l, |h-prev_c|, |l-prev_c|)
    前 period 根取 TR 簡單平均為種子，之後 ATR = (前ATR*(n-1) + TR)/n
    回傳與 kl 等長的 [(atr14, atr14_ratio), ...]，資料不足處為 (None, None)。"""
    n = len(kl)
    out = [(None, None)] * n
    if n == 0: return out
    trs = []
    for i, k in enumerate(kl):
        if i == 0:
            tr = k["h"] - k["l"]
        else:
            pc = kl[i-1]["c"]
            tr = max(k["h"] - k["l"], abs(k["h"] - pc), abs(k["l"] - pc))
        trs.append(tr)
    if n < period: return out
    prev = sum(trs[:period], Decimal(0)) / Decimal(period)
    c0 = kl[period-1]["c"]
    out[period-1] = (prev, (prev / c0 * 100) if c0 else None)
    for i in range(period, n):
        prev = (prev * Decimal(period - 1) + trs[i]) / Decimal(period)
        ci = kl[i]["c"]
        out[i] = (prev, (prev / ci * 100) if ci else None)
    return out

# ==================== 行情資料庫（決策時只讀 DB，不打 API）====================
# 設計目的：均K 是趨勢型策略，需要多根歷史才能判斷。若等到訊號當下才抓 K 線，
# 會多花數百毫秒甚至更久。收集器在每根 K 線收線後就把結果算好寫入 SQLite，
# 下單時只做一次本機讀取。
DB = None
COLLECT_LAST = {}    # (sym, tf) -> 最近已寫入的 bar ts
COLLECT_ERR = {}     # (sym, tf) -> 最近一次錯誤訊息

DDL = """
CREATE TABLE IF NOT EXISTS ha_bars (
  sym TEXT NOT NULL,
  tf  TEXT NOT NULL,
  ts  INTEGER NOT NULL,
  dt  TEXT,
  color TEXT,
  dir INTEGER,
  ha_o REAL, ha_h REAL, ha_l REAL, ha_c REAL,
  body_pct REAL,
  range_pct REAL,
  chg_pct REAL,
  o REAL, h REAL, l REAL, c REAL,
  atr14 REAL,
  atr14_ratio REAL,
  streak INTEGER,
  streak_body REAL,
  streak_range REAL,
  updated INTEGER,
  PRIMARY KEY (sym, tf, ts)
);
CREATE INDEX IF NOT EXISTS idx_ha_bars_lookup ON ha_bars (sym, tf, ts DESC);
CREATE TABLE IF NOT EXISTS ha_meta (
  sym TEXT NOT NULL,
  tf  TEXT NOT NULL,
  last_ts INTEGER,
  last_run INTEGER,
  bars INTEGER,
  err TEXT,
  PRIMARY KEY (sym, tf)
);
"""

def db_open():
    global DB
    os.makedirs(os.path.dirname(DB_FILE), exist_ok=True)
    DB = sqlite3.connect(DB_FILE, timeout=10, check_same_thread=False)
    DB.row_factory = sqlite3.Row
    DB.execute("PRAGMA journal_mode=WAL")      # 讀寫不互鎖
    DB.execute("PRAGMA synchronous=NORMAL")
    DB.executescript(DDL)
    DB.commit()
    n = DB.execute("SELECT COUNT(*) FROM ha_bars").fetchone()[0]
    print(f"行情DB 就緒：{DB_FILE}（現有 {n} 筆）")

def _f(v):
    try: return float(v)
    except Exception: return None

def build_rows(sym, tf, kl, ha, atrs):
    """把 K 線 / HA / ATR 併成一列列可寫入 DB 的紀錄，並算好連續段累計。"""
    rows = []
    streak = 0; sbody = 0.0; srange = 0.0; prev_color = None
    now = int(time.time())
    for i, x in enumerate(ha):
        k = kl[i]
        ho, hh, hl, hc = x["ho"], x["hh"], x["hl"], x["hc"]
        body = float((hc - ho) / ho * 100) if ho else 0.0
        rng = float((hh - hl) / k["c"] * 100) if k["c"] else 0.0
        if i == 0:
            chg = 0.0
        else:
            pc = ha[i-1]["hc"]
            chg = float((hc - pc) / pc * 100) if pc else 0.0
        if x["color"] == prev_color:
            streak += 1; sbody += body; srange += rng
        else:
            streak = 1; sbody = body; srange = rng
        prev_color = x["color"]
        a, r = atrs[i]
        # 一律用 K 線「開盤」時間，方便與交易所圖表對照
        dt = datetime.fromtimestamp(int(x["ts"]) / 1000, TZ8).strftime("%Y-%m-%d %H:%M")
        rows.append((sym, tf, int(x["ts"]), dt, x["color"],
                     1 if x["color"] == "G" else -1,
                     _f(ho), _f(hh), _f(hl), _f(hc),
                     body, rng, chg,
                     _f(k["o"]), _f(k["h"]), _f(k["l"]), _f(k["c"]),
                     _f(a) if a is not None else None,
                     _f(r) if r is not None else None,
                     streak, sbody, srange, now))
    return rows

def db_write(sym, tf, rows, err=None):
    if DB is None or not rows: return
    rows = rows[-DB_KEEP:]          # 前面幾根只是 ATR14/連續段的暖身，不必落地
    DB.executemany(
        "INSERT OR REPLACE INTO ha_bars VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", rows)
    DB.execute("DELETE FROM ha_bars WHERE sym=? AND tf=? AND ts NOT IN "
               "(SELECT ts FROM ha_bars WHERE sym=? AND tf=? ORDER BY ts DESC LIMIT ?)",
               (sym, tf, sym, tf, DB_KEEP))
    n = DB.execute("SELECT COUNT(*) FROM ha_bars WHERE sym=? AND tf=?", (sym, tf)).fetchone()[0]
    DB.execute("INSERT OR REPLACE INTO ha_meta VALUES (?,?,?,?,?,?)",
               (sym, tf, rows[-1][2], int(time.time()), n, err))
    DB.commit()

def db_latest(sym, tf, n=1):
    """取最近 n 根（回傳舊->新的 dict 清單）。決策時就是呼叫這個，不打 API。"""
    if DB is None: return []
    cur = DB.execute("SELECT * FROM ha_bars WHERE sym=? AND tf=? ORDER BY ts DESC LIMIT ?",
                     (sym, tf, n))
    return [dict(r) for r in cur.fetchall()][::-1]

def db_meta(sym=None, tf=None):
    if DB is None: return []
    q = "SELECT * FROM ha_meta"; a = []
    if sym: q += " WHERE sym=?"; a.append(sym)
    if sym and tf: q += " AND tf=?"; a.append(tf)
    cur = DB.execute(q + " ORDER BY sym, tf", a)
    return [dict(r) for r in cur.fetchall()]

def db_fresh(sym, tf, tol=2):
    """DB 是否新鮮：最近一根是否就是剛收線那根（容許落後 tol 根）。"""
    m = db_meta(sym, tf)
    if not m or not m[0].get("last_ts"): return False
    sec = tf_sec(tf)
    want = next_open_epoch(int(time.time()), tf) - 2 * sec
    return int(m[0]["last_ts"]) >= (want - tol * sec) * 1000

async def collect_one(sym, tf, warm=False, fast=False):
    """抓一個 (幣種,週期) 並寫入 DB。回傳寫入根數，失敗回 0。
    fast=True 時只抓最少必要根數，縮短傳輸時間（下單前的密集輪詢用）。"""
    try:
        spec = await get_spec(sym)
    except Exception as e:
        COLLECT_ERR[(sym, tf)] = f"spec:{type(e).__name__}"; return 0
    try:
        lim = DB_WARM if warm else (60 if fast else HA_HIST)
        kl, ha = await klines_and_ha(spec["iid"], tf, lim)
        if not ha:
            COLLECT_ERR[(sym, tf)] = "無K線"; return 0
        atrs = calc_atr(kl, 14)
        rows = build_rows(sym, tf, kl, ha, atrs)
        db_write(sym, tf, rows)
        COLLECT_LAST[(sym, tf)] = rows[-1][2]
        COLLECT_ERR.pop((sym, tf), None)
        return len(rows)
    except Exception as e:
        msg = f"{type(e).__name__}: {e}"
        COLLECT_ERR[(sym, tf)] = msg
        try: db_write(sym, tf, [], err=msg)
        except Exception: pass
        print("collect fail", sym, tf, msg)
        return 0

def db_symbols():
    return [s["symbol"] for s in SYMS if s.get("enabled")]

async def collector(app):
    """背景收集器：每根 K 線收線後把該週期所有幣種算好寫入 DB。"""
    await asyncio.sleep(3)
    syms = db_symbols()
    print(f"行情收集器啟動：{len(syms)} 幣種 x {len(TF_LIST)} 週期")
    # 開機先補歷史
    ok = 0
    for tf in TF_LIST:
        for sym in syms:
            if await collect_one(sym, tf, warm=True): ok += 1
            await asyncio.sleep(DB_GAP)
    print(f"行情DB 初始化完成：{ok}/{len(syms)*len(TF_LIST)} 組（不發 TG，用 /db 查看）")
    while True:
        try:
            await asyncio.sleep(DB_TICK)
            now = time.time()
            for tf in TF_LIST:
                sec = tf_sec(tf)
                closed = next_open_epoch(int(now), tf) - 2 * sec   # 最後一根已收線的開盤 epoch
                if now < closed + sec + COLLECT_LAG:               # 還沒到可抓的時間
                    continue
                want = closed * 1000
                todo = [s for s in db_symbols() if COLLECT_LAST.get((s, tf), 0) < want]
                if not todo: continue
                for sym in todo:
                    await collect_one(sym, tf)
                    await asyncio.sleep(DB_GAP)
        except asyncio.CancelledError:
            raise
        except Exception as e:
            print("collector loop error", type(e).__name__, e)
            await asyncio.sleep(5)

async def get_klines_live(iid, bar, limit=HA_HIST):
    """取 K 線並保留「進行中」那根（confirm=0），標記 live=True。"""
    r = await pub(f"/api/v5/market/candles?instId={iid}&bar={bar}&limit={min(300, limit)}")
    if r.get("code") != "0": return []
    out = []
    for c in (r.get("data") or []):
        try:
            out.append({"ts": int(c[0]), "o": Decimal(c[1]), "h": Decimal(c[2]),
                        "l": Decimal(c[3]), "c": Decimal(c[4]),
                        "live": (len(c) >= 9 and str(c[8]) != "1")})
        except Exception:
            continue
    out.reverse()
    return out

async def live_rows(sym, tf, need):
    """回傳最近 need 根（最後一根是進行中的即時快照）。
    只給 /status 預覽用；真正的進出場判斷仍只用已收線資料。"""
    try:
        spec = await get_spec(sym)
    except Exception:
        return []
    sec, bar, mul = HA_TF[tf]
    if mul == 1:
        kl = await get_klines_live(spec["iid"], bar, HA_HIST)
    else:
        raw = await get_klines_live(spec["iid"], bar, min(300, HA_HIST * mul + mul * 2))
        pms = sec * 1000
        while raw and int(raw[0]["ts"]) % pms != 0:
            raw.pop(0)
        kl = []
        for i in range(0, len(raw), mul):
            g = raw[i:i+mul]
            kl.append({"ts": g[0]["ts"], "o": g[0]["o"],
                       "h": max(x["h"] for x in g), "l": min(x["l"] for x in g),
                       "c": g[-1]["c"],
                       "live": (len(g) < mul) or any(x.get("live") for x in g)})
    if len(kl) < 15: return []
    ha = calc_ha(kl)
    atrs = calc_atr(kl, 14)
    rows = []
    for i, x in enumerate(ha):
        a, rr = atrs[i]
        ho, hc = x["ho"], x["hc"]
        body = float((hc - ho) / ho * 100) if ho else 0.0
        lv = kl[i].get("live", False)
        # 一律用該根 K 線自己的開盤時間，方便與交易所圖表對照；
        # 進行中那根由 bar_line 於時間後補上 * 記號
        dt = datetime.fromtimestamp(int(x["ts"]) / 1000, TZ8).strftime("%Y-%m-%d %H:%M:00")
        rows.append({"ts": int(x["ts"]), "dt": dt, "color": x["color"],
                     "body_pct": body, "c": float(kl[i]["c"]),
                     "atr14_ratio": float(rr) if rr is not None else None,
                     "live": lv})
    return rows[-need:] if len(rows) > need else rows

# ---------- Telegram（旁路） ----------
_BG = set()

async def _send_bg(app, chat, t):
    try:
        await app.bot.send_message(chat, t)
    except Exception as e:
        print("notify fail", type(e).__name__, e)

async def notify(app, chat, t):
    try:
        tk = asyncio.create_task(_send_bg(app, chat, t))
        _BG.add(tk); tk.add_done_callback(_BG.discard)
    except Exception as e:
        print("notify schedule fail", e)

async def reply(u, t):
    for i in range(2):
        try:
            await u.message.reply_text(t); return True
        except Exception as e:
            print("reply fail", i, type(e).__name__, e)
            await asyncio.sleep(2)
    return False

# ---------- OKX 事實查詢 ----------
async def okx_pos(iid, ps):
    r = await api("GET", "/api/v5/account/positions")
    if r.get("code") != "0": return None
    for p in (r.get("data") or []):
        if p.get("instId") == iid and p.get("posSide") == ps:
            try:
                if float(p.get("pos") or 0) != 0: return p
            except Exception: pass
    return None

async def okx_orders(iid=None, ps=None, prefix="h"):
    """只認均K 自己的前綴，絕不碰普K 的 n / x 單。"""
    r = await api("GET", "/api/v5/trade/orders-pending")
    if r.get("code") != "0": return []
    out = []
    for o in (r.get("data") or []):
        if iid and o.get("instId") != iid: continue
        if ps and o.get("posSide") != ps: continue
        cid = str(o.get("clOrdId") or "")
        if prefix and not cid.startswith(prefix): continue
        out.append(o)
    return out

async def sweep_h(iid=None, ps=None):
    """清掉均K 自己殘留的掛單（h/y 前綴）。市價單正常不會留單，此為保險。"""
    n = 0
    for pre in ("h", "y"):
        for o in await okx_orders(iid, ps, prefix=pre):
            cr = await api("POST", "/api/v5/trade/cancel-order",
                           {"instId": o["instId"], "ordId": o["ordId"]})
            if cr.get("code") == "0": n += 1
    return n

async def close_record(iid, ps, after_ms, tries=10):
    for i in range(tries):
        r = await api("GET", f"/api/v5/account/positions-history?instType=SWAP&instId={iid}&limit=10")
        if r.get("code") == "0":
            for p in (r.get("data") or []):
                if p.get("posSide") == ps and int(p.get("uTime") or 0) >= after_ms:
                    return p
        await asyncio.sleep(1)
    return None

async def notify_long(app, chat, head, lines, tail):
    """把長明細拆成多則 TG 訊息（單則上限 4096）。"""
    LIM = 3400
    head = [x for x in head if x != ""]
    lines = [x for x in lines if x != ""]
    tail = [x for x in tail if x != ""]
    buf = list(head); msgs = []
    for ln in lines:
        if sum(len(x) + 1 for x in buf) + len(ln) + 1 > LIM and len(buf) > len(head):
            msgs.append("\n".join(buf)); buf = list(head)
        buf.append(ln)
    buf += tail
    msgs.append("\n".join(buf))
    for i, m in enumerate(msgs):
        if len(msgs) > 1:
            m = f"（{i+1}/{len(msgs)}）\n" + m
        await notify(app, chat, m)
        await asyncio.sleep(0.3)

GRN = "\U0001F7E9"
RED = "\U0001F7E5"

def bar_line(r, want=None, v1=None, v2=None, sec=True):
    """一根明細：要求色 時間 實際色 值1｜值2（顏色不符則行尾標 ❌）
    預設 值1=本根 ATR14 ratio、值2=累計 ATR14 ratio。"""
    act = GRN if r.get("color") == "G" else RED
    a = r.get("atr14_ratio")
    n1 = v1 if v1 is not None else (("%.4f%%" % float(a)) if a is not None else "-")
    n2 = v2 if v2 is not None else "-"
    dt = str(r.get("dt") or "")
    hm_ = dt[11:19] if sec and len(dt) >= 19 else dt[11:16]
    lv = "*" if r.get("live") else ""
    if want is None:
        return f"{act}{hm_}{lv} {n1}｜{n2}"
    wl = GRN if want == "G" else RED
    bad = "\u274c" if r.get("color") != want else ""
    return f"{wl}{hm_}{lv}{act} {n1}｜{n2}{bad}"

def entry_lines(info, dr):
    """進場條件明細：前段 → 虛線 → 後段。
    每根顯示 本根ATR14r｜累計ATR14r（累計只在順向段內累加）。"""
    opp = "R" if dr == "L" else "G"
    want = "G" if dr == "L" else "R"
    L = [bar_line(x, opp) for x in info["pre_seg"]]
    L.append("\u2504" * 18)
    acc = 0.0
    for x in info["post_seg"]:
        v = float(x.get("atr14_ratio") or 0)
        acc += v
        L.append(bar_line(x, want, "%.4f%%" % v, "%.4f%%" % acc))
    return L

def judge_entry_db(rows, d, pre, post):
    """rows 為舊->新，至少 pre+post 根。
    進場條件（兩項同時成立）：
      1. 前 pre 根全為反向色（確認是真反轉，不是順勢中途）
      2. 後 post 根全為順勢色（做多綠、做空紅）
    ATR 仍照常計算並顯示，但目前不參與判定。"""
    need = pre + post
    if len(rows) < need: return None
    seg = rows[-need:]
    pre_seg, post_seg = seg[:pre], seg[pre:]
    want = "G" if d == "L" else "R"
    opp = "R" if d == "L" else "G"
    pre_ok = all(r.get("color") == opp for r in pre_seg)
    color_ok = all(r.get("color") == want for r in post_seg)
    atr_sum = sum(float(r.get("atr14_ratio") or 0) for r in post_seg)
    return {"pre_seg": pre_seg, "post_seg": post_seg,
            "pre_ok": pre_ok, "color_ok": color_ok, "atr_sum": atr_sum,
            "hit": pre_ok and color_ok}

# ---------- 止盈止損框架（OKX algo 單，交易所端保護） ----------
def frame_px(entry_px, d, tp_pct, sl_pct, tick):
    """由進場價算出初始止盈 / 止損觸發價，並對齊 tickSz。"""
    ep = Decimal(str(entry_px))
    if d == "L":
        tp = ep * (Decimal(1) + Decimal(str(tp_pct)) / 100)
        sl = ep * (Decimal(1) - Decimal(str(sl_pct)) / 100)
        return align(tp, tick, "S"), align(sl, tick, "L")
    tp = ep * (Decimal(1) - Decimal(str(tp_pct)) / 100)
    sl = ep * (Decimal(1) + Decimal(str(sl_pct)) / 100)
    return align(tp, tick, "L"), align(sl, tick, "S")

async def place_frame(S, spec, iid, d, pos, size, entry_px):
    """掛一張 OCO：止盈 + 止損同時存在，先觸發的平倉、另一邊由交易所自動取消。
    掛在交易所端，程式掛掉保護仍在。"""
    tp, sl = frame_px(entry_px, d, S["tp_pct"], S["sl_pct"], spec["tick"])
    cs = "sell" if d == "L" else "buy"
    body = {"instId": iid, "tdMode": "isolated", "side": cs, "posSide": pos,
            "ordType": "oco", "sz": str(size),
            "tpTriggerPx": str(tp), "tpOrdPx": "-1",
            "slTriggerPx": str(sl), "slOrdPx": "-1",
            "tpTriggerPxType": "last", "slTriggerPxType": "last",
            "algoClOrdId": ALGO_PFX + uuid.uuid4().hex[:14]}
    r = await api("POST", "/api/v5/trade/order-algo", body)
    if r.get("code") != "0":
        em = (r.get("data") or [{}])[0].get("sMsg") or r.get("msg")
        return None, str(tp), str(sl), em
    aid = (r.get("data") or [{}])[0].get("algoId")
    return aid, str(tp), str(sl), None

async def cancel_frame(iid, algo_id):
    """撤掉 algo 單。回傳是否成功。"""
    if not algo_id: return True
    r = await api("POST", "/api/v5/trade/cancel-algos",
                  [{"instId": iid, "algoId": str(algo_id)}])
    return r.get("code") == "0"

async def algo_pending(iid=None):
    """查未觸發的 algo 單（只認自己的前綴）。"""
    q = "/api/v5/trade/orders-algo-pending?ordType=oco"
    if iid: q += f"&instId={iid}"
    r = await api("GET", q)
    if r.get("code") != "0": return []
    out = []
    for o in (r.get("data") or []):
        if str(o.get("algoClOrdId") or "").startswith(ALGO_PFX):
            out.append(o)
    return out

def frame_shift(S, px, d):
    """依現價決定框架是否上移（做多）/下移（做空）。
    移動幅度＝相對上次基準價的漲跌幅，止盈止損整體同步平移。
    回傳 (新止盈, 新止損, 新基準, 移動幅度%) 或 None（不需移動）。"""
    try:
        base = Decimal(str(S.get("frame_base") or S.get("pos_px")))
        cur = Decimal(str(px))
        tp = Decimal(str(S["tp_px"])); sl = Decimal(str(S["sl_px"]))
    except Exception:
        return None
    if base <= 0: return None
    thr = Decimal(str(S["move_pct"])) / 100
    if d == "L":
        gain = (cur - base) / base
        if gain < thr: return None
    else:
        gain = (base - cur) / base
        if gain < thr: return None
    ratio = cur / base                      # 框架整體按同樣比例平移
    tick = S["spec"]["tick"]
    ntp = align(tp * ratio, tick, "S" if d == "L" else "L")
    nsl = align(sl * ratio, tick, "L" if d == "L" else "S")
    return ntp, nsl, cur, float(gain * 100)

async def amend_frames(items):
    """批次修改 algo 單的止盈止損觸發價。items: [(instId, algoId, tp, sl)]
    OKX 一次最多 10 筆，超過自動分批。回傳成功筆數。"""
    ok = 0
    for i in range(0, len(items), 10):
        chunk = items[i:i+10]
        body = [{"instId": a, "algoId": str(b),
                 "newTpTriggerPx": str(c), "newSlTriggerPx": str(e)}
                for a, b, c, e in chunk]
        r = await api("POST", "/api/v5/trade/amend-algos", body)
        if r.get("code") == "0":
            ok += len(chunk)
        else:
            print("amend-algos fail", r.get("code"), r.get("msg"),
                  (r.get("data") or [{}])[0].get("sMsg") if r.get("data") else "")
    return ok

# ---------- 進場 ----------
async def h_open(app, S, spec, iid, d, pos, info, k):
    # 用剛收線那根的收盤價估張數，省掉一次 ticker API（越少往返越貼近開盤價）
    ref_px = None; ref_ts = None
    try:
        lastrow = (info.get("post_seg") or [])[-1]
        ref_px = Decimal(str(lastrow["c"])); ref_ts = int(lastrow["ts"])
    except Exception:
        pass
    last = ref_px if ref_px else await get_last(iid)
    size = csize(S["margin"], Decimal(S["lev"]), last, spec["ctval"], spec["lot"])
    if size < spec["minsz"]:
        await notify(app, S["chat"], f"{E.BOT} {E.LOSS} {S['sym']} {E.dir_word(d)} 保證金不足，循環停止")
        S["alive"] = False
        return False
    r = await api("POST", "/api/v5/trade/order",
                  {"instId": iid, "tdMode": "isolated", "side": "buy" if d == "L" else "sell",
                   "posSide": pos, "ordType": "market", "sz": str(size),
                   "clOrdId": "h" + uuid.uuid4().hex[:14]})
    bump(k, "placed")
    if r.get("code") != "0":
        em = (r.get("data") or [{}])[0].get("sMsg") or r.get("msg")
        await notify(app, S["chat"], f"{E.BOT} {E.LOSS} {S['sym']} {E.dir_word(d)} 進場失敗：{em}")
        return False
    oid = r["data"][0]["ordId"]
    fpx = None
    for _ in range(8):
        st = await api("GET", f"/api/v5/trade/order?instId={iid}&ordId={oid}")
        if st.get("code") == "0" and st.get("data"):
            dd = st["data"][0]
            if dd.get("state") == "filled" and dd.get("avgPx"):
                fpx = Decimal(dd["avgPx"]); break
            if dd.get("state") == "canceled":
                await notify(app, S["chat"], f"{E.BOT} {E.LOSS} {S['sym']} {E.dir_word(d)} 市價單被取消")
                return False
        await asyncio.sleep(1)
    if fpx is None:
        # 查單沒回 filled 不代表沒成交；多查幾次持倉再判定，
        # 否則會出現「OKX 已有倉、程式卻以為空手」而下一根重複進場。
        p = None
        for _ in range(3):
            p = await okx_pos(iid, pos)
            if p: break
            await asyncio.sleep(2)
        if not p:
            await notify(app, S["chat"], f"{E.BOT} {E.LOSS} {S['sym']} {E.dir_word(d)} 進場未確認成交，跳過本輪\n⚠ 若 OKX 實際有倉請手動處理")
            return False
        fpx = Decimal(p.get("avgPx") or await get_last(iid))
        try:
            rsz = abs(Decimal(str(p.get("pos") or "0")))
            if rsz > 0: size = rsz
        except Exception:
            pass
    bump(k, "entered")
    ee = time.time()
    S["state"] = "持倉中"; S["pos_open"] = True
    S["pos_px"] = str(fpx); S["pos_ee"] = ee; S["pos_sz"] = str(size)
    S["pnl_hist"] = []                       # 逐根損益紀錄，供 /status 與出場核對
    S["frame_base"] = str(fpx); S["move_n"] = 0; S["last_move"] = ee
    save_state()
    # 成交後立刻在交易所端掛上止盈止損（OCO），程式掛掉保護仍在
    aid, tpx, spx, aerr = await place_frame(S, spec, iid, d, pos, size, fpx)
    S["algo_id"] = aid; S["tp_px"] = tpx; S["sl_px"] = spx
    save_state()
    if aerr:
        await notify(app, S["chat"],
            f"{E.BOT} {E.LOSS} {S['sym']} {E.dir_word(d)} 止盈止損掛單失敗：{aerr}\n"
            f"⚠ 目前是裸倉，請至 OKX 手動設定或平倉")
    head = [f"{E.BOT} OKX均K｜{ACCT}", "事件：🔔 訊號進場",
            f"{E.dir_emoji(d)} {d} {S['tf']}",
            strat_params(S["sym"], d) or ""]
    lines = entry_lines(info, d) if info else []
    ein = datetime.fromtimestamp(ee, TZ8).strftime("%H:%M:%S")
    lag = slip = None
    if ref_ts is not None:
        lag = t_send - (ref_ts / 1000 + tf_sec(S["tf"]))   # 以「送單時刻」計
    if ref_px:
        slip = (fpx - ref_px) / ref_px * 100
        if d == "S": slip = -slip
    # 存進策略狀態，出場時一併寫入交易紀錄（/replay 報表要用）
    _pn, _ps = (S.get("_poll") or (None, None))
    S["in_lag"] = round(lag, 3) if lag is not None else None
    S["in_slip"] = round(float(slip), 4) if slip is not None else None
    S["in_ref"] = str(ref_px) if ref_px else None
    S["in_poll_n"] = _pn
    S["in_poll_s"] = round(_ps, 3) if _ps is not None else None
    tail = ["━" * 10,
            f"進場{ein}｜{fpx}｜{size}張",
            (f"參考收盤{ref_px}｜偏離{slip:+.4f}%"
             + (f"｜延遲{lag:.2f}s" if lag is not None else "")) if ref_px else "",
            (f"取線輪詢{S['_poll'][0]}次／{S['_poll'][1]:.2f}s" if S.get("_poll") else ""),
            f"止盈{S.get('tp_px')}（{pct(S['tp_pct'])}%）",
            f"止損{S.get('sl_px')}（{pct(S['sl_pct'])}%）",
            f"每{S['interval']}秒檢查，達{pct(S['move_pct'])}%整體平移",
            f"時間：{hhmmss()}"]
    await notify_long(app, S["chat"], head, lines, tail)
    return True

# ---------- 重啟接管 ----------
async def h_takeover(app, S, spec, iid, d, pos):
    """重啟後接回既有持倉，並把交易所端的止盈止損單接回來。"""
    p = await okx_pos(iid, pos)
    if not p:
        for a in ("pos_open", "pos_px", "pos_ee", "pos_sz", "pnl_hist",
                  "algo_id", "tp_px", "sl_px", "frame_base", "move_n", "last_move",
                  "in_lag", "in_slip", "in_ref", "in_poll_n", "in_poll_s"):
            S.pop(a, None)
        save_state()
        return None
    fpx = Decimal(S.get("pos_px") or p.get("avgPx") or "0")
    ee = float(S.get("pos_ee") or time.time())
    size = abs(Decimal(p.get("pos") or "0"))
    S["state"] = "持倉中"
    try:
        for o in await algo_pending(iid):
            if o.get("posSide") == pos:
                S["algo_id"] = o.get("algoId")
                if o.get("tpTriggerPx"): S["tp_px"] = o["tpTriggerPx"]
                if o.get("slTriggerPx"): S["sl_px"] = o["slTriggerPx"]
                break
    except Exception as e:
        print("algo takeover fail", e)
    if not S.get("frame_base"): S["frame_base"] = str(fpx)
    save_state()
    await notify(app, S["chat"],
        f"{E.BOT} {E.dir_emoji(d)} {S['sym']} {E.dir_word(d)} 已接管既有持倉\n"
        f"進場價 {fpx}｜{size} 張\n"
        f"止盈{S.get('tp_px', '-')}／止損{S.get('sl_px', '-')}"
        + ("｜已接回交易所單" if S.get("algo_id") else "｜⚠ 交易所無止盈止損單"))
    return (size, fpx, ee)

# ---------- 主迴圈 ----------
async def wait_db_bar(sym, tf, want_ts):
    """收線瞬間就開始密集輪詢，直到 OKX 把那根標記為已收線（confirm=1）。
    回傳 (是否成功, 輪詢次數, 耗時秒)。這是下單延遲的主要來源，
    間隔設得很短是為了讓成交價盡量貼近下一根的開盤價。"""
    t0 = time.time()
    r = db_latest(sym, tf, 1)
    if r and int(r[-1]["ts"]) >= want_ts:
        return True, 0, 0.0
    n = 0
    while time.time() - t0 < POLL_MAX:
        n += 1
        await collect_one(sym, tf, fast=True)
        r = db_latest(sym, tf, 1)
        if r and int(r[-1]["ts"]) >= want_ts:
            return True, n, time.time() - t0
        await asyncio.sleep(POLL_MS / 1000.0)
    return False, n, time.time() - t0

async def hloop(app, chat, S):
    """均K 主迴圈：每根 K 收線後讀 DB 判燈號，決策不打 API。"""
    spec = S["spec"]; iid = spec["iid"]; d = S["dir"]
    pos = "long" if d == "L" else "short"
    k = skey(S["sym"], d)
    need = int(S["pre"]) + int(S["post"])
    try:
        if S.get("pos_open"):
            await h_takeover(app, S, spec, iid, d, pos)

        while S["alive"]:
            tf_sec_v = tf_sec(S["tf"])
            oe = next_open_epoch(int(time.time()), S["tf"])
            S["state"] = "持倉中" if S.get("pos_open") else "等訊號"
            save_state()
            # 分段睡眠：每 5 秒檢查一次 alive，/stop 後最多 5 秒收工
            while S["alive"]:
                w = oe + HA_LAG - time.time()
                if w <= 0: break
                await asyncio.sleep(min(5.0, w))
            if not S["alive"]: break

            want = (oe - tf_sec_v) * 1000
            ok, poll_n, poll_s = await wait_db_bar(S["sym"], S["tf"], want)
            S["_poll"] = (poll_n, poll_s)
            S["hb"] = time.time(); S["hb_warned"] = False
            if not ok:
                await notify(app, chat, f"{E.BOT} {E.LOSS} {S['sym']} {E.dir_word(d)} "
                                        f"{S['tf']} 行情DB 未更新，本輪跳過（不下單）")
                continue
            rows = db_latest(S["sym"], S["tf"], max(need, 2))
            if not rows: continue
            last = rows[-1]
            bar_ts = int(last["ts"])
            if bar_ts == S.get("last_bar"):
                continue                      # 這根已判過，不重複
            S["last_bar"] = bar_ts; save_state()

            if S.get("pos_open"):
                # --- 持倉中：出場已交給交易所端的止盈止損框架，這裡只記錄逐根損益 ---
                try:
                    ep2 = Decimal(str(S.get("pos_px") or "0"))
                    c2 = Decimal(str(last["c"]))
                    if ep2:
                        pnl2 = float((c2 - ep2) / ep2 * 100) if d == "L" \
                               else float((ep2 - c2) / ep2 * 100)
                        base2 = Decimal(str(rows[-2]["c"])) if len(rows) >= 2 else ep2
                        one2 = float((c2 - base2) / base2 * 100) if d == "L" \
                               else float((base2 - c2) / base2 * 100)
                        hist = S.get("pnl_hist") or []
                        hist.append({"dt": str(last.get("dt") or ""),
                                     "c": float(last["c"]),
                                     "pnl": round(pnl2, 4), "one": round(one2, 4),
                                     "color": last.get("color")})
                        S["pnl_hist"] = hist[-120:]
                        save_state()
                except Exception as e:
                    print("pnl_hist fail", e)
            else:
                # --- 空手：進場判斷 ---
                if len(rows) < need: continue
                r = judge_entry_db(rows, d, int(S["pre"]), int(S["post"]))
                if r and r["hit"]:
                    await h_open(app, S, spec, iid, d, pos, r, k)
    except asyncio.CancelledError:
        raise
    except Exception as e:
        print("hloop error", S.get("sym"), S.get("dir"), type(e).__name__, e)
        await notify(app, chat, f"{E.BOT} {E.LOSS} {S['sym']} {E.dir_word(d)} 循環錯誤：{type(e).__name__}: {e}")
    finally:
        S["state"] = "已停止"; S["alive"] = False
        try:
            nrm = await sweep_h(iid, pos)
            if nrm:
                await notify(app, chat, f"{E.BOT} {S['sym']} {E.dir_word(d)} 結束前清除殘留掛單 {nrm} 筆")
        except Exception as e:
            print("finally sweep fail", e)
        if not SHUTTING_DOWN:
            if STRATS.get(k) is S:
                STRATS.pop(k, None)
            try:
                me = asyncio.current_task()
            except Exception:
                me = None
            if TASKS.get(k) is None or TASKS.get(k) is me:
                TASKS.pop(k, None)
            save_state(True)
        try:
            if await okx_pos(iid, pos):
                await notify(app, chat, f"{E.BOT} ⚠ {S['sym']} {E.dir_word(d)} 已停止但仍有持倉，請至 OKX 處理")
        except Exception:
            pass

# ---------- 止盈止損框架移動 + 出場偵測 ----------
async def close_bookkeeping(app, S, reason):
    """偵測到倉位已不在（止盈或止損觸發）後的收尾：取真實損益、撤殘單、寫紀錄。"""
    spec = S["spec"]; iid = spec["iid"]; d = S["dir"]
    pos = "long" if d == "L" else "short"
    try:
        fpx = Decimal(str(S.get("pos_px") or "0"))
    except Exception:
        fpx = Decimal(0)
    ee = float(S.get("pos_ee") or time.time())
    size = Decimal(str(S.get("pos_sz") or "0"))
    t0 = int(ee * 1000)
    # 撤掉可能殘留的 algo 單（OCO 通常會自動撤另一邊，這裡保險）
    try:
        await cancel_frame(iid, S.get("algo_id"))
    except Exception as e:
        print("cancel_frame fail", e)
    ph = await close_record(iid, pos, t0)
    src = "OKX"
    if ph:
        g = Decimal(ph.get("pnl") or "0")
        fee = Decimal(ph.get("fee") or "0") + Decimal(ph.get("fundingFee") or "0")
        net = Decimal(ph.get("realizedPnl") or "0")
        xpx = Decimal(ph.get("closeAvgPx") or "0") or await get_last(iid)
        nv = Decimal(ph.get("openAvgPx") or fpx) * Decimal(ph.get("closeTotalPos") or size) * spec["ctval"]
    else:
        src = "估算"
        xpx = await get_last(iid)
        g = (xpx - fpx) * size * spec["ctval"] if d == "L" else (fpx - xpx) * size * spec["ctval"]
        fee = Decimal(0); net = g
        nv = fpx * size * spec["ctval"]
    gp = (g / nv * 100) if nv else Decimal(0)
    fp = (fee / nv * 100) if nv else Decimal(0)
    npv = (net / nv * 100) if nv else Decimal(0)
    hs = int(time.time() - ee)
    tfs = tf_sec(S["tf"])
    log_trade({"date": today8(), "sym": S["sym"], "dir": d, "reason": reason,
               "hold_s": hs, "bars": round(hs / tfs, 1),
               "gross": str(g), "fee": str(fee), "net": str(net), "nv": str(nv),
               "src": src, "ts": hhmmss(),
               "in_ts": datetime.fromtimestamp(ee, TZ8).strftime("%H:%M:%S"),
               "tf": S["tf"], "pre": str(S["pre"]), "post": str(S["post"]),
               "tp_pct": str(S["tp_pct"]), "sl_pct": str(S["sl_pct"]),
               "move_pct": str(S["move_pct"]), "interval": S["interval"],
               "move_n": S.get("move_n", 0),
               "tp_px": S.get("tp_px"), "sl_px": S.get("sl_px"),
               "lev": S["lev"], "margin": str(S["margin"]),
               "in_px": str(fpx), "out_px": str(xpx),
               "in_lag": S.get("in_lag"), "in_slip": S.get("in_slip"),
               "in_ref": S.get("in_ref"),
               "in_poll_n": S.get("in_poll_n"), "in_poll_s": S.get("in_poll_s")})
    hist = S.get("pnl_hist") or []
    ico = "🟢" if net >= 0 else "🔴"
    head = [f"{E.BOT} OKX均K｜{ACCT}", f"事件：{ico} 已出場",
            f"{E.dir_emoji(d)} {S['sym']} {d} {S['tf']}",
            strat_params(S["sym"], d) or "",
            f"進場{datetime.fromtimestamp(ee, TZ8).strftime('%H:%M')}｜{fpx}"
            f"→出場{hhmm()}｜{xpx}",
            f"出場原因：{reason}",
            f"持倉{hs}s（約{hs / tfs:.1f}根）｜框架移動{S.get('move_n', 0)}次", "━" * 10]
    lines = []
    show = hist[-40:]
    for i, hrec in enumerate(show, start=len(hist) - len(show) + 1):
        hm2 = str(hrec.get("dt") or "")[11:16]
        lg2 = "\U0001F7E9" if hrec.get("color") == "G" else "\U0001F7E5"
        lines.append(f"{i} {hm2}{lg2}本{hrec.get('one', 0):+.3f}% 累{hrec['pnl']:+.3f}%")
    tail = ["━" * 10,
            f"毛損益{g:+.6f}({gp:+.3f}%)",
            f"手續費{fee:+.6f}({fp:+.3f}%)",
            f"淨損益{net:+.6f}({npv:+.3f}%){E.pnl_emoji(net)}",
            f"時間：{hhmmss()}"]
    await notify_long(app, S["chat"], head, lines, tail)
    for a in ("pos_open", "pos_px", "pos_ee", "pos_sz", "pnl_hist",
              "algo_id", "tp_px", "sl_px", "frame_base", "move_n", "last_move",
              "in_lag", "in_slip", "in_ref", "in_poll_n", "in_poll_s"):
        S.pop(a, None)
    S["state"] = "等訊號"
    save_state(True)

async def frame_mover(app):
    """每秒對齊整秒；到了各策略的間隔就查一次全部持倉，
    需要平移的批次送出 amend-algos。同時偵測倉位消失（止盈/止損已觸發）。"""
    await asyncio.sleep(5)
    print("止盈止損框架移動任務已啟動")
    while True:
        try:
            await asyncio.sleep(MOVE_TICK - (time.time() % MOVE_TICK))
            now = int(time.time())
            held = [S for S in list(STRATS.values())
                    if S.get("alive") and S.get("pos_open")]
            if not held: continue
            due = [S for S in held if now % max(1, int(S.get("interval", 5))) == 0]
            if not due: continue
            r = await api("GET", "/api/v5/account/positions")   # 一次拿全部
            if r.get("code") != "0": continue
            live = {}
            for p in (r.get("data") or []):
                try:
                    if float(p.get("pos") or 0) == 0: continue
                except Exception:
                    continue
                sy = p["instId"].replace("-USDT-SWAP", "USDT")
                live[(sy, "L" if p["posSide"] == "long" else "S")] = p
            amends = []
            for S in due:
                key = (S["sym"], S["dir"])
                p = live.get(key)
                if p is None:
                    # 倉位不在了 -> 止盈或止損已觸發
                    try:
                        await close_bookkeeping(app, S, "Frame_Exit")
                    except Exception as e:
                        print("close_bookkeeping fail", S["sym"], type(e).__name__, e)
                    continue
                px = p.get("last") or p.get("markPx")
                if not px or not S.get("algo_id"): continue
                mv = frame_shift(S, px, S["dir"])
                if not mv: continue
                ntp, nsl, nbase, gain = mv
                S["_pending"] = (str(ntp), str(nsl), str(nbase), gain)
                amends.append((S["spec"]["iid"], S["algo_id"], ntp, nsl, S))
            if amends:
                items = [(a, b, c, d2) for a, b, c, d2, _ in amends]
                okn = await amend_frames(items)
                for a, b, c, d2, S in amends:
                    if okn:
                        ntp, nsl, nbase, gain = S.pop("_pending")
                        S["tp_px"] = ntp; S["sl_px"] = nsl
                        S["frame_base"] = nbase
                        S["move_n"] = int(S.get("move_n", 0)) + 1
                        S["last_move"] = time.time()
                    else:
                        S.pop("_pending", None)
                if okn: save_state()
        except asyncio.CancelledError:
            raise
        except Exception as e:
            print("frame_mover error", type(e).__name__, e)
            await asyncio.sleep(2)

# ---------- 心跳看門狗 ----------
async def hb_watch(app):
    while True:
        await asyncio.sleep(60)
        try:
            for k, S in list(STRATS.items()):
                if not S.get("alive"): continue
                hb = S.get("hb")
                if not hb: continue
                lim = tf_sec(S["tf"]) * HB_BARS
                gap = time.time() - hb
                if gap > lim and not S.get("hb_warned"):
                    S["hb_warned"] = True
                    await notify(app, S.get("chat") or CHAT_ID,
                        f"{E.BOT} {E.LOSS} 心跳異常\n"
                        f"{E.dir_emoji(S['dir'])} {S['sym']} {E.dir_word(S['dir'])}\n"
                        f"已 {int(gap)}s（>{HB_BARS} 根 {S['tf']}）未推進燈號判定。\n"
                        f"⚠ 均K 無停損，出場依賴程式運作，請確認服務狀態\n時間：{hhmmss()}")
                elif gap <= lim and S.get("hb_warned"):
                    S["hb_warned"] = False
                    await notify(app, S.get("chat") or CHAT_ID,
                        f"{E.BOT} ✅ 心跳恢復｜{S['sym']} {E.dir_word(S['dir'])}")
        except Exception as e:
            print("hb_watch fail", e)

# ---------- 啟動接管 ----------
async def rebuild_strat(d):
    spec = await get_spec(d["sym"])
    S = {"sym": d["sym"], "dir": d["dir"], "tf": d.get("tf", ACCOUNT_TF),
         "lev": int(d["lev"]), "margin": Decimal(str(d["margin"])),
         "pre": int(d["pre"]), "post": int(d["post"]),
         "tp_pct": Decimal(str(d["tp_pct"])),
         "sl_pct": Decimal(str(d["sl_pct"])),
         "move_pct": Decimal(str(d["move_pct"])),
         "interval": int(d["interval"]), "spec": spec,
         "alive": True, "state": "等訊號", "chat": d.get("chat", CHAT_ID),
         "hb": time.time(), "hb_warned": False}
    for a in ("pos_open", "pos_px", "pos_ee", "pos_sz", "pnl_hist", "last_bar",
              "algo_id", "tp_px", "sl_px", "frame_base", "move_n", "last_move",
              "in_lag", "in_slip", "in_ref", "in_poll_n", "in_poll_s"):
        if a in d: S[a] = d[a]
    return S

async def startup_recover(app):
    global CHAT_ID, ACCOUNT_TF, STATS
    if not os.path.exists(STATE_FILE):
        print("無存檔"); return
    try:
        data = json.load(open(STATE_FILE))
    except Exception as e:
        print("讀存檔失敗", e); return
    CHAT_ID = data.get("chat"); ACCOUNT_TF = data.get("tf", "5m"); STATS = data.get("stats", {})
    saved = data.get("strats", [])
    if not saved:
        print("存檔無策略"); return
    rec = []
    for d in saved:
        try:
            S = await rebuild_strat(d)
            k = skey(S["sym"], S["dir"])
            if k in STRATS:
                print("存檔有重複策略，略過", k); continue
            STRATS[k] = S
            TASKS[k] = asyncio.create_task(hloop(app, S["chat"], S))
            rec.append(f"{E.dir_emoji(S['dir'])} {S['sym']} {E.dir_word(S['dir'])}")
        except Exception as e:
            print("重建失敗", d, e)
    print(f"已接管均K 策略 {len(rec)}")
    if CHAT_ID and rec:
        await notify(app, CHAT_ID,
            f"{E.BOT} OKX均K｜{ACCT}\n事件：🔄 重啟認領完成\n━━━━━━━━━━\n"
            f"已接管策略（{len(rec)}）：\n" + "\n".join("・" + x for x in rec) +
            f"\n循環已接管，繼續運作\n時間：{hhmmss()}")

# ---------- TG 指令 ----------
def strat_params(sym, dr):
    """回傳與 /run 輸入完全相同的參數字串，方便直接複製重下。"""
    S = STRATS.get(skey(sym, dr))
    if not S or not S.get("alive"):
        return None
    return (f"/run {sym} {dr} {S['lev']}x {pct(S['margin'])} "
            f"{S['pre']} {S['post']} {pct(S['tp_pct'])}% {pct(S['sl_pct'])}% "
            f"{pct(S['move_pct'])}% {S['interval']}")

async def strat_detail(S, sym, dr, mark_px=None):
    """單一策略現況。空手列進場燈號，持倉列止盈止損框架。"""
    L = []
    tf = S["tf"]
    L.append(f"{E.dir_emoji(dr)} {sym} {dr} {tf}")
    L.append(strat_params(sym, dr) or "")
    if not S.get("pos_open"):
        pre = int(S["pre"]); post = int(S["post"])
        rows = await live_rows(sym, tf, pre + post)
        src_live = bool(rows) and len(rows) >= pre + post
        if not src_live:
            rows = db_latest(sym, tf, pre + post)
            if len(rows) < pre + post:
                L.append(f"狀態：⚠ 資料不足（{len(rows)}/{pre+post} 根）")
                return L
        r = judge_entry_db(rows, dr, pre, post)
        L.append("狀態：" + ("有機會進場（差收線）" if r and r["hit"] else "尚無符合條件"))
        if not src_live:
            L.append("⚠ 即時取價失敗，以下為最後已收線資料")
        L += entry_lines(r, dr)
        if src_live and rows[-1].get("live"):
            L.append("* 為進行中，收線前仍會變動")
        return L
    # ── 持倉中 ──
    L.append("狀態：📌 持倉中")
    now = hhmmss()
    ein = datetime.fromtimestamp(float(S.get("pos_ee") or time.time()),
                                 TZ8).strftime("%H:%M:%S")
    cur = None
    if mark_px:
        try: cur = Decimal(str(mark_px))
        except Exception: cur = None
    if cur is None:
        rr = db_latest(sym, tf, 1)
        if rr: cur = Decimal(str(rr[-1]["c"]))
    tp = S.get("tp_px"); sl = S.get("sl_px"); ep = S.get("pos_px")
    if dr == "L":
        L.append(f"止盈價：{tp}｜{now}")
        L.append(f"進場價：{ep}｜{ein}")
        L.append(f"止損價：{sl}｜{now}")
    else:
        L.append(f"止損價：{sl}｜{now}")
        L.append(f"進場價：{ep}｜{ein}")
        L.append(f"止盈價：{tp}｜{now}")
    L.append(f"目前價：{cur if cur is not None else '-'}｜{now}")
    if cur is not None and ep:
        try:
            e2 = Decimal(str(ep))
            pnl = (cur - e2) / e2 * 100 if dr == "L" else (e2 - cur) / e2 * 100
            L.append(f"未實現：{pnl:+.4f}%")
        except Exception:
            pass
    lm = S.get("last_move")
    lms = datetime.fromtimestamp(float(lm), TZ8).strftime("%H:%M:%S") if lm else "-"
    L.append(f"框架移動 {S.get('move_n', 0)} 次｜上次 {lms}｜每 {S['interval']}s 檢查")
    if not S.get("algo_id"):
        L.append("⚠ 交易所端無止盈止損單，目前是裸倉")
    hist = S.get("pnl_hist") or []
    if hist:
        show = hist[-5:]
        for h in show:
            hm_ = str(h.get("dt") or "")[11:16]
            lg = GRN if h.get("color") == "G" else RED
            L.append(f"{hm_}{lg}本{h.get('one', 0):+.3f}% 累{h['pnl']:+.3f}%")
    return L

def _pctarg(v):
    """必須帶 % 的參數。回傳 Decimal，格式不符回 None。"""
    v = str(v).strip()
    if not v.endswith("%"): return None
    try: return Decimal(v[:-1])
    except Exception: return None

async def cmd_run(u, c):
    global CHAT_ID; CHAT_ID = u.effective_chat.id
    a = c.args
    fmt = (f"用法：/run 商品 方向 槓桿x 資金 反向燈號數 順向燈號數 止盈% 止損% 移動門檻% 間隔秒\n"
           f"例：/run BTCUSDT L 1x 10 1 5 0.4% 0.2% 0.05% 5\n"
           f"━━━━━━━━━━\n"
           f"槓桿要帶 x；止盈/止損/移動門檻要帶 %\n"
           f"反向燈號數：做多=紅燈數，做空=綠燈數\n"
           f"順向燈號數：做多=綠燈數，做空=紅燈數\n"
           f"止盈止損掛在交易所端（OCO），程式掛掉仍有效\n"
           f"間隔秒：每幾秒查價一次，達門檻就把止盈止損框架整體平移（1~60）\n"
           f"週期依 /timeframe，目前 {ACCOUNT_TF}")
    if len(a) != 10:
        await reply(u, f"{E.BOT} 參數數量錯誤（需10個，收到{len(a)}個）\n{fmt}"); return
    sym = a[0].upper(); dr = a[1].upper()
    if dr not in ("L", "S"):
        await reply(u, f"{E.BOT} 方向須 L 或 S"); return
    lv = str(a[2]).strip().lower()
    if not lv.endswith("x"):
        await reply(u, f"{E.BOT} 槓桿要帶 x，例如 1x"); return
    try:
        lev = int(lv[:-1]); margin = Decimal(a[3])
        pre = int(a[4]); post = int(a[5]); interval = int(a[9])
    except Exception:
        await reply(u, f"{E.BOT} 參數格式錯誤\n{fmt}"); return
    tp_pct = _pctarg(a[6]); sl_pct = _pctarg(a[7]); move_pct = _pctarg(a[8])
    for nm, v in (("止盈", tp_pct), ("止損", sl_pct), ("移動門檻", move_pct)):
        if v is None:
            await reply(u, f"{E.BOT} {nm} 要帶 %，例如 0.4%"); return
        if v <= 0:
            await reply(u, f"{E.BOT} {nm} 須大於 0"); return
    tf = ACCOUNT_TF
    if tf not in HA_TF:
        await reply(u, f"{E.BOT} 目前週期 {tf} 無效，請先 /timeframe 設定"); return
    for nm, v in (("反向燈號數", pre), ("順向燈號數", post)):
        if not 1 <= v <= DB_KEEP:
            await reply(u, f"{E.BOT} {nm} 須 1~{DB_KEEP}"); return
    if pre + post > DB_KEEP:
        await reply(u, f"{E.BOT} 兩個燈號數相加不可超過 {DB_KEEP}（行情DB 保留上限）"); return
    if not 1 <= interval <= 60:
        await reply(u, f"{E.BOT} 間隔秒須 1~60"); return
    if lev < 1:
        await reply(u, f"{E.BOT} 槓桿須 ≥ 1"); return
    k = skey(sym, dr)
    if k in STRATS and STRATS[k].get("alive"):
        await reply(u, f"{E.BOT} {sym} {E.dir_word(dr)} 已在運行"); return
    try: spec = await get_spec(sym)
    except Exception: await reply(u, f"{E.LOSS} 找不到商品 {sym}"); return
    op = await get_last(spec["iid"])
    size = csize(margin, Decimal(lev), op, spec["ctval"], spec["lot"])
    if size < spec["minsz"]:
        nd = spec["minsz"] * spec["ctval"] * op / Decimal(lev)
        await reply(u, f"{E.BOT} {E.LOSS} 保證金不足：算出 {size} 張 < 最小 {spec['minsz']}\n此槓桿下至少需約 {nd:.4f} USDT"); return
    rows = db_latest(sym, tf, pre + post)
    if len(rows) < pre + post:
        await reply(u, f"{E.BOT} {E.LOSS} 行情DB 資料不足（{sym} {tf} 目前 {len(rows)} 根，需 {pre+post} 根）"); return
    r = judge_entry_db(rows, dr, pre, post)
    prev = "\n".join(entry_lines(r, dr))
    tp0, sl0 = frame_px(op, dr, tp_pct, sl_pct, spec["tick"])
    ps = "long" if dr == "L" else "short"
    exist = await okx_pos(spec["iid"], ps)
    warn = f"\n⚠ OKX 上 {sym} {E.dir_word(dr)} 已有 {exist['pos']} 張持倉" if exist else ""
    if not db_fresh(sym, tf): warn += "\n⚠ 行情DB 落後，啟動後會等資料補齊才判斷"
    PENDING[u.effective_chat.id] = {"act": "run", "t": time.time(), "sym": sym, "dir": dr,
        "tf": tf, "lev": lev, "margin": margin, "pre": pre, "post": post,
        "tp_pct": tp_pct, "sl_pct": sl_pct, "move_pct": move_pct,
        "interval": interval, "spec": spec}
    await reply(u, f"{E.BOT} OKX均K｜{ACCT}\n事件：交易參數預覽\n"
        f"{E.dir_emoji(dr)} {sym} {dr} {tf}\n"
        f"/run {sym} {dr} {lev}x {pct(margin)} {pre} {post} "
        f"{pct(tp_pct)}% {pct(sl_pct)}% {pct(move_pct)}% {interval}\n"
        f"目前價{op}｜預估{size}張\n"
        f"━━━━━━━━━━\n{prev}\n"
        f"━━━━━━━━━━\n"
        f"以現價估：止盈{tp0}／止損{sl0}\n"
        f"每{interval}秒檢查，達{pct(move_pct)}%整體平移{warn}\n"
        f"下一步：60秒內 /confirm\n時間：{hhmmss()}")
    asyncio.create_task(_to(c.application, u.effective_chat.id, PENDING[u.effective_chat.id]["t"]))

ACT_NAME = {"run": "/run 建立策略", "stop": "/stop 停止策略", "stopall": "/stopall 停止全部"}

async def _to(app, chat, stamp):
    await asyncio.sleep(61)
    p = PENDING.get(chat)
    if p and p["t"] == stamp:
        nm = ACT_NAME.get(p.get("act", "run"), "動作")
        del PENDING[chat]
        await notify(app, chat, f"{E.BOT} 逾時已取消：{nm}\n如仍要執行請重新輸入")

async def cmd_confirm(u, c):
    global CHAT_ID; CHAT_ID = u.effective_chat.id
    p = PENDING.get(u.effective_chat.id)
    if not p: await reply(u, f"{E.BOT} 沒有待確認的動作"); return
    if time.time() - p["t"] > 60:
        nm = ACT_NAME.get(p.get("act", "run"), "動作")
        del PENDING[u.effective_chat.id]
        await reply(u, f"{E.BOT} 確認逾時（{nm}），請重新輸入"); return
    act = p.get("act", "run")
    if act == "stop":
        del PENDING[u.effective_chat.id]
        await do_stop(u, p["key"]); return
    if act == "stopall":
        del PENDING[u.effective_chat.id]
        await do_stopall(u); return
    del PENDING[u.effective_chat.id]
    k = skey(p["sym"], p["dir"])
    # 同 key 若還有舊策略/舊 task，先停掉並等它收工，避免兩個 task 同時對同一倉位下單
    old_S = STRATS.get(k)
    if old_S is not None:
        old_S["alive"] = False
    old_T = TASKS.get(k)
    if old_T is not None and not old_T.done():
        old_T.cancel()
        try:
            await asyncio.wait([old_T], timeout=8)
        except Exception as e:
            print("cancel old task fail", e)
        await reply(u, f"{E.BOT} 已先停止 {p['sym']} {E.dir_word(p['dir'])} 的舊策略")
    STRATS[k] = S = {**p, "alive": True, "state": "等訊號", "chat": u.effective_chat.id,
                     "hb": time.time(), "hb_warned": False}
    TASKS[k] = asyncio.create_task(hloop(c.application, u.effective_chat.id, S))
    save_state(True)
    cnt = sum(1 for s in STRATS.values() if s.get("alive"))
    await reply(u, f"{E.BOT} ✅ 已確認，{p['sym']} {E.dir_word(p['dir'])} 啟動\n運行中策略：{cnt} 個")

async def cmd_stop(u, c):
    """只做預覽並登記待確認，實際停止在 /confirm 之後。"""
    global CHAT_ID; CHAT_ID = u.effective_chat.id
    a = c.args
    alive = [k for k, s in STRATS.items() if s.get("alive")]
    if not alive: await reply(u, f"{E.BOT} 目前無運行中策略"); return
    if not a:
        lst = "\n".join(f"・/stop {STRATS[k]['sym']} {STRATS[k]['dir']}" for k in alive)
        await reply(u, f"{E.BOT} 請指定：\n{lst}\n或 /stopall"); return
    sym = a[0].upper()
    tg = [skey(sym, a[1].upper())] if len(a) >= 2 and skey(sym, a[1].upper()) in alive \
         else [k for k in alive if STRATS[k]["sym"] == sym]
    if not tg: await reply(u, f"{E.BOT} 找不到運行中的 {sym}"); return
    if len(tg) > 1: await reply(u, f"{E.BOT} {sym} 有多方向，請指定 /stop {sym} L 或 S"); return
    key = tg[0]
    S = STRATS[key]; d = S["dir"]; iid = S["spec"]["iid"]
    ps = "long" if d == "L" else "short"
    p = await okx_pos(iid, ps)
    L = [f"{E.BOT} OKX均K｜{ACCT}", "事件：停止策略確認", "━" * 10,
         f"{E.dir_emoji(d)} {S['sym']} {E.dir_word(d)} {S['tf']}",
         strat_params(S["sym"], d) or ""]
    if p:
        L += ["━" * 10,
              f"⚠ 目前持倉 {p['pos']} 張",
              f"均價 {p.get('avgPx','?')}｜浮動 {p.get('upl','?')}",
              "停止後不會自動平倉，也不再監控回吐",
              "倉位將完全交由你手動處理"]
    else:
        L += ["━" * 10, "目前空手，停止後不再等訊號"]
    L += ["━" * 10, "下一步：60秒內 /confirm 確認停止", f"時間：{hhmmss()}"]
    PENDING[u.effective_chat.id] = {"act": "stop", "t": time.time(), "key": key}
    await reply(u, "\n".join(L))
    asyncio.create_task(_to(c.application, u.effective_chat.id, PENDING[u.effective_chat.id]["t"]))

async def do_stop(u, key):
    """真正執行停止（由 /confirm 呼叫）。"""
    S = STRATS.get(key)
    if not S or not S.get("alive"):
        await reply(u, f"{E.BOT} 該策略已不在運行中"); return
    d = S["dir"]; iid = S["spec"]["iid"]
    ps = "long" if d == "L" else "short"
    p = await okx_pos(iid, ps)
    S["alive"] = False
    # 立刻移出並寫回存檔，不等迴圈醒來——否則重啟會把停掉的策略撈回來（幽靈策略）
    STRATS.pop(key, None); TASKS.pop(key, None)
    save_state(True)
    if p:
        await reply(u, f"{E.BOT} ✅ 已停止\n{E.dir_emoji(d)} {S['sym']} {E.dir_word(d)}\n"
                       f"⚠ 已進場不自動平倉\n倉位 {p['pos']} 張 均價 {p.get('avgPx','?')} 浮 {p.get('upl','?')}\n"
                       f"止盈止損單仍留在交易所端，但框架停止移動\n請至 OKX 手動處理")
    else:
        await reply(u, f"{E.BOT} ✅ 已停止\n{E.dir_emoji(d)} {S['sym']} {E.dir_word(d)}")

async def cmd_stopall(u, c):
    """只做預覽並登記待確認，實際停止在 /confirm 之後。"""
    global CHAT_ID; CHAT_ID = u.effective_chat.id
    alive = [k for k, s in STRATS.items() if s.get("alive")]
    if not alive: await reply(u, f"{E.BOT} 目前無運行中策略"); return
    held = []; flat = []
    for k in alive:
        S = STRATS[k]; d = S["dir"]
        ps = "long" if d == "L" else "short"
        p = await okx_pos(S["spec"]["iid"], ps)
        line = f"{E.dir_emoji(d)} {S['sym']} {d} {S['tf']}"
        if p: held.append(f"{line}｜持倉 {p['pos']} 張")
        else: flat.append(line)
    L = [f"{E.BOT} OKX均K｜{ACCT}", "事件：停止全部確認", "━" * 10,
         f"將停止 {len(alive)} 個策略"]
    if flat: L += ["━" * 10, f"空手（{len(flat)}）："] + flat
    if held:
        L += ["━" * 10, f"⚠ 持倉中（{len(held)}）："] + held
        L += ["停止後不會自動平倉，也不再監控回吐", "倉位將完全交由你手動處理"]
    L += ["━" * 10, "下一步：60秒內 /confirm 確認停止", f"時間：{hhmmss()}"]
    PENDING[u.effective_chat.id] = {"act": "stopall", "t": time.time()}
    await reply(u, "\n".join(L))
    asyncio.create_task(_to(c.application, u.effective_chat.id, PENDING[u.effective_chat.id]["t"]))

async def do_stopall(u):
    """真正執行停止全部（由 /confirm 呼叫）。"""
    alive = [k for k, s in STRATS.items() if s.get("alive")]
    if not alive:
        await reply(u, f"{E.BOT} 目前無運行中策略"); return
    held = []; done = []
    for k in list(alive):
        S = STRATS[k]; d = S["dir"]; iid = S["spec"]["iid"]
        ps = "long" if d == "L" else "short"
        p = await okx_pos(iid, ps)
        S["alive"] = False
        STRATS.pop(k, None); TASKS.pop(k, None)
        (held if p else done).append(f"{S['sym']} {S['dir']}")
    orphan = await sweep_h()
    save_state(True)
    m = f"{E.BOT} ✅ /stopall 已執行\n━━━━━━━━━━\n"
    if done: m += f"已停止策略（{len(done)}）：\n" + "\n".join("・" + x for x in done) + "\n"
    if orphan: m += f"另清除均K 殘留掛單：{orphan} 筆\n"
    if held: m += f"⚠ 持倉需手動平倉（{len(held)}）：\n" + "\n".join("・" + x for x in held) + "\n"
    await reply(u, m + f"時間：{hhmmss()}")

async def cmd_status(u, c):
    """總覽一則 + 每個幣種各一則（兩個方向）。空手列進場條件，持倉列出場距離。"""
    global CHAT_ID; CHAT_ID = u.effective_chat.id
    posr = await api("GET", "/api/v5/account/positions")
    bal = await api("GET", "/api/v5/account/balance")
    eq = av = "?"
    if bal.get("code") == "0":
        x = next((d for d in bal["data"][0].get("details", []) if d["ccy"] == "USDT"), None)
        if x:
            eq = f"{Decimal(x.get('eq','0')):.4f}"
            av = f"{Decimal(x.get('availEq') or x.get('availBal') or '0'):.4f}"
    pl = [p for p in posr.get("data", []) if float(p.get("pos", "0")) != 0] if posr.get("code") == "0" else []
    # 只取均K 自己策略對應的持倉；帳戶上其他倉位（原K 或手動單）一律不列入
    mark = {}; okxpos = {}
    for p in pl:
        sy = p["instId"].replace("-USDT-SWAP", "USDT")
        dk = "L" if p["posSide"] == "long" else "S"
        mark[(sy, dk)] = p.get("markPx") or p.get("last")
        okxpos[(sy, dk)] = p
    alive = [s for s in STRATS.values() if s.get("alive")]
    syms = sorted({s["sym"] for s in alive})
    held = [s for s in alive if s.get("pos_open")]
    L = [f"{E.BOT} OKX均K｜{ACCT}", "事件：現況總覽", "━" * 10,
         f"USDT權益：{eq}", f"可用餘額：{av}", f"帳戶週期：{ACCOUNT_TF}",
         "━" * 10,
         f"均K 策略：{len(alive)} 個｜幣種 {len(syms)} 個",
         f"均K 持倉：{len(held)} 個"]
    for s in held:
        dk = s["dir"]; key = (s["sym"], dk)
        pinfo = okxpos.get(key)
        szs = (pinfo or {}).get("pos") or s.get("pos_sz") or "?"
        L.append(f"{E.dir_emoji(dk)} {s['sym']} {dk} {szs}張")
        if pinfo is None:
            L.append("　⚠ OKX 查無此倉，可能已手動平倉")
    L += ["━" * 10, f"時間：{hhmmss()} UTC+8"]
    await reply(u, "\n".join(L))
    for i, sym in enumerate(syms, 1):
        M = [f"{E.BOT} {sym}（{i}/{len(syms)}）", "━" * 10]
        for dr in ("L", "S"):
            S = STRATS.get(skey(sym, dr))
            if not S or not S.get("alive"):
                M.append(f"{E.dir_emoji(dr)} {E.dir_word(dr)}：未設定")
                M.append("━" * 10)
                continue
            try:
                M += await strat_detail(S, sym, dr, mark.get((sym, dr)))
            except Exception as e:
                M.append(f"{E.LOSS} 明細產生失敗：{type(e).__name__}: {e}")
            M.append("━" * 10)
        M.append(f"時間：{hhmmss()}")
        await reply(u, "\n".join(M))
        await asyncio.sleep(0.3)

# ---------- /summary ----------
def sum_lines(rs, entered):
    L = []
    m = len(rs)
    win = [r for r in rs if Decimal(str(r.get("net") or "0")) > 0]
    los = [r for r in rs if Decimal(str(r.get("net") or "0")) < 0]
    wr = (len(win) / m * 100) if m else 0
    L.append("進場數：%d | 已出場：%d" % (entered, m))
    L.append("勝率：%.2f%%（勝 %d / 敗 %d）" % (wr, len(win), len(los)))
    if m:
        L.append("平均持倉：%d秒（約 %.1f 根）" % (
            sum(int(r.get("hold_s") or 0) for r in rs) / m,
            sum(float(r.get("bars") or 0) for r in rs) / m))
    else:
        L.append("平均持倉：-")
    if win:
        L.append("　平均獲利：%+.6f" % (sum(Decimal(str(r.get("net") or "0")) for r in win) / len(win)))
    if los:
        L.append("　平均虧損：%+.6f" % (sum(Decimal(str(r.get("net") or "0")) for r in los) / len(los)))
    L.append("━" * 10)
    tg = sum((Decimal(str(r.get("gross") or "0")) for r in rs), Decimal(0))
    tf = sum((Decimal(str(r.get("fee") or "0")) for r in rs), Decimal(0))
    tn = sum((Decimal(str(r.get("net") or "0")) for r in rs), Decimal(0))
    nv = sum((Decimal(str(r.get("nv") or "0")) for r in rs), Decimal(0))
    gp = (tg / nv * 100) if nv else Decimal(0)
    fp = (tf / nv * 100) if nv else Decimal(0)
    npc = (tn / nv * 100) if nv else Decimal(0)
    L.append("毛損益：%+.6f (%+.3f%%)" % (tg, gp))
    L.append("手續費：%+.6f (%+.3f%%)" % (tf, fp))
    L.append("淨損益：%+.6f (%+.3f%%) %s" % (tn, npc, E.pnl_emoji(tn)))
    return L

async def cmd_summary(u, c):
    t = today8(); recs = load_trades(t)
    ts = {k: v for k, v in STATS.items() if str(v.get("date")) == str(t)}
    L = [f"{E.BOT} OKX均K｜{ACCT}", f"📊📊📊 Summary {t}"]
    for dr in ("L", "S"):
        rows = [r for r in recs if r["dir"] == dr]
        en = sum(v["entered"] for k, v in ts.items() if k.endswith("_" + dr))
        L.append("━" * 10)
        L.append(f"{E.dir_emoji(dr)} {E.dir_word(dr)}")
        L += sum_lines(rows, en)
    L += ["━" * 10, f"時間：{hhmmss()}"]
    await reply(u, "\n".join(L))
    for sy in sorted({r["sym"] for r in recs}):
        D = [f"\U0001f49a\U0001f499\U0001fa75\U0001f49c {sy} {t}"]
        for dr in ("L", "S"):
            rows = [r for r in recs if r["sym"] == sy and r["dir"] == dr]
            st_ = ts.get(skey(sy, dr)) or {"placed": 0, "entered": 0}
            D.append("━" * 10)
            D.append(f"策略：{E.dir_emoji(dr)} {strat_params(sy, dr) or (dr + '（已停止）')}")
            D += sum_lines(rows, st_["entered"])
        D += ["━" * 10, f"時間：{hhmmss()}"]
        await reply(u, "\n".join(D))

def _fmt_atr(v, tick):
    """ATR 依商品 tick 決定小數位，避免大幣印一堆 0 或小幣被截斷。"""
    if v is None: return "-"
    exp = -tick.as_tuple().exponent
    q = Decimal(1).scaleb(-max(2, min(8, exp + 1)))
    return str(v.quantize(q))

def build_ha_xlsx(sym, tf, ha, atrs, tick, path, kl=None):
    """明細＋統計兩張表。
    欄位：幣種 / 週期 / 日期 / 時間 / 原K OHLC / 原K燈 / 原K漲跌幅 /
          均K OHLC / 均K燈 / 均K漲跌幅 / ATR14 / ATR14 ratio"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active; ws.title = "明細"
    heads = ["幣種", "週期", "日期", "時間",
             "原K開", "原K高", "原K低", "原K收", "原K燈", "原K漲跌幅",
             "均K開", "均K高", "均K低", "均K收", "均K燈", "均K漲跌幅",
             "ATR14", "ATR14 ratio"]
    ws.append(heads)
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="404040")
    for i in range(1, len(heads) + 1):
        cc = ws.cell(row=1, column=i); cc.font = hf; cc.fill = hfill
        cc.alignment = Alignment(horizontal="center")
    exp = -tick.as_tuple().exponent
    ndp = max(2, min(8, exp + 1))
    afmt = "0." + "0" * ndp
    pfmt = "0." + "0" * ndp
    for i, x in enumerate(ha):
        dt = datetime.fromtimestamp(int(x["ts"]) / 1000, TZ8)          # K 線開盤時間
        av, rv = atrs[i]
        up = x["color"] == "G"
        k = kl[i] if (kl and i < len(kl)) else None
        ro = float(k["o"]) if k else None
        rh = float(k["h"]) if k else None
        rl = float(k["l"]) if k else None
        rc = float(k["c"]) if k else None
        rbody = ((rc - ro) / ro * 100) if (ro and rc is not None) else None
        hbody = float((x["hc"] - x["ho"]) / x["ho"] * 100) if x["ho"] else None
        ws.append([sym, tf, dt.strftime("%Y/%m/%d"), dt.strftime("%H:%M"),
                   ro, rh, rl, rc,
                   ("\U0001F7E9" if rc >= ro else "\U0001F7E5") if (ro is not None and rc is not None) else None,
                   rbody,
                   float(x["ho"]), float(x["hh"]), float(x["hl"]), float(x["hc"]),
                   "\U0001F7E9" if up else "\U0001F7E5",
                   hbody,
                   float(av) if av is not None else None,
                   float(rv) if rv is not None else None])
        r = ws.max_row
        for col in (2, 9, 15): ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")
        for col in (5, 6, 7, 8, 11, 12, 13, 14): ws.cell(row=r, column=col).number_format = pfmt
        for col in (10, 16): ws.cell(row=r, column=col).number_format = '0.0000"%"'
        ws.cell(row=r, column=17).number_format = afmt
        ws.cell(row=r, column=18).number_format = '0.0000"%"' 
    ws.freeze_panes = "A2"
    for i, w in enumerate([12, 8, 12, 8,
                           13, 13, 13, 13, 8, 12,
                           13, 13, 13, 13, 8, 12,
                           14, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    va = [float(a) for a, _ in atrs if a is not None]
    vr = [float(r) for _, r in atrs if r is not None]
    ng = sum(1 for x in ha if x["color"] == "G")
    def stat(v):
        if not v: return ("-", "-", "-", "-")
        sv = sorted(v); m = len(sv)
        med = sv[m // 2] if m % 2 else (sv[m // 2 - 1] + sv[m // 2]) / 2
        return (sum(v) / m, med, max(v), min(v))
    aA, aM, aX, aN = stat(va)
    rA, rM, rX, rN = stat(vr)
    t0 = datetime.fromtimestamp(int(ha[0]["ts"]) / 1000, TZ8).strftime("%Y/%m/%d %H:%M")
    t1 = datetime.fromtimestamp(int(ha[-1]["ts"]) / 1000, TZ8).strftime("%Y/%m/%d %H:%M")
    w2 = wb.create_sheet("統計")
    rows = [["項目", "數值"],
            ["幣種", sym], ["週期", tf], ["根數", len(ha)],
            ["期間起", t0], ["期間迄", t1],
            ["\U0001F7E9 根數", ng], ["\U0001F7E5 根數", len(ha) - ng],
            ["\U0001F7E9 佔比", round(ng / len(ha) * 100, 2) if ha else 0],
            ["ATR14 平均", aA], ["ATR14 中位", aM], ["ATR14 最大", aX], ["ATR14 最小", aN],
            ["ATR14 ratio 平均", rA], ["ATR14 ratio 中位", rM],
            ["ATR14 ratio 最大", rX], ["ATR14 ratio 最小", rN],
            ["燈號來源", "均K燈=Heikin-Ashi｜原K燈=原始收≥開"],
            ["ATR 來源", "原始 K 線（Wilder 平滑，14 週期）"],
            ["ATR14 ratio", "ATR14 ÷ 該根原始收盤價 × 100%"],
            ["產生時間", now8().strftime("%Y/%m/%d %H:%M:%S")]]
    for r in rows: w2.append(r)
    for rr in range(2, w2.max_row + 1):
        lab = w2.cell(row=rr, column=1).value or ""
        cell = w2.cell(row=rr, column=2)
        if "ratio" in str(lab) and isinstance(cell.value, (int, float)):
            cell.number_format = '0.0000"%"'
        elif "佔比" in str(lab) and isinstance(cell.value, (int, float)):
            cell.number_format = '0.00"%"'
        elif str(lab).startswith("ATR14 ") and isinstance(cell.value, (int, float)):
            cell.number_format = afmt
    for i in range(1, 3):
        cc = w2.cell(row=1, column=i); cc.font = hf; cc.fill = hfill
    w2.column_dimensions["A"].width = 24
    w2.column_dimensions["B"].width = 24
    wb.save(path)

def send_ha_mail(path, name, sym, tf, m, aA, rA):
    """寄出均K 燈號+ATR 報表。回傳 (ok, 訊息)。"""
    import smtplib
    from email.message import EmailMessage
    env = {}
    try:
        for line in open("/srv/1111bot/.env"):
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip().strip('"').strip("'")
    except Exception as e:
        return False, "讀 .env 失敗：%s" % e
    user = env.get("GMAIL_USER")
    pwd = (env.get("GMAIL_APP_PASSWORD") or "").replace(" ", "")
    to = env.get("REPORT_TO") or user
    if not user or not pwd:
        return False, "未設定 GMAIL_USER / GMAIL_APP_PASSWORD"
    msg = EmailMessage()
    msg["Subject"] = "OKX %s 均K燈號+ATR %s %s（%d 根）" % (ACCT, sym, tf, m)
    msg["From"] = user; msg["To"] = to
    msg.set_content(
        "商品 %s\n週期 %s\n根數 %d\n"
        "ATR14 平均 %s\nATR14 ratio 平均 %s%%\n"
        "欄位：幣種 / 日期 / 時間 / 燈號 / ATR14 / ATR14 ratio\n"
        "燈號取 Heikin-Ashi，ATR 取原始 K 線\n" % (sym, tf, m, aA, rA))
    msg.add_attachment(open(path, "rb").read(),
                       maintype="application",
                       subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       filename=name)
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as sv:
        sv.login(user, pwd); sv.send_message(msg)
    return True, to

async def cmd_ha(u, c):
    """均K 燈號+ATR 報表（Excel 寄信）。用法：/ha ETHUSDT [根數 3~2000]"""
    if not c.args:
        await reply(u, f"{E.BOT} 用法：/ha ETHUSDT 900\n"
                       f"根數 3~{HA_MAX}（預設 300）\n"
                       f"產生 Excel（明細＋統計）寄到信箱\n"
                       f"欄位：幣種/日期/時間/燈號/ATR14/ATR14 ratio\n"
                       f"※ 燈號取 HA，ATR 取原始 K 線\n"
                       f"目前週期：{ACCOUNT_TF}")
        return
    sym = c.args[0].upper()
    n = 300
    if len(c.args) >= 2:
        try: n = max(3, min(HA_MAX, int(c.args[1])))
        except Exception: pass
    try:
        spec = await get_spec(sym)
    except Exception:
        await reply(u, f"{E.LOSS} 找不到商品 {sym}"); return
    await reply(u, f"{E.BOT} 產生 {sym} {ACCOUNT_TF} 燈號+ATR 報表中（{n} 根），請稍候…")
    try:
        kl = await klines_paged_for_tf(spec["iid"], ACCOUNT_TF, n + 20)   # +20 給 ATR14 暖身
    except Exception as e:
        await reply(u, f"{E.LOSS} K 線取得失敗：{type(e).__name__}: {e}"); return
    if not kl:
        await reply(u, f"{E.BOT} {sym} K 線取得失敗"); return
    ha_all = calc_ha(kl)
    atr_all = calc_atr(kl, 14)
    st = max(0, len(ha_all) - n)
    ha = ha_all[st:]; atrs = atr_all[st:]
    m = len(ha)
    if m < 3:
        await reply(u, f"{E.BOT} {sym} {ACCOUNT_TF} 可用 K 線僅 {m} 根，不足以產表"); return
    day = now8().strftime("%Y%m%d")
    name = f"OKX_{ACCT}_均K_{sym}_{ACCOUNT_TF}_{m}根_{day}.xlsx"
    path = f"/srv/1111bot/data/{name}"
    try:
        build_ha_xlsx(sym, ACCOUNT_TF, ha, atrs, spec["tick"], path, kl[st:])
    except Exception as e:
        await reply(u, f"{E.LOSS} 產生 Excel 失敗：{type(e).__name__}: {e}"); return
    va = [a for a, _ in atrs if a is not None]
    vr = [r for _, r in atrs if r is not None]
    aA = (sum(va) / len(va)) if va else Decimal(0)
    rA = (sum(vr) / len(vr)) if vr else Decimal(0)
    aAs = _fmt_atr(aA, spec["tick"]); rAs = f"{rA:.4f}"
    try:
        ok, info = send_ha_mail(path, name, sym, ACCOUNT_TF, m, aAs, rAs)
    except Exception as e:
        await reply(u, f"{E.LOSS} 寄送失敗：{type(e).__name__}: {e}\n檔案已存於 VPS：{name}"); return
    if not ok:
        await reply(u, f"{E.LOSS} 未寄送：{info}\n檔案已存於 VPS：{name}"); return
    t0 = datetime.fromtimestamp(int(ha[0]["ts"]) / 1000, TZ8).strftime("%m/%d %H:%M")
    t1 = datetime.fromtimestamp(int(ha[-1]["ts"]) / 1000, TZ8).strftime("%m/%d %H:%M")
    ng = sum(1 for x in ha if x["color"] == "G")
    await reply(u, f"{E.BOT} ✅ 均K 報表已寄出\n"
                   f"{sym} {ACCOUNT_TF}｜{m} 根\n"
                   f"期間：{t0} ~ {t1}\n"
                   f"燈號：🟩{ng} / 🟥{m - ng}\n"
                   f"ATR14 平均 {aAs}\n"
                   f"ATR14 ratio 平均 {rAs}%\n"
                   f"時間：{hhmmss()}")

async def cmd_db(u, c):
    """行情DB 狀態。用法：/db 或 /db ETHUSDT 或 /db ETHUSDT 5m"""
    if DB is None:
        await reply(u, f"{E.BOT} 行情DB 尚未就緒"); return
    sym = c.args[0].upper() if c.args else None
    tf = c.args[1] if len(c.args) >= 2 else None
    if tf and tf not in HA_TF:
        await reply(u, f"{E.BOT} 週期須為：" + " / ".join(TF_LIST)); return
    ms = db_meta(sym, tf)
    if not ms:
        await reply(u, f"{E.BOT} 查無資料（收集器可能還在初始化）"); return
    if sym and tf:
        rows = db_latest(sym, tf, 5)
        L = [f"{E.BOT} OKX均K｜{ACCT}", f"事件：行情DB {sym} {tf}", "━" * 10,
             f"筆數：{ms[0]['bars']}｜新鮮：{'✅' if db_fresh(sym, tf) else '⚠ 落後'}",
             "━" * 10, "最近 5 根："]
        for r in rows:
            lg = "\U0001F7E9" if r["color"] == "G" else "\U0001F7E5"
            ar = f"{r['atr14_ratio']:.4f}%" if r["atr14_ratio"] is not None else "-"
            L.append(f"{r['dt']} {lg} 連{r['streak']}")
            L.append(f"　實體{r['body_pct']:+.4f}% 振幅{r['range_pct']:.4f}% ATRr {ar}")
            L.append(f"　連續段累計 實體{r['streak_body']:+.4f}% 振幅{r['streak_range']:.4f}%")
        L += ["━" * 10, f"時間：{hhmmss()}"]
        await reply(u, "\n".join(L)); return
    tot = sum(m["bars"] or 0 for m in ms)
    stale = [m for m in ms if not db_fresh(m["sym"], m["tf"])]
    errs = [m for m in ms if m.get("err")]
    L = [f"{E.BOT} OKX均K｜{ACCT}", "事件：行情DB 狀態", "━" * 10,
         f"組合數：{len(ms)}（{len(db_symbols())} 幣種 x {len(TF_LIST)} 週期）",
         f"總筆數：{tot}",
         f"新鮮：{len(ms) - len(stale)}｜落後：{len(stale)}",
         f"錯誤：{len(errs)}"]
    if sym:
        L += ["━" * 10, f"{sym} 各週期："]
        for m in ms:
            fr = "✅" if db_fresh(m["sym"], m["tf"]) else "⚠"
            last = datetime.fromtimestamp((m["last_ts"] or 0) / 1000, TZ8).strftime("%m/%d %H:%M") if m["last_ts"] else "-"
            L.append(f"{fr} {m['tf']:>6}｜{m['bars']} 根｜最新 {last}")
    else:
        if stale:
            L += ["━" * 10, "落後的組合（前 10）："]
            for m in stale[:10]:
                L.append(f"⚠ {m['sym']} {m['tf']}")
        if errs:
            L += ["━" * 10, "錯誤（前 5）："]
            for m in errs[:5]:
                L.append(f"{E.LOSS} {m['sym']} {m['tf']}：{str(m['err'])[:40]}")
        L += ["━" * 10, "細節：/db ETHUSDT 或 /db ETHUSDT 5m"]
    L += ["━" * 10, f"時間：{hhmmss()}"]
    await reply(u, "\n".join(L))

async def cmd_coins(u, c):
    on = [s["symbol"] for s in SYMS if s["enabled"]]
    L = [f"{E.BOT} OKX均K｜{ACCT}", "事件：幣種清單（即時）", "━━━━━━━━━━"]
    for sym in on:
        try:
            sp = await get_spec(sym); last = await get_last(sp["iid"])
            mm = sp["minsz"] * sp["ctval"] * last
            L.append(f"{sym}｜最小{sp['minsz']}張｜{mm:.4f}U")
        except Exception:
            L.append(f"{sym}｜查詢失敗")
    L += ["━━━━━━━━━━", f"時間：{hhmmss()}"]
    await reply(u, "\n".join(L))

async def cmd_timeframe(u, c):
    global ACCOUNT_TF
    opts = " / ".join(TF_LIST)
    if not c.args:
        L = [f"{E.BOT} 目前週期：{ACCOUNT_TF}", "━" * 10, "可選週期："]
        for t in TF_LIST:
            sec, bar, mul = HA_TF[t]
            src = f"OKX {bar}" if mul == 1 else f"{mul}x OKX {bar} 合成"
            L.append(f"・{t}（{src}）")
        L += ["━" * 10, "變更：/timeframe 60m",
              "※ 僅影響之後新建立的策略",
              "※ 720m/1440m 以 UTC+8 為日界，與 OKX App 一致"]
        await reply(u, "\n".join(L)); return
    tf = c.args[0]
    if tf not in HA_TF:
        await reply(u, f"{E.BOT} 週期須為：{opts}"); return
    ACCOUNT_TF = tf; save_state()
    await reply(u, f"{E.BOT} ✅ 帳戶週期已設為 {tf}\n（僅影響之後新建立的策略）")

async def cmd_menu(u, c):
    await reply(u, f"{E.BOT} OKX均K｜{ACCT}\n使用說明\n━━━━━━━━━━\n"
        "/status 策略現況\n"
        "/summary 當日戰報\n"
        "/coins 幣種\n"
        "/ha 商品 根數  燈號+ATR 報表 Excel 寄信（3~2000根）\n"
        "/replay 逐筆回溯報表 Excel 寄信（/replay 20260904）\n"
        "━━━━━━━━━━\n"
        "/run 商品 方向 槓桿x 資金 反向燈號數 順向燈號數 止盈% 止損% 移動門檻% 間隔秒\n"
        "例：/run BTCUSDT L 1x 10 1 5 0.4% 0.2% 0.05% 5\n"
        f"　週期依 /timeframe（目前 {ACCOUNT_TF}）\n"
        "/confirm 確認執行（run/stop/stopall 共用）\n"
        "/stop 商品 方向（需 /confirm）\n"
        "/stopall 停全部（需 /confirm）\n"
        "━━━━━━━━━━\n"
        "/db 行情DB狀態（/db ETHUSDT 5m 看細節）\n"
        f"/timeframe 查看/設定週期 共{len(TF_LIST)}種\n"
        "━━━━━━━━━━\n"
        "進場：前段全反向色 + 後段全順勢色（taker）\n"
        "出場：止盈止損掛在交易所端（OCO）\n"
        "　　　每 N 秒查價，達門檻就把框架整體平移\n"
        "　　　程式掛掉止盈止損仍然有效\n"
        "⚠ 無 TP/SL/TE，僅靠回吐出場\n"
        "✅ 判斷全讀本機行情DB，下單不等 API\n"
        "✅ 重啟接管持倉\n"
        "✅ 獨立進程，不影響其他策略")

async def cmd_unknown(u, c):
    await reply(u, f"{E.BOT} 指令無法辨識：{u.message.text}\n請用 /menu")

# ---------- /replay 逐筆回溯報表 ----------
def send_xlsx_mail(path, name, subject, body):
    """通用：寄出一個 xlsx 附件。回傳 (ok, 訊息)。"""
    import smtplib
    from email.message import EmailMessage
    env = {}
    try:
        for line in open("/srv/1111bot/.env"):
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                kk, vv = line.split("=", 1)
                env[kk.strip()] = vv.strip().strip('"').strip("'")
    except Exception as e:
        return False, "讀 .env 失敗：%s" % e
    user = env.get("GMAIL_USER")
    pwd = (env.get("GMAIL_APP_PASSWORD") or "").replace(" ", "")
    to = env.get("REPORT_TO") or user
    if not user or not pwd:
        return False, "未設定 GMAIL_USER / GMAIL_APP_PASSWORD"
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = user; msg["To"] = to
    msg.set_content(body)
    msg.add_attachment(open(path, "rb").read(),
                       maintype="application",
                       subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       filename=name)
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as sv:
        sv.login(user, pwd); sv.send_message(msg)
    return True, to

def _epoch(day, hms):
    """day 'YYYY-MM-DD' + hms 'HH:MM:SS' -> epoch 秒（UTC+8）"""
    try:
        return datetime.strptime(day + " " + str(hms)[:8],
                                 "%Y-%m-%d %H:%M:%S").replace(tzinfo=TZ8).timestamp()
    except Exception:
        return None

async def build_replay_xlsx(day, trades, path):
    """每筆成交一個分頁，逐根列出燈號 / ATR / 損益 / 出場條件檢查。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HF = Font(bold=True, color="FFFFFF"); HFILL = PatternFill("solid", fgColor="404040")
    CEN = Alignment(horizontal="center")

    # 先把每筆的進出場 epoch 算出來（跨日則進場歸前一天）
    for t in trades:
        et = _epoch(day, t.get("in_ts")); xt = _epoch(day, t.get("ts"))
        if et is not None and xt is not None and et > xt:
            et -= 86400
        t["_ent"] = et; t["_exi"] = xt

    # 依 (幣種,週期) 分組，一組只抓一次 K 線
    groups = {}
    for t in trades:
        groups.setdefault((t.get("sym"), t.get("tf") or ACCOUNT_TF), []).append(t)
    series = {}
    for (sym, tf), ts_ in groups.items():
        if tf not in HA_TF: continue
        try:
            spec = await get_spec(sym)
        except Exception:
            continue
        sec = tf_sec(tf)
        ents = [t["_ent"] for t in ts_ if t.get("_ent")]
        if not ents: continue
        need = int((time.time() - min(ents)) / sec) + 60
        kl = await klines_paged_for_tf(spec["iid"], tf, max(60, min(2000, need)))
        if not kl: continue
        series[(sym, tf)] = (kl, calc_ha(kl), calc_atr(kl, 14))

    wb = Workbook(); ws = wb.active; ws.title = "總表"
    SH = ["#", "幣種", "週期", "方向", "進場時間", "進場價", "出場時間", "出場價",
          "持倉秒數", "持倉根數", "累計損益%", "淨損益",
          "進場參考價", "進場延遲s", "進場偏離%", "進場輪詢次", "進場輪詢s",
          "出場參考價", "出場延遲s", "出場偏離%", "出場輪詢次", "出場輪詢s",
          "反向燈號數", "順向燈號數", "止盈%", "止損%", "移動門檻%", "間隔秒",
          "框架移動次數", "分頁"]
    ws.append(SH)
    for i in range(1, len(SH) + 1):
        cc = ws.cell(1, i); cc.font = HF; cc.fill = HFILL; cc.alignment = CEN

    def fl(v, d=0.0):
        try: return float(v)
        except Exception: return d

    made = 0
    for idx, t in enumerate(trades, 1):
        sym = t.get("sym"); tf = t.get("tf") or ACCOUNT_TF; dr = t.get("dir")
        key = (sym, tf)
        ep = fl(t.get("in_px")); ent = t.get("_ent"); exi = t.get("_exi")
        hm_in = str(t.get("in_ts") or "")[:5]
        tab = f"{idx}_{str(sym)[:6]}_{dr}_{hm_in.replace(':', '')}"[:31]
        nv = fl(t.get("nv")); net = fl(t.get("net"))
        def nn(v):
            try: return float(v) if v is not None else None
            except Exception: return None
        ws.append([idx, sym, tf, dr, t.get("in_ts"), ep, t.get("ts"), fl(t.get("out_px")),
                   int(fl(t.get("hold_s"))), fl(t.get("bars")),
                   (net / nv * 100) if nv else 0.0, net,
                   nn(t.get("in_ref")), nn(t.get("in_lag")), nn(t.get("in_slip")),
                   t.get("in_poll_n"), nn(t.get("in_poll_s")),
                   nn(t.get("out_ref")), nn(t.get("out_lag")), nn(t.get("out_slip")),
                   t.get("out_poll_n"), nn(t.get("out_poll_s")),
                   int(fl(t.get("pre"), 0)), int(fl(t.get("post"), 0)),
                   fl(t.get("tp_pct")), fl(t.get("sl_pct")), fl(t.get("move_pct")),
                   t.get("interval"), t.get("move_n"), tab])
        r = ws.max_row
        for col in (11, 15, 20, 25, 26, 27): ws.cell(r, col).number_format = '0.0000"%"'
        ws.cell(r, 12).number_format = "0.000000"
        for col in (14, 17, 19, 22): ws.cell(r, col).number_format = "0.000"
        for col in (6, 8, 13, 18): ws.cell(r, col).number_format = "0.######"
        if key not in series or ent is None or exi is None: continue

        kl, ha, atrs = series[key]
        sec = tf_sec(tf)
        pre = int(fl(t.get("pre"), 1)); post = int(fl(t.get("post"), 1))

        ent_bar = int(ent // sec) * sec
        exi_bar = int(exi // sec) * sec
        lo = ent_bar - (pre + post + 3) * sec
        hi = exi_bar + 2 * sec

        w2 = wb.create_sheet(tab)
        H = ["標記", "開盤時間",
             "均K開", "原K開", "均K收", "原K收",
             "均K燈", "原K燈", "均K漲跌幅", "原K漲跌幅", "ATR14r",
             "本根損益", "累計損益",
             "進場價", "出場價", "參考收盤", "偏離%", "延遲s"]
        w2.append(H)
        for i in range(1, len(H) + 1):
            cc = w2.cell(1, i); cc.font = HF; cc.fill = HFILL; cc.alignment = CEN
        prev_c = None
        for i, k in enumerate(kl):
            bts = int(k["ts"]) // 1000
            if bts < lo or bts > hi: continue
            x = ha[i]; a, rr = atrs[i]
            c = float(k["c"])
            body = float((x["hc"] - x["ho"]) / x["ho"] * 100) if x["ho"] else 0.0
            if bts < ent_bar:
                tag = "訊號"; one = pnl = None
            else:
                tag = ("進場" if bts == ent_bar else
                       "出場" if bts == exi_bar else
                       "出場後" if bts > exi_bar else "持倉")
                pnl = (c - ep) / ep * 100 if dr == "L" else (ep - c) / ep * 100
                base = prev_c if prev_c is not None else ep
                one = (c - base) / base * 100 if dr == "L" else (base - c) / base * 100
            if bts >= ent_bar: prev_c = c
            # 進出場那兩根補上實際成交價與滑價
            ref_c = in_p = out_p = slip = lagv = None
            if bts == ent_bar:
                ref_c = fl(t.get("in_ref")) or None
                in_p = ep
                slip = fl(t.get("in_slip")) if t.get("in_slip") is not None else None
                lagv = fl(t.get("in_lag")) if t.get("in_lag") is not None else None
            elif bts == exi_bar:
                ref_c = fl(t.get("out_ref")) or None
                out_p = fl(t.get("out_px")) or None
                slip = fl(t.get("out_slip")) if t.get("out_slip") is not None else None
                lagv = fl(t.get("out_lag")) if t.get("out_lag") is not None else None
            # 原K 的燈號與漲跌幅（以原始開收計，與交易所K棒一致）
            ro = float(k["o"]); rc = c
            rbody = (rc - ro) / ro * 100 if ro else 0.0
            w2.append([tag,
                       datetime.fromtimestamp(bts, TZ8).strftime("%m/%d %H:%M"),
                       float(x["ho"]), ro, float(x["hc"]), rc,
                       "\U0001F7E9" if x["color"] == "G" else "\U0001F7E5",
                       "\U0001F7E9" if rc >= ro else "\U0001F7E5",
                       body, rbody, float(rr) if rr is not None else None,
                       one, pnl,
                       in_p, out_p, ref_c, slip, lagv])
            rr2 = w2.max_row
            for col in (9, 10, 11, 12, 13, 17): w2.cell(rr2, col).number_format = '0.0000"%"'
            for col in (3, 4, 5, 6, 14, 15, 16): w2.cell(rr2, col).number_format = "0.######"
            w2.cell(rr2, 18).number_format = "0.000"
            for col in (1, 7, 8): w2.cell(rr2, col).alignment = CEN
        w2.freeze_panes = "A2"
        for i, wdt in enumerate([9, 13,
                                 13, 13, 13, 13,
                                 7, 7, 12, 12, 10,
                                 12, 12,
                                 12, 12, 12, 10, 9], 1):
            w2.column_dimensions[get_column_letter(i)].width = wdt
        made += 1

    # ── 執行品質統計 ──
    def col(name):
        v = [float(t[name]) for t in trades if t.get(name) is not None]
        return v
    w3 = wb.create_sheet("執行品質")
    w3.append(["項目", "筆數", "平均", "中位", "最大", "最小"])
    for i in range(1, 7):
        cc = w3.cell(1, i); cc.font = HF; cc.fill = HFILL; cc.alignment = CEN
    for lab, key, fmtn in [("進場延遲(s)", "in_lag", "0.000"),
                           ("出場延遲(s)", "out_lag", "0.000"),
                           ("進場偏離(%)", "in_slip", '0.0000"%"'),
                           ("出場偏離(%)", "out_slip", '0.0000"%"'),
                           ("進場輪詢(次)", "in_poll_n", "0"),
                           ("出場輪詢(次)", "out_poll_n", "0"),
                           ("進場輪詢(s)", "in_poll_s", "0.000"),
                           ("出場輪詢(s)", "out_poll_s", "0.000")]:
        v = col(key)
        if not v:
            w3.append([lab, 0, None, None, None, None]); continue
        sv = sorted(v); m = len(sv)
        med = sv[m // 2] if m % 2 else (sv[m // 2 - 1] + sv[m // 2]) / 2
        w3.append([lab, m, sum(v) / m, med, max(v), min(v)])
        for c2 in range(3, 7):
            w3.cell(w3.max_row, c2).number_format = fmtn
    w3.append([])
    w3.append(["說明"])
    w3.cell(w3.max_row, 1).font = Font(bold=True)
    for line in ["延遲＝K線收線到送出委託的秒數",
                 "偏離＝成交價相對該根收盤價的差距，正值代表對你有利",
                 "輪詢＝等 OKX 標記該根已收線所打的次數與耗時"]:
        w3.append([line])
    for c3, wd in (("A", 16), ("B", 9), ("C", 12), ("D", 12), ("E", 12), ("F", 12)):
        w3.column_dimensions[c3].width = wd

    ws.freeze_panes = "A2"
    for i, wdt in enumerate([5, 11, 7, 6, 11, 12, 11, 12, 10, 10, 12, 12,
                             12, 11, 11, 11, 11,
                             12, 11, 11, 11, 11,
                             12, 12, 10, 10, 12, 10, 12, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    wb.save(path)
    return made

async def cmd_replay(u, c):
    """逐筆回溯報表（Excel 寄信）。用法：/replay 或 /replay 20260904"""
    global CHAT_ID; CHAT_ID = u.effective_chat.id
    day = today8()
    if c.args:
        a0 = str(c.args[0]).strip()
        if a0 in ("yesterday", "昨天"):
            day = (now8() - timedelta(days=1)).strftime("%Y-%m-%d")
        elif len(a0) == 8 and a0.isdigit():
            day = f"{a0[:4]}-{a0[4:6]}-{a0[6:]}"
        elif len(a0) == 10 and a0.count("-") == 2:
            day = a0
        else:
            await reply(u, f"{E.BOT} 日期格式：/replay 20260904 或 /replay 昨天"); return
    trades = load_trades(day)
    if not trades:
        await reply(u, f"{E.BOT} {day} 無成交紀錄"); return
    await reply(u, f"{E.BOT} 產生 {day} 逐筆回溯報表中（{len(trades)} 筆），請稍候…")
    name = f"OKX_{ACCT}_均K回溯_{day.replace('-', '')}.xlsx"
    path = f"/srv/1111bot/data/{name}"
    try:
        made = await build_replay_xlsx(day, trades, path)
    except Exception as e:
        await reply(u, f"{E.LOSS} 產生失敗：{type(e).__name__}: {e}"); return
    tn = sum((float(t.get("net") or 0) for t in trades), 0.0)
    body = ("帳號 %s\n策略 均K\n日期 %s\n成交筆數 %d\n明細分頁 %d\n淨損益 %+.6f USDT\n\n"
            "每筆一個分頁，逐根列出燈號 / ATR14r / 本根損益 / 累計損益，\n"
            "並標示三個出場條件是否成立。\n" % (ACCT, day, len(trades), made, tn))
    try:
        ok, info = send_xlsx_mail(path, name, f"OKX {ACCT} 均K逐筆回溯 {day}（{len(trades)} 筆）", body)
    except Exception as e:
        await reply(u, f"{E.LOSS} 寄送失敗：{type(e).__name__}: {e}\n檔案已存於 VPS：{name}"); return
    if not ok:
        await reply(u, f"{E.LOSS} 未寄送：{info}\n檔案已存於 VPS：{name}"); return
    await reply(u, f"{E.BOT} ✅ 回溯報表已寄出\n{day}｜{len(trades)} 筆｜明細分頁 {made}\n"
                   f"淨損益 {tn:+.6f} USDT\n時間：{hhmmss()}")

# ---------- 每日 Email 日報（00:05 寄前一日） ----------
def build_daily_xlsx(day, trades, path):
    """均K 日報：明細 + 統計。無底色。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    HF = Font(bold=True, color="FFFFFF"); HFILL = PatternFill("solid", fgColor="404040")
    CEN = Alignment(horizontal="center")
    ws = wb.active; ws.title = "明細"
    H = ["日期", "幣種", "週期", "方向", "進場時間", "出場時間", "持倉秒數", "持倉根數",
         "進場價", "出場價", "最高利潤", "淨損益%", "毛損益", "手續費", "淨損益"]
    ws.append(H)
    for i in range(1, len(H) + 1):
        cc = ws.cell(1, i); cc.font = HF; cc.fill = HFILL; cc.alignment = CEN
    def fl(v, d=0.0):
        try: return float(v)
        except Exception: return d
    def it(v, d=0):
        try: return int(v)
        except Exception: return d
    for t in trades:
        net = fl(t.get("net")); nv = fl(t.get("nv"))
        ws.append([day, t.get("sym"), t.get("tf"), t.get("dir"),
                   t.get("in_ts"), t.get("ts"), it(t.get("hold_s")), fl(t.get("bars")),
                   fl(t.get("in_px")), fl(t.get("out_px")),
                   fl(t.get("peak_pct")), (net / nv * 100) if nv else 0.0,
                   fl(t.get("gross")), fl(t.get("fee")), net])
        r = ws.max_row
        ws.cell(r, 4).alignment = CEN
        for col in (11, 12): ws.cell(r, col).number_format = '0.0000"%"'
        for col in (13, 14, 15): ws.cell(r, col).number_format = "0.000000"
        for col in (9, 10): ws.cell(r, col).number_format = "0.######"
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(H))}{ws.max_row}"
    for i, w in enumerate([11, 10, 7, 6, 10, 10, 10, 10, 12, 12, 11, 11,
                           12, 12, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    w2 = wb.create_sheet("統計")
    w2.append(["項目", "數值"])
    for i in (1, 2):
        cc = w2.cell(1, i); cc.font = HF; cc.fill = HFILL
    nets = [fl(t.get("net")) for t in trades]
    n = len(nets) or 1
    win = [x for x in nets if x > 0]; los = [x for x in nets if x < 0]
    tg_ = sum(fl(t.get("gross")) for t in trades)
    tf_ = sum(fl(t.get("fee")) for t in trades)
    tn_ = sum(nets); tnv = sum(fl(t.get("nv")) for t in trades)
    rows = [["日期", day], ["帳號", ACCT], ["策略", "均K（Heikin-Ashi）"],
            ["交易筆數", len(trades)], ["獲利筆數", len(win)], ["虧損筆數", len(los)],
            ["勝率", len(win) / n * 100],
            ["平均持倉秒數", sum(it(t.get("hold_s")) for t in trades) / n],
            ["平均持倉根數", sum(fl(t.get("bars")) for t in trades) / n],
            ["平均最高利潤", sum(fl(t.get("peak_pct")) for t in trades) / n],
            ["毛損益", tg_], ["手續費", tf_], ["淨損益", tn_],
            ["淨損益%", (tn_ / tnv * 100) if tnv else 0.0],
            ["最大單筆獲利", max(nets) if nets else 0.0],
            ["最大單筆虧損", min(nets) if nets else 0.0]]
    for r in rows: w2.append(r)
    for r in range(2, w2.max_row + 1):
        lab = str(w2.cell(r, 1).value or ""); cc = w2.cell(r, 2)
        if lab in ("勝率", "淨損益%", "平均最高利潤"): cc.number_format = '0.0000"%"'
        elif lab in ("毛損益", "手續費", "淨損益", "最大單筆獲利", "最大單筆虧損"):
            cc.number_format = "0.000000"
        elif lab.startswith("平均持倉"): cc.number_format = "0.0"
    def blk(title, header):
        w2.append([]); w2.append([title])
        w2.cell(w2.max_row, 1).font = Font(bold=True)
        w2.append(header)
    blk("── 依方向 ──", ["方向", "筆數", "勝率", "淨損益"])
    for d in ("L", "S"):
        sub = [t for t in trades if t.get("dir") == d]
        if not sub: continue
        sn = [fl(t.get("net")) for t in sub]
        w2.append(["多" if d == "L" else "空", len(sub),
                   len([x for x in sn if x > 0]) / len(sn) * 100, sum(sn)])
        w2.cell(w2.max_row, 3).number_format = '0.00"%"'
        w2.cell(w2.max_row, 4).number_format = "0.000000"
    blk("── 依幣種 ──", ["幣種", "筆數", "勝率", "淨損益"])
    for sy in sorted({t.get("sym") for t in trades if t.get("sym")}):
        sub = [t for t in trades if t.get("sym") == sy]
        sn = [fl(t.get("net")) for t in sub]
        w2.append([sy, len(sub), len([x for x in sn if x > 0]) / len(sn) * 100, sum(sn)])
        w2.cell(w2.max_row, 3).number_format = '0.00"%"'
        w2.cell(w2.max_row, 4).number_format = "0.000000"
    for col, w in (("A", 20), ("B", 14), ("C", 12), ("D", 14)):
        w2.column_dimensions[col].width = w
    wb.save(path)

def send_report_mail(path, name, day, cnt, tn):
    """寄出均K 日報。回傳 (ok, 訊息)。"""
    import smtplib
    from email.message import EmailMessage
    env = {}
    try:
        for line in open("/srv/1111bot/.env"):
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip().strip('"').strip("'")
    except Exception as e:
        return False, "讀 .env 失敗：%s" % e
    user = env.get("GMAIL_USER")
    pwd = (env.get("GMAIL_APP_PASSWORD") or "").replace(" ", "")
    to = env.get("REPORT_TO") or user
    if not user or not pwd:
        return False, "未設定 GMAIL_USER / GMAIL_APP_PASSWORD"
    msg = EmailMessage()
    msg["Subject"] = "OKX %s 均K日報 %s（%d 筆）" % (ACCT, day, cnt)
    msg["From"] = user; msg["To"] = to
    msg.set_content("帳號 %s\n策略 均K\n日期 %s\n交易筆數 %d\n淨損益 %+.6f USDT\n"
                    % (ACCT, day, cnt, tn))
    msg.add_attachment(open(path, "rb").read(),
                       maintype="application",
                       subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       filename=name)
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as sv:
        sv.login(user, pwd); sv.send_message(msg)
    return True, to

async def job_report(ctx):
    """00:05 寄出前一日日報（不發 TG，結果只寫 log）。"""
    day = (now8() - timedelta(days=1)).strftime("%Y-%m-%d")
    trades = load_trades(day)
    if not trades:
        print("均K 日報：%s 無交易，未寄送" % day); return
    name = f"OKX_{ACCT}_均K日報_{day.replace('-', '')}.xlsx"
    path = f"/srv/1111bot/data/{name}"
    try:
        build_daily_xlsx(day, trades, path)
    except Exception as e:
        print("均K 日報產生失敗", type(e).__name__, e); return
    tn = 0.0
    for t in trades:
        try: tn += float(t.get("net") or 0)
        except Exception: pass
    try:
        ok, info = send_report_mail(path, name, day, len(trades), tn)
    except Exception as e:
        print("均K 日報寄送失敗", type(e).__name__, e, "檔案：", path); return
    if ok:
        print("均K 日報已寄出 %s（%d 筆，淨損益 %+.6f）-> %s" % (day, len(trades), tn, info))
    else:
        print("均K 日報未寄送：%s，檔案：%s" % (info, path))

# ---------- 每日自動 summary ----------
class _M:
    def __init__(self, app, chat): self._a = app; self._c = chat
    async def reply_text(self, t): await self._a.bot.send_message(self._c, t)
class _U:
    def __init__(self, app, chat): self.message = _M(app, chat)

async def job_summary(ctx):
    if not CHAT_ID: return
    try: await cmd_summary(_U(ctx.application, CHAT_ID), ctx)
    except Exception as e: print("auto summary fail", e)

# ---------- 啟動 ----------
async def _post_stop(app):
    global SHUTTING_DOWN
    SHUTTING_DOWN = True
    print("收到停止訊號，保留策略存檔供重啟接管")

async def _post_init(app):
    global HTTP
    HTTP = httpx.AsyncClient(timeout=httpx.Timeout(20.0, connect=10.0), limits=httpx.Limits(max_connections=40))
    CMDS = [BotCommand("status", "策略現況"), BotCommand("summary", "當日戰報"),
            BotCommand("coins", "幣種"), BotCommand("ha", "燈號+ATR 報表寄信"),
            BotCommand("run", "建立均K策略"), BotCommand("confirm", "確認執行"),
            BotCommand("stop", "停指定"), BotCommand("stopall", "停全部"),
            BotCommand("replay", "逐筆回溯報表寄信"), BotCommand("db", "行情DB狀態"), BotCommand("timeframe", "週期"),
            BotCommand("menu", "說明")]
    scopes = [BotCommandScopeDefault(), BotCommandScopeAllPrivateChats()]
    try:
        saved = json.load(open(STATE_FILE)) if os.path.exists(STATE_FILE) else {}
        ch = saved.get("chat")
        if ch: scopes.append(BotCommandScopeChat(ch))
    except Exception:
        pass
    for sc in scopes:
        try: await app.bot.delete_my_commands(scope=sc)
        except Exception as e: print("delete cmds fail", type(sc).__name__, e)
    await app.bot.set_my_commands(CMDS)
    print(f"左下 Menu 已更新（已清除 {len(scopes)} 個 scope 的舊指令）")
    try:
        jq = app.job_queue
        if jq:
            t2359 = datetime.strptime("23:59", "%H:%M").time().replace(tzinfo=TZ8)
            jq.run_daily(job_summary, time=t2359, name="daily_summary")
            t0005 = datetime.strptime("00:05", "%H:%M").time().replace(tzinfo=TZ8)
            jq.run_daily(job_report, time=t0005, name="daily_report")
            print("已排程：23:59 TG /summary｜00:05 Email 前一日日報")
    except Exception as e:
        print("schedule fail", e)
    try:
        db_open()
    except Exception as e:
        print("行情DB 開啟失敗", type(e).__name__, e)
    await startup_recover(app)
    hb = asyncio.create_task(hb_watch(app))
    _BG.add(hb); hb.add_done_callback(_BG.discard)
    print("心跳看門狗已啟動")
    col = asyncio.create_task(collector(app))
    _BG.add(col); col.add_done_callback(_BG.discard)
    fm = asyncio.create_task(frame_mover(app))
    _BG.add(fm); fm.add_done_callback(_BG.discard)

def main():
    os.makedirs(os.path.dirname(STATE_FILE), exist_ok=True)
    print(f"啟動 {ACCT} 均K B6-2（token ...{TOKEN[-6:]}）")
    app = (Application.builder().token(TOKEN).post_init(_post_init).post_stop(_post_stop)
           .connect_timeout(30.0).read_timeout(30.0).write_timeout(30.0)
           .pool_timeout(30.0).get_updates_read_timeout(40.0)
           .get_updates_connect_timeout(30.0).build())
    for cmd, fn in [(["menu", "start"], cmd_menu), ("run", cmd_run), ("confirm", cmd_confirm),
                    ("stop", cmd_stop), ("stopall", cmd_stopall), ("status", cmd_status),
                    ("summary", cmd_summary), ("ha", cmd_ha), ("replay", cmd_replay), ("db", cmd_db),
                    ("timeframe", cmd_timeframe), ("coins", cmd_coins)]:
        app.add_handler(CommandHandler(cmd, fn))
    app.add_handler(MessageHandler(filters.COMMAND, cmd_unknown))
    app.run_polling()

if __name__ == "__main__":
    main()
