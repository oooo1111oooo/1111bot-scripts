# -*- coding: utf-8 -*-
import sys, os, json, smtplib, datetime
from decimal import Decimal
from email.message import EmailMessage
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

BASE = os.environ.get("REPORT_BASE", "/srv/1111bot")
ACCT = os.environ.get("REPORT_ACCT", "o3333o")
KIND = os.environ.get("REPORT_KIND", "原K")
FONT = "PingFang TC"
SIZE = 12

HEAD = ["交易日期","進場時間","出場時間","進場參數","保證金","幣種","方向","出場原因",
        "埋伏秒數","持倉秒數","進場價","出場價","毛損益","毛損益率(%)","手續費","淨損益"]
NUM8 = '#,##0.00000000_ ;[Red]\\-#,##0.00000000\\ '
NUM6 = '#,##0.000000_ ;[Red]\\-#,##0.000000\\ '
NUM0 = '0_ ;[Red]\\-0\\ '
PCT  = '0.000000%'
WIDTH = {"A":3.3,"B":10.7,"C":8.6,"D":8.6,"E":8.6,"F":8.6,"G":11.1,"H":4.7,
         "I":9.4,"J":8.6,"K":8.6,"L":12.7,"M":12.7,"N":12.7,"O":11.9,"P":11.7,"Q":12.7}
REASON = {"Take_Profit":"TP","Stop_Loss":"SP","Time_Exit":"TE"}

def load_env(p):
    d = {}
    try:
        for line in open(p):
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                d[k.strip()] = v.strip().strip('"').strip("'")
    except Exception:
        pass
    return d

def dnum(v, default="0"):
    try: return float(Decimal(str(v if v not in (None, "") else default)))
    except Exception: return 0.0

def build(day, recs, out):
    wb = Workbook(); ws = wb.active; ws.title = "工作表1"
    for col, w in WIDTH.items(): ws.column_dimensions[col].width = w
    for i, h in enumerate(HEAD):
        c = ws.cell(row=2, column=2+i, value=h)
        c.font = Font(name=FONT, size=SIZE); c.alignment = Alignment(horizontal="center")
    d0 = datetime.datetime.strptime(day, "%Y%m%d")
    r = 3
    for rec in recs:
        gross = dnum(rec.get("gross")); fee = dnum(rec.get("fee"))
        net = dnum(rec.get("net")); nv = dnum(rec.get("nv"))
        rate = (gross / nv) if nv else 0.0
        vals = [d0, rec.get("in_ts",""), rec.get("ts",""), rec.get("tf",""), dnum(rec.get("margin")),
                rec.get("sym",""), rec.get("dir",""), REASON.get(rec.get("reason"), rec.get("reason","")),
                dnum(rec.get("ambush_s")), dnum(rec.get("hold_s")),
                dnum(rec.get("in_px")), dnum(rec.get("out_px")),
                gross, rate, fee, net]
        for i, v in enumerate(vals):
            c = ws.cell(row=r, column=2+i, value=v)
            c.font = Font(name=FONT, size=SIZE)
        ws.cell(row=r, column=2).number_format = "mm-dd-yy"
        for col in (3, 4): ws.cell(row=r, column=col).number_format = "h:mm:ss;@"
        for col in (10, 11): ws.cell(row=r, column=col).number_format = NUM0
        for col in (12, 13): ws.cell(row=r, column=col).number_format = NUM6
        for col in (14, 16, 17): ws.cell(row=r, column=col).number_format = NUM8
        ws.cell(row=r, column=15).number_format = PCT
        r += 1
    ws.freeze_panes = "A3"
    wb.save(out)
    return r - 3

def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    day = args[0] if args else datetime.datetime.now().strftime("%Y%m%d")
    nomail = "--no-mail" in sys.argv
    src = "%s/data/trades_%s_%s.json" % (BASE, ACCT, day)
    try: recs = json.load(open(src))
    except Exception as e:
        print("找不到紀錄檔:", src, e); return 1
    recs.sort(key=lambda x: str(x.get("ts") or ""))
    name = "OKX_%s_%s_%s.xlsx" % (ACCT, KIND, day)
    out = "%s/data/%s" % (BASE, name)
    n = build(day, recs, out)
    print("已產生 %s（%d 筆）" % (out, n))
    if nomail: return 0
    env = load_env("%s/.env" % BASE)
    user = env.get("GMAIL_USER"); pwd = (env.get("GMAIL_APP_PASSWORD") or "").replace(" ", "")
    to = env.get("REPORT_TO") or user
    if not user or not pwd:
        print("未設定 GMAIL_USER / GMAIL_APP_PASSWORD，略過寄送"); return 0
    m = EmailMessage()
    m["Subject"] = "OKX %s %s 日報 %s（%d 筆）" % (ACCT, KIND, day, n)
    m["From"] = user; m["To"] = to
    m.set_content("附件為 %s 的交易明細，共 %d 筆。" % (day, n))
    m.add_attachment(open(out, "rb").read(),
                     maintype="application",
                     subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     filename=name)
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
        s.login(user, pwd); s.send_message(m)
    print("已寄送至", to)
    return 0

if __name__ == "__main__":
    sys.exit(main())
