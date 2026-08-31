#!/usr/bin/env python3
"""B6-1 原K｜o3333o — 核心重寫版
規格：
  1. 每根 K 線開盤即掛限價埋伏單；未成交於收線前 3 秒撤單。
  2. 遲到一律立刻補掛，除非距離收線不足 30 秒（避免與下一根碰撞）才跳過。
  3. 一根 K 線只掛一次。進場後依 TP/SL/TE 出場，出場後等下一根開盤。
  4. OKX 為唯一真相來源：撤單、持倉、損益一律回查 OKX 確認。
  5. 重啟時接管 OKX 上的既有持倉與掛單，不留孤兒。
  6. Telegram 為旁路：發送失敗絕不影響交易流程。
"""
import sys, hmac, base64, hashlib, json, time, asyncio, uuid, os
from decimal import Decimal, ROUND_FLOOR, ROUND_CEILING, ROUND_DOWN
from datetime import datetime, timezone, timedelta
import httpx
from telegram import BotCommand, BotCommandScopeDefault, BotCommandScopeAllPrivateChats, BotCommandScopeChat
from telegram.ext import Application, CommandHandler, MessageHandler, filters
sys.path.insert(0, "/srv/1111bot")
from app.core import emoji as E
from app.strategy.normal import next_open_epoch as _noe_unused, TF_SEC as _TFS_unused

# 原K 專用時間框架（皆整除 60 分鐘，起訖時刻自然對齊整點）
TF_SEC = {"3m": 180, "4m": 240, "5m": 300, "6m": 360, "10m": 600}

def next_open_epoch(now_epoch, tf):
    sec = TF_SEC[tf]
    return ((now_epoch // sec) + 1) * sec

BASE = "https://www.okx.com"
ACCT = "o3333o"
TZ8 = timezone(timedelta(hours=8))
ACCOUNT_TF = "5m"
STATE_FILE = "/srv/1111bot/data/strategies_o3333o.json"
HOLD_SEC = 120       # 固定持倉秒數（TE）：進場後最多持有幾秒
ENTRY_CUTOFF = 120   # TF 剩餘不足幾秒就放棄進場（＝HOLD_SEC，確保持倉能跑滿）
CLOSE_LEAD = 2       # 安全上限：無論如何不晚於 TF 結束前幾秒平倉

def load_env(p):
    d = {}
    for line in open(p):
        line = line.strip()
        if "=" in line and not line.startswith("#"):
            k, v = line.split("=", 1); d[k] = v
    return d

ACC = load_env("/srv/1111bot/config/accounts.env")
BOTS = load_env("/srv/1111bot/config/bots.env")
TOKEN = BOTS["BOT_o3333o_NORMAL"]
SYMS = json.load(open("/srv/1111bot/config/symbols.json"))["symbols"]

PENDING = {}; STRATS = {}; TASKS = {}; STATS = {}
CHAT_ID = None
SHUTTING_DOWN = False
HTTP = None
SPEC_CACHE = {}

def skey(s, d): return f"{s}_{d}"
def inst_id(s): return s.replace("USDT", "") + "-USDT-SWAP"
def now8(): return datetime.now(TZ8)
def hhmmss(): return now8().strftime("%H:%M:%S")
def today8(): return now8().strftime("%Y-%m-%d")
def pct(v):
    """去尾零但不用科學記號：10 -> "10"（非 "1E+1"），0.50 -> "0.5"。"""
    d = Decimal(str(v)).normalize()
    if d == d.to_integral_value():
        d = d.quantize(Decimal(1))
    return str(d)

# ---------- 狀態持久化（原子寫入） ----------
SAVE_FIELDS = ("sym","dir","tf","lev","margin","offset","tp","sl","chat",
               "pos_open","pos_px","pos_tp","pos_sl","pos_ee","pos_pt","pos_oid","last_open","catchup")

def save_state(_open=open, _replace=os.replace, _fsync=os.fsync, _dump=json.dump):
    # 關閉流程中絕不寫檔：此時 loop() 的 finally 會逐一 pop 掉 STRATS，
    # 若照常寫入就會把存檔覆蓋成空的，導致重啟後策略全滅。
    if SHUTTING_DOWN:
        return
    try:
        data = {"chat": CHAT_ID, "tf": ACCOUNT_TF, "stats": STATS, "strats": []}
        for k, S in STRATS.items():
            if S.get("alive"):
                data["strats"].append({a: S[a] for a in SAVE_FIELDS if a in S})
        def enc(o): return str(o) if isinstance(o, Decimal) else o
        tmp = STATE_FILE + ".tmp"
        with _open(tmp, "w") as f:
            _dump(data, f, default=enc); f.flush(); _fsync(f.fileno())
        _replace(tmp, STATE_FILE)
    except Exception as e:
        print("save_state fail", e)

def bump(k, field):
    t = today8()
    if k not in STATS or STATS[k]["date"] != t:
        STATS[k] = {"date": t, "placed": 0, "entered": 0}
    STATS[k][field] += 1
    save_state()

def get_stat(k):
    t = today8()
    if k not in STATS or STATS[k]["date"] != t: return (0, 0)
    return (STATS[k]["placed"], STATS[k]["entered"])

# ---------- 交易紀錄 ----------
def trade_file(t):
    return STATE_FILE.replace("strategies_", "trades_").replace(".json", "_" + str(t).replace("-", "") + ".json")

def load_trades(t):
    try: return json.load(open(trade_file(t)))
    except Exception: return []

def log_trade(rec):
    try:
        fp = trade_file(rec.get("date"))
        try: arr = json.load(open(fp))
        except Exception: arr = []
        arr.append(rec)
        tmp = fp + ".tmp"
        with open(tmp, "w") as f:
            json.dump(arr, f, default=str); f.flush(); os.fsync(f.fileno())
        os.replace(tmp, fp)
    except Exception as e:
        print("log_trade fail", e)

# ---------- OKX API（全非同步，不阻塞事件迴圈） ----------
def ts_now():
    n = datetime.now(timezone.utc)
    return n.strftime("%Y-%m-%dT%H:%M:%S.") + f"{n.microsecond//1000:03d}Z"

def sign(sec, ts, m, p, b=""):
    return base64.b64encode(hmac.new(sec.encode(), f"{ts}{m}{p}{b}".encode(), hashlib.sha256).digest()).decode()

async def api(method, path, body=None):
    b = json.dumps(body) if body else ""
    ts = ts_now()
    h = {"OK-ACCESS-KEY": ACC[f"OKX_{ACCT}_API_KEY"],
         "OK-ACCESS-SIGN": sign(ACC[f"OKX_{ACCT}_SECRET"], ts, method, path, b),
         "OK-ACCESS-TIMESTAMP": ts,
         "OK-ACCESS-PASSPHRASE": ACC[f"OKX_{ACCT}_PASSPHRASE"],
         "Content-Type": "application/json"}
    try:
        r = await HTTP.request(method, BASE + path, headers=h, content=b)
        return r.json()
    except Exception as e:
        print("api fail", path, type(e).__name__, e)
        return {"code": "-1", "msg": str(e), "data": []}

async def pub(path):
    try:
        r = await HTTP.get(BASE + path)
        return r.json()
    except Exception as e:
        print("pub fail", path, type(e).__name__, e)
        return {"code": "-1", "msg": str(e), "data": []}

async def get_spec(s):
    if s in SPEC_CACHE: return SPEC_CACHE[s]
    iid = inst_id(s)
    r = await pub(f"/api/v5/public/instruments?instType=SWAP&instId={iid}")
    d = r["data"][0]
    spec = {"iid": iid, "tick": Decimal(d["tickSz"]), "lot": Decimal(d["lotSz"]),
            "minsz": Decimal(d["minSz"]), "ctval": Decimal(d["ctVal"]),
            "maxlev": Decimal(d["lever"]), "ctvalccy": d["ctValCcy"]}
    SPEC_CACHE[s] = spec
    return spec

async def get_last(iid):
    r = await pub(f"/api/v5/market/ticker?instId={iid}")
    return Decimal(r["data"][0]["last"])

def align(px, tick, d):
    return (px / tick).to_integral_value(rounding=ROUND_FLOOR if d == "L" else ROUND_CEILING) * tick

def csize(m, lev, px, cv, lot):
    return ((m * lev / px) / cv / lot).to_integral_value(rounding=ROUND_DOWN) * lot

# ---------- Telegram（旁路：永不阻塞交易） ----------
_BG = set()

async def _send_bg(app, chat, t):
    try:
        await app.bot.send_message(chat, t)
    except Exception as e:
        print("notify fail", type(e).__name__, e)

async def notify(app, chat, t):
    try:
        tk = asyncio.create_task(_send_bg(app, chat, t))
        _BG.add(tk); tk.add_done_callback(_BG.discard)
    except Exception as e:
        print("notify schedule fail", e)

async def reply(u, t):
    """指令回覆：失敗重試一次，再失敗只記錄，不拋出。"""
    for i in range(2):
        try:
            await u.message.reply_text(t); return True
        except Exception as e:
            print("reply fail", i, type(e).__name__, e)
            await asyncio.sleep(2)
    return False

# ---------- OKX 事實查詢 ----------
async def okx_pos(iid, ps):
    r = await api("GET", "/api/v5/account/positions")
    if r.get("code") != "0": return None
    for p in (r.get("data") or []):
        if p.get("instId") == iid and p.get("posSide") == ps:
            try:
                if float(p.get("pos") or 0) != 0: return p
            except Exception: pass
    return None

async def okx_orders(iid=None, ps=None, prefix="n"):
    r = await api("GET", "/api/v5/trade/orders-pending")
    if r.get("code") != "0": return []
    out = []
    for o in (r.get("data") or []):
        if iid and o.get("instId") != iid: continue
        if ps and o.get("posSide") != ps: continue
        if prefix and not str(o.get("clOrdId") or "").startswith(prefix): continue
        out.append(o)
    return out

async def cancel_verified(iid, oid, tries=4):
    """撤單並回查 OKX 確認。回傳 canceled / filled / fail"""
    for i in range(tries):
        await api("POST", "/api/v5/trade/cancel-order", {"instId": iid, "ordId": oid})
        await asyncio.sleep(0.8)
        st = await api("GET", f"/api/v5/trade/order?instId={iid}&ordId={oid}")
        if st.get("code") == "0" and st.get("data"):
            s2 = st["data"][0].get("state")
            if s2 == "canceled": return "canceled"
            if s2 == "filled": return "filled"
    return "fail"

async def sweep(iid, pos, keep=None):
    """清掉本 bot 在該幣種該方向的所有殘留掛單。"""
    n = 0
    for o in await okx_orders(iid, pos):
        if keep and o.get("ordId") == keep: continue
        await api("POST", "/api/v5/trade/cancel-order", {"instId": iid, "ordId": o["ordId"]})
        n += 1
    return n

async def order_detail(iid, oid, tries=10):
    """查訂單成交明細（成交後即時可得）。回傳 {avgPx, fee, sz} 或 None。"""
    if not oid:
        return None
    for _ in range(tries):
        st = await api("GET", f"/api/v5/trade/order?instId={iid}&ordId={oid}")
        if st.get("code") == "0" and st.get("data"):
            dd = st["data"][0]
            if dd.get("state") == "filled" and dd.get("avgPx"):
                return {"avgPx": Decimal(dd["avgPx"]),
                        "fee": Decimal(dd.get("fee") or "0"),
                        "sz": Decimal(dd.get("accFillSz") or dd.get("sz") or "0")}
        await asyncio.sleep(1)
    return None

def _from_ph(ph, size, ctval, fpx):
    """由 positions-history 取值（OKX 官方口徑）。"""
    g = Decimal(ph.get("pnl") or "0")
    fee = Decimal(ph.get("fee") or "0") + Decimal(ph.get("fundingFee") or "0")
    net = Decimal(ph.get("realizedPnl") or "0")
    xpx = Decimal(ph.get("closeAvgPx") or "0")
    opx = Decimal(ph.get("openAvgPx") or "0") or fpx
    csz = Decimal(ph.get("closeTotalPos") or "0") or size
    nv = opx * csz * ctval
    return g, fee, net, nv, xpx

def _from_orders(od_in, od_out, d, size, ctval):
    """由進出場訂單成交明細取值（OKX 實收手續費與成交均價）。"""
    opx = od_in["avgPx"]; xpx = od_out["avgPx"]
    fee = od_in["fee"] + od_out["fee"]
    g = (xpx - opx) * size * ctval if d == "L" else (opx - xpx) * size * ctval
    return g, fee, g + fee, opx * size * ctval, xpx

def _rec_epoch_ms(rec):
    """由紀錄的 date + ts 還原出場時刻（毫秒）。"""
    try:
        s = str(rec.get("date")) + " " + str(rec.get("ts"))
        return int(datetime.strptime(s, "%Y-%m-%d %H:%M:%S").replace(tzinfo=TZ8).timestamp() * 1000)
    except Exception:
        return 0

async def _find_ph(iid, ps, t_ms, before=180000, after=1800000):
    """在 positions-history 找出時間最接近 t_ms 的那筆平倉紀錄。"""
    if not t_ms:
        return None
    r = await api("GET", f"/api/v5/account/positions-history?instType=SWAP&instId={iid}&limit=100")
    if r.get("code") != "0":
        return None
    best = None; bd = None
    for p in (r.get("data") or []):
        if p.get("posSide") != ps:
            continue
        ut = int(p.get("uTime") or 0)
        if ut < t_ms - before or ut > t_ms + after:
            continue
        dd = abs(ut - t_ms)
        if bd is None or dd < bd:
            best = p; bd = dd
    return best

async def reconcile_pending(app, days=3):
    """背景補帳：掃描近幾日紀錄檔，把 PENDING 的財務數字向 OKX 補齊。"""
    filled = 0
    for back in range(days):
        t = (now8() - timedelta(days=back)).strftime("%Y-%m-%d")
        fp = trade_file(t)
        try:
            arr = json.load(open(fp))
        except Exception:
            continue
        changed = False
        for rec in arr:
            if rec.get("src") != "PENDING":
                continue
            # 查詢鍵可能缺（舊紀錄），一律從 sym/dir/date/ts 推導，確保都能補
            iid = rec.get("iid") or inst_id(rec.get("sym", ""))
            pos = rec.get("pos") or ("long" if rec.get("dir") == "L" else "short")
            t_ms = int(rec.get("t0") or 0) or _rec_epoch_ms(rec)
            if not iid or not pos or not t_ms:
                continue
            try:
                ctv = (await get_spec(rec["sym"]))["ctval"]
            except Exception:
                ctv = Decimal("1")
            fpx = Decimal(rec.get("in_px") or "0")
            g = fee = net = nv = xpx = None; src = None
            ph = await _find_ph(iid, pos, t_ms)
            if ph:
                g, fee, net, nv, xpx = _from_ph(ph, Decimal("0"), ctv, fpx)
                src = "OKX"
            elif rec.get("xoid") and rec.get("eoid"):
                od_out = await order_detail(iid, rec.get("xoid"), tries=1)
                od_in = await order_detail(iid, rec.get("eoid"), tries=1)
                if od_out and od_in:
                    sz = od_out.get("sz") or Decimal("0")
                    g, fee, net, nv, xpx = _from_orders(od_in, od_out, rec["dir"], sz, ctv)
                    src = "訂單"
            if src:
                rec["gross"] = str(g); rec["fee"] = str(fee)
                rec["net"] = str(net); rec["nv"] = str(nv)
                rec["out_px"] = str(xpx); rec["src"] = src
                rec["filled_at"] = hhmmss()
                changed = True; filled += 1
            await asyncio.sleep(0.2)
        if changed:
            try:
                tmp = fp + ".tmp"
                with open(tmp, "w") as f:
                    json.dump(arr, f, default=str); f.flush(); os.fsync(f.fileno())
                os.replace(tmp, fp)
            except Exception as e:
                print("reconcile write fail", e)
    return filled

async def reconcile_watch(app):
    """每 5 分鐘嘗試補帳一次；補到就通知。"""
    while True:
        await asyncio.sleep(300)
        try:
            n = await reconcile_pending(app)
            if n and CHAT_ID:
                await notify(app, CHAT_ID, f"{E.BOT} ✅ 補帳完成：已向 OKX 回填 {n} 筆待補紀錄")
        except Exception as e:
            print("reconcile_watch fail", e)

async def close_record(iid, ps, after_ms, tries=3):
    """出場後取 OKX 真實平倉紀錄。"""
    for i in range(tries):
        r = await api("GET", f"/api/v5/account/positions-history?instType=SWAP&instId={iid}&limit=10")
        if r.get("code") == "0":
            for p in (r.get("data") or []):
                if p.get("posSide") == ps and int(p.get("uTime") or 0) >= after_ms:
                    return p
        await asyncio.sleep(1)
    return None

# ---------- 出場處理（共用：正常出場與重啟接管都走這裡） ----------
async def do_exit(app, S, spec, iid, d, pos, size, fpx, tp, sl, ee, pt, reason, k):
    cs = "sell" if d == "L" else "buy"
    # 平倉張數一律以 OKX 實際持倉為準（避免重複掛單造成倉位加倍時只平掉一半）
    p_now = await okx_pos(iid, pos)
    if p_now:
        try:
            real = abs(Decimal(str(p_now.get("availPos") or p_now.get("pos") or "0")))
            if real > 0:
                if real != size:
                    print("size mismatch", S.get("sym"), d, "local", size, "okx", real)
                size = real
        except Exception:
            pass
    else:
        await notify(app, S["chat"], f"{E.BOT} {S['sym']} {E.dir_word(d)} 平倉時 OKX 已無持倉，略過")
        for a in ("pos_open", "pos_px", "pos_tp", "pos_sl", "pos_ee", "pos_pt", "pos_oid"):
            S.pop(a, None)
        save_state()
        return True
    t0 = int(time.time() * 1000)
    xr = await api("POST", "/api/v5/trade/order",
                   {"instId": iid, "tdMode": "isolated", "side": cs, "posSide": pos,
                    "ordType": "market", "sz": str(size),
                    "clOrdId": "x" + uuid.uuid4().hex[:14]})
    if xr.get("code") != "0":
        em = (xr.get("data") or [{}])[0].get("sMsg") or xr.get("msg")
        await notify(app, S["chat"], f"{E.BOT} {E.LOSS} {S['sym']} {E.dir_word(d)} 平倉失敗：{em}\n⚠ 倉位可能仍在，請至 OKX 確認")
        return False
    # 財務數字一律以 OKX 為準，絕不估算。
    #   1. positions-history（官方口徑，含資金費）
    #   2. 訂單成交明細（OKX 實收手續費與成交均價）
    #   3. 兩者皆取不到 -> 寫入 null 並標記 pending，由背景補帳任務事後回填
    xoid = (xr.get("data") or [{}])[0].get("ordId")
    eoid = S.get("pos_oid")
    g = fee = net = nv = xpx = None
    src = "PENDING"
    ph = await close_record(iid, pos, t0, tries=3)
    if ph:
        g, fee, net, nv, xpx = _from_ph(ph, size, spec["ctval"], fpx)
        src = "OKX"
    else:
        od_out = await order_detail(iid, xoid, tries=8)
        od_in = await order_detail(iid, eoid, tries=2) if eoid else None
        if od_out and od_in:
            g, fee, net, nv, xpx = _from_orders(od_in, od_out, d, size, spec["ctval"])
            src = "訂單"
    hs = int(time.time() - ee)
    amb_s = f"{round(ee - pt)}s" if pt else "-"
    rec = {"date": today8(), "sym": S["sym"], "dir": d, "reason": reason,
           "ambush_s": round(ee - pt) if pt else 0, "hold_s": hs,
           "gross": None if g is None else str(g),
           "fee": None if fee is None else str(fee),
           "net": None if net is None else str(net),
           "nv": None if nv is None else str(nv),
           "src": src, "ts": hhmmss(),
           "in_ts": datetime.fromtimestamp(ee, TZ8).strftime("%H:%M:%S"),
           "tf": S["tf"], "margin": str(S["margin"]),
           "in_px": str(fpx), "out_px": None if xpx is None else str(xpx),
           "iid": iid, "pos": pos, "t0": t0, "eoid": eoid, "xoid": xoid}
    log_trade(rec)
    if src == "PENDING":
        await notify(app, S["chat"],
            f"{E.BOT} {E.LOSS} {S['sym']} {E.dir_word(d)} 已平倉，但 OKX 損益尚未回報\n"
            f"出場原因：{reason}｜持倉 {hs}s\n"
            f"已標記待補帳，背景任務會自動回填（可用 /audit 追蹤）\n時間：{hhmmss()}")
        for a in ("pos_open", "pos_px", "pos_tp", "pos_sl", "pos_ee", "pos_pt", "pos_oid"):
            S.pop(a, None)
        save_state()
        return True
    gp = (g / nv * 100) if nv else Decimal(0)
    fp = (fee / nv * 100) if nv else Decimal(0)
    npv = (net / nv * 100) if nv else Decimal(0)
    await notify(app, S["chat"],
        f"{E.BOT} OKX原K｜{ACCT}\n事件：{'🟢' if net >= 0 else '🔴'} 已出場\n"
        f"商　　品：{E.dir_emoji(d)} {S['sym']} {E.dir_word(d)}\n出場原因：{reason}\n"
        f"進場:{fpx}({pct(S['offset'])}%) | {datetime.fromtimestamp(ee, TZ8).strftime('%H:%M:%S')} | {amb_s}\n"
        f"止盈:{tp}({pct(S['tp'])}%)\n止損:{sl}({pct(S['sl'])}%)\n"
        f"出場:{xpx}({gp:+.3f}%) | {hhmmss()} | {hs}s\n"
        f"毛損益：{g:+.6f} ({gp:+.3f}%)\n手續費：{fee:+.6f} ({fp:+.3f}%)\n"
        f"淨損益：{net:+.6f} ({npv:+.3f}%) {E.pnl_emoji(net)}\n"
        + ("" if src == "OKX" else f"⚠ 來源：{src}\n")
        + f"時間：{hhmmss()}")
    for a in ("pos_open", "pos_px", "pos_tp", "pos_sl", "pos_ee", "pos_pt", "pos_oid"):
        S.pop(a, None)
    save_state()
    return True

# ---------- 持倉監控 ----------
async def monitor(app, S, spec, iid, d, pos, size, fpx, tp, sl, ee, pt, k, tf_end):
    """固定持倉 HOLD_SEC 秒即平倉（TE）；tf_end 僅作為不跨輪的安全上限。"""
    chk = 0
    while S["alive"]:
        await asyncio.sleep(1)
        chk += 1
        last = await get_last(iid)
        reason = None
        if d == "L":
            if last >= tp: reason = "Take_Profit"
            elif last <= sl: reason = "Stop_Loss"
        else:
            if last <= tp: reason = "Take_Profit"
            elif last >= sl: reason = "Stop_Loss"
        if not reason and (time.time() - ee >= HOLD_SEC or time.time() >= tf_end - CLOSE_LEAD):
            reason = "Time_Exit"
        if not reason and chk % 16 == 0:
            p_chk = await okx_pos(iid, pos)
            if p_chk:
                try:
                    rs = abs(Decimal(str(p_chk.get("pos") or "0")))
                    if rs > 0 and rs != size:
                        print("position size changed", S["sym"], d, size, "->", rs)
                        size = rs
                except Exception:
                    pass
            if not p_chk:
                # 手動平倉也必須入 DB，否則 DB 與 OKX 永遠對不起來
                log_trade({"date": today8(), "sym": S["sym"], "dir": d, "reason": "Manual_Close",
                           "ambush_s": round(ee - pt) if pt else 0,
                           "hold_s": int(time.time() - ee),
                           "gross": None, "fee": None, "net": None, "nv": None,
                           "src": "PENDING", "ts": hhmmss(),
                           "in_ts": datetime.fromtimestamp(ee, TZ8).strftime("%H:%M:%S"),
                           "tf": S["tf"], "margin": str(S["margin"]),
                           "in_px": str(fpx), "out_px": None,
                           "iid": iid, "pos": pos, "t0": int(time.time() * 1000),
                           "eoid": S.get("pos_oid"), "xoid": None})
                await notify(app, S["chat"],
                    f"{E.BOT} {S['sym']} {E.dir_word(d)} OKX 已無持倉（可能手動平倉）\n"
                    f"已記入 DB 並標記待補帳，背景任務會向 OKX 取回損益\n時間：{hhmmss()}")
                for a in ("pos_open", "pos_px", "pos_tp", "pos_sl", "pos_ee", "pos_pt", "pos_oid"):
                    S.pop(a, None)
                save_state(); return
        if reason:
            await do_exit(app, S, spec, iid, d, pos, size, fpx, tp, sl, ee, pt, reason, k)
            return
    # 策略被停止但仍持倉
    if await okx_pos(iid, pos):
        await notify(app, S["chat"], f"{E.BOT} {S['sym']} {E.dir_word(d)} 策略停止但仍有持倉，請至 OKX 處理")

# ---------- 主迴圈：一根 K 線一輪 ----------
async def loop(app, chat, S):
    spec = S["spec"]; iid = spec["iid"]; d = S["dir"]
    pos = "long" if d == "L" else "short"
    k = skey(S["sym"], d)
    try:
        # 重啟接管：OKX 上已有持倉 -> 直接進監控
        if S.get("pos_open"):
            p = await okx_pos(iid, pos)
            if p:
                fpx = Decimal(S.get("pos_px") or p.get("avgPx") or "0")
                tp = Decimal(S["pos_tp"]); sl = Decimal(S["pos_sl"])
                ee = float(S.get("pos_ee") or time.time())
                pt0 = float(S["pos_pt"]) if S.get("pos_pt") else None
                size = abs(Decimal(p.get("pos") or "0"))
                S["state"] = "持倉中"; save_state()
                tfs = TF_SEC[S["tf"]]
                tf_end0 = (int(time.time()) // tfs + 1) * tfs
                await notify(app, chat, f"{E.BOT} {E.dir_emoji(d)} {S['sym']} {E.dir_word(d)} 已接管既有持倉，恢復 TP/SL/TE 監控")
                await monitor(app, S, spec, iid, d, pos, size, fpx, tp, sl, ee, pt0, k, tf_end0)
            else:
                for a in ("pos_open", "pos_px", "pos_tp", "pos_sl", "pos_ee", "pos_pt", "pos_oid"):
                    S.pop(a, None)
                save_state()

        while S["alive"]:
            tf_sec = TF_SEC[S["tf"]]
            now = time.time()
            cur = int(now // tf_sec) * tf_sec
            room = cur + tf_sec - now
            # 僅在「上一輪未成交、剛撤完單」的情況下才允許盤中補掛；
            # 新建策略與出場後一律等下一根 K 線開盤，嚴守一根 K 線一輪。
            if S.get("catchup") and cur != S.get("last_open") and room >= ENTRY_CUTOFF:
                oe = cur
            else:
                S["state"] = "等下輪"; save_state()
                oe = next_open_epoch(int(time.time()), S["tf"])
                w = oe - time.time()
                if w > 0: await asyncio.sleep(w)
                if not S["alive"]: break
            S["last_open"] = oe; S["catchup"] = False; save_state()

            # 開盤取價 -> 埋伏價 -> 張數
            op = await get_last(iid)
            amb = align(op * (1 - S["offset"] / 100) if d == "L" else op * (1 + S["offset"] / 100), spec["tick"], d)
            size = csize(S["margin"], Decimal(S["lev"]), amb, spec["ctval"], spec["lot"])
            if size < spec["minsz"]:
                await notify(app, chat, f"{E.BOT} {E.LOSS} {S['sym']} {E.dir_word(d)} 保證金不足，循環停止")
                break

            await sweep(iid, pos)                 # 掛新單前先清乾淨
            r = await api("POST", "/api/v5/trade/order",
                          {"instId": iid, "tdMode": "isolated", "side": "buy" if d == "L" else "sell",
                           "posSide": pos, "ordType": "limit", "px": str(amb), "sz": str(size),
                           "clOrdId": "n" + uuid.uuid4().hex[:14]})
            if r.get("code") != "0":
                em = (r.get("data") or [{}])[0].get("sMsg") or r.get("msg")
                await notify(app, chat, f"{E.BOT} {E.LOSS} {S['sym']} {E.dir_word(d)} 掛單失敗：{em}")
                await asyncio.sleep(3); continue
            oid = r["data"][0]["ordId"]
            S["state"] = "委託中"; bump(k, "placed")
            pt = time.time()
            # 掛單後確認同幣同向只剩這一張，撤掉任何重複單
            dup = await sweep(iid, pos, keep=oid)
            if dup:
                print("dup orders cleared", S["sym"], d, dup)
                await notify(app, chat, f"{E.BOT} {S['sym']} {E.dir_word(d)} 已清除殘留掛單 {dup} 筆")

            # 輪詢成交，直到 TF 剩餘不足 ENTRY_CUTOFF 秒
            tf_end = oe + tf_sec
            deadline = tf_end - ENTRY_CUTOFF   # 剩 60 秒就不再等成交
            filled = False; fpx = None
            while S["alive"] and time.time() < deadline:
                await asyncio.sleep(2)
                st = await api("GET", f"/api/v5/trade/order?instId={iid}&ordId={oid}")
                if st.get("code") == "0" and st.get("data"):
                    s2 = st["data"][0]["state"]
                    if s2 == "filled":
                        filled = True; fpx = Decimal(st["data"][0]["avgPx"]); break
                    if s2 == "canceled":
                        break

            if not filled:
                rc = await cancel_verified(iid, oid)
                if rc == "filled":
                    st = await api("GET", f"/api/v5/trade/order?instId={iid}&ordId={oid}")
                    try: fpx = Decimal(st["data"][0]["avgPx"]); filled = True
                    except Exception: pass
                elif rc != "canceled":
                    await notify(app, chat, f"{E.BOT} {E.LOSS} {S['sym']} {E.dir_word(d)} 撤單未確認，本策略停止以免重複掛單")
                    S["alive"] = False; break
            if not S["alive"]: break
            if not filled:
                S["catchup"] = True
                continue

            # 成交 -> 算 TP/SL -> 監控
            bump(k, "entered")
            if d == "L":
                tp = align(fpx * (1 + S["tp"] / 100), spec["tick"], "S")
                sl = align(fpx * (1 - S["sl"] / 100), spec["tick"], "L")
            else:
                tp = align(fpx * (1 - S["tp"] / 100), spec["tick"], "L")
                sl = align(fpx * (1 + S["sl"] / 100), spec["tick"], "S")
            ee = time.time()
            S["state"] = "持倉中"
            S["pos_open"] = True; S["pos_px"] = str(fpx)
            S["pos_tp"] = str(tp); S["pos_sl"] = str(sl); S["pos_ee"] = ee; S["pos_pt"] = pt
            S["pos_oid"] = oid
            save_state()
            await notify(app, chat,
                f"{E.BOT} OKX原K｜{ACCT}\n事件：🔔 已進場成交\n"
                f"商　　品：{E.dir_emoji(d)} {S['sym']} {E.dir_word(d)}\n"
                f"進場:{fpx}({pct(S['offset'])}%) | {datetime.fromtimestamp(ee, TZ8).strftime('%H:%M:%S')} | {int(ee - pt)}s\n"
                f"止盈:{tp}({pct(S['tp'])}%)\n止損:{sl}({pct(S['sl'])}%)\n"
                f"狀　　態：📌 持倉中\n時間：{hhmmss()}")
            await monitor(app, S, spec, iid, d, pos, size, fpx, tp, sl, ee, pt, k, tf_end)
            # 出場後允許補掛：出場流程（查 OKX 真實損益）可能耗時數秒而跨進新 TF，
            # 若新 TF 尚未掛過且剩餘 >= ENTRY_CUTOFF 就立刻掛，避免整輪被跳過。
            # 若仍在同一個 TF（cur == last_open），迴圈頂端會照常睡到下一個 TF 開始。
            S["catchup"] = True
    except asyncio.CancelledError:
        raise
    except Exception as e:
        print("loop error", S.get("sym"), S.get("dir"), type(e).__name__, e)
        await notify(app, chat, f"{E.BOT} {E.LOSS} {S['sym']} {E.dir_word(d)} 循環錯誤：{type(e).__name__}: {e}")
    finally:
        if SHUTTING_DOWN:
            # 服務關閉：保留 STRATS 與存檔原狀，讓重啟後能完整認領
            S["state"] = "已停止"
        else:
            # 策略結束前先清掉自己掛在 OKX 上的單，否則會變成沒人認領的孤兒單
            try:
                left = await sweep(iid, pos)
                if left:
                    await notify(app, chat, f"{E.BOT} {E.LOSS} {S['sym']} {E.dir_word(d)} 策略結束，已清除自身殘留掛單 {left} 筆")
            except Exception as e:
                print("finally sweep fail", S.get("sym"), d, e)
            S["state"] = "已停止"; S["alive"] = False
            # 身分檢查：若 STRATS[k] 已被新策略取代，絕不能誤刪，
            # 否則新策略會從清單消失卻仍在背景掛單（幽靈策略）。
            if STRATS.get(k) is S:
                STRATS.pop(k, None)
            try:
                if TASKS.get(k) is asyncio.current_task():
                    TASKS.pop(k, None)
            except Exception:
                pass
            save_state()

# ---------- 啟動接管 ----------
async def rebuild_strat(d):
    spec = await get_spec(d["sym"])
    S = {"sym": d["sym"], "dir": d["dir"], "tf": d.get("tf", ACCOUNT_TF),
         "lev": int(d["lev"]), "margin": Decimal(str(d["margin"])),
         "offset": Decimal(str(d["offset"])), "tp": Decimal(str(d["tp"])),
         "sl": Decimal(str(d["sl"])), "spec": spec,
         "alive": True, "state": "等下輪", "chat": d.get("chat", CHAT_ID)}
    for a in ("pos_open", "pos_px", "pos_tp", "pos_sl", "pos_ee", "pos_pt", "pos_oid", "last_open", "catchup"):
        if a in d: S[a] = d[a]
    return S

async def startup_recover(app):
    global CHAT_ID, ACCOUNT_TF, STATS
    if not os.path.exists(STATE_FILE):
        print("無存檔"); return
    try:
        data = json.load(open(STATE_FILE))
    except Exception as e:
        print("讀存檔失敗", e); return
    CHAT_ID = data.get("chat"); ACCOUNT_TF = data.get("tf", "5m"); STATS = data.get("stats", {})
    saved = data.get("strats", [])
    if not saved:
        print("存檔無策略"); return
    rec = []; failed = []
    for d in saved:
        try:
            S = await rebuild_strat(d)
            k = skey(S["sym"], S["dir"])
            STRATS[k] = S
            TASKS[k] = asyncio.create_task(loop(app, S["chat"], S))
            rec.append(f"{E.dir_emoji(S['dir'])} {S['sym']} {E.dir_word(S['dir'])}")
        except Exception as e:
            print("重建失敗", d, e)
            failed.append(f"{d.get('sym')} {d.get('dir')}：{type(e).__name__}")
    pend = await api("GET", "/api/v5/trade/orders-pending")
    posr = await api("GET", "/api/v5/account/positions")
    n_ord = len(pend.get("data", [])) if pend.get("code") == "0" else 0
    n_pos = len([p for p in posr.get("data", []) if float(p.get("pos", "0")) != 0]) if posr.get("code") == "0" else 0
    print(f"已接管策略 {len(rec)}｜OKX 掛單{n_ord} 持倉{n_pos}")
    if failed:
        print("重建失敗清單:", failed)
        if CHAT_ID:
            await notify(app, CHAT_ID, f"{E.BOT} {E.LOSS} 重啟時有 {len(failed)} 個策略重建失敗：\n" +
                         "\n".join("・" + x for x in failed) + "\n⚠ 這些策略已消失，請確認 OKX 是否有殘留掛單")
    if CHAT_ID and rec and n_ord > len(rec):
        await notify(app, CHAT_ID, f"{E.BOT} {E.LOSS} OKX 掛單 {n_ord} 筆 > 策略 {len(rec)} 個，可能有孤兒單，請查 /status")
    if CHAT_ID and rec:
        await notify(app, CHAT_ID,
            f"{E.BOT} OKX原K｜{ACCT}\n事件：🔄 重啟認領完成\n━━━━━━━━━━\n"
            f"已接管策略（{len(rec)}）：\n" + "\n".join("・" + x for x in rec) +
            f"\nOKX 現況：掛單{n_ord} 持倉{n_pos}\n循環已接管，繼續運作\n時間：{hhmmss()}")

# ---------- TG 指令 ----------
# ---------- K 線 / 振幅（/amp 用） ----------
NATIVE_BARS = {"3m": "3m", "5m": "5m"}   # 4m/6m 無原生 K 線；10m 由兩根 5m 合成

async def get_klines(iid, bar, limit=300):
    """只取已收線（confirm=1）的 K 線，回傳舊->新。"""
    r = await pub(f"/api/v5/market/candles?instId={iid}&bar={bar}&limit={min(300, limit)}")
    if r.get("code") != "0":
        return []
    out = []
    for c in (r.get("data") or []):
        try:
            if len(c) >= 9 and str(c[8]) != "1":
                continue
            out.append({"ts": int(c[0]), "o": Decimal(c[1]), "h": Decimal(c[2]),
                        "l": Decimal(c[3]), "c": Decimal(c[4])})
        except Exception:
            continue
    out.reverse()
    return out

def _merge2(kl5):
    """兩根 5m 合成一根 10m，強制對齊 10 分鐘邊界。"""
    out = []
    i = 0
    n = len(kl5)
    while i < n:
        a = kl5[i]
        ts_a = int(a["ts"])
        if ts_a % 600000 != 0:
            i += 1; continue
        if i + 1 >= n:
            break
        b = kl5[i + 1]
        if int(b["ts"]) - ts_a != 300000:
            i += 1; continue
        out.append({"ts": ts_a, "o": a["o"], "h": max(a["h"], b["h"]),
                    "l": min(a["l"], b["l"]), "c": b["c"]})
        i += 2
    return out

async def klines_for_tf(iid, tf, want=300):
    """依 TF 取原始 K 線。10m 由兩根 5m 合成；其餘須為 OKX 原生週期。"""
    if tf == "10m":
        return _merge2(await get_klines(iid, "5m", 300))
    bar = NATIVE_BARS.get(tf)
    if not bar:
        return None
    return await get_klines(iid, bar, want)

def calc_amp(kl):
    """每根回傳 (振幅%, 漲跌幅%)。
    振幅% = (高-低) / 前一根收盤 * 100（恆正）
    漲跌幅% = (收盤-前一根收盤) / 前一根收盤 * 100（帶正負）
    第一根無前收，以本根開盤價替代。"""
    out = []
    for i, k in enumerate(kl):
        base = kl[i-1]["c"] if i > 0 else k["o"]
        if not base:
            out.append((Decimal(0), Decimal(0))); continue
        amp = (k["h"] - k["l"]) / base * 100
        chg = (k["c"] - base) / base * 100
        out.append((amp, chg))
    return out

async def _reply_long(u, head, lines, tail):
    """長清單拆多則送出（Telegram 單則上限 4096 字元）。"""
    LIM = 3500
    buf = list(head); msgs = []
    for ln in lines:
        if sum(len(x) + 1 for x in buf) + len(ln) + 1 > LIM and len(buf) > len(head):
            msgs.append("\n".join(buf)); buf = list(head)
        buf.append(ln)
    buf += tail
    msgs.append("\n".join(buf))
    for i, m in enumerate(msgs):
        if len(msgs) > 1:
            m = m.replace("事件：振幅檢視", "事件：振幅檢視（%d/%d）" % (i+1, len(msgs)), 1)
        await reply(u, m)

def strat_params(sym, dr):
    S = STRATS.get(skey(sym, dr))
    if not S or not S.get("alive"):
        return f"{dr}（已停止）"
    return (f"{dr} {S['lev']}x {pct(S['margin'])} {pct(S['offset'])} "
            f"{pct(S['tp'])} {pct(S['sl'])}")

async def cmd_run(u, c):
    global CHAT_ID; CHAT_ID = u.effective_chat.id
    a = c.args
    fmt = f"用法：/run 商品 方向 槓桿 保證金 埋伏 TP SL\n例：/run ETHUSDT L 1x 3 0.5 0.5 0.5\n（TE 已固定為 TF 結束，週期依 /timeframe，目前 {ACCOUNT_TF}）"
    if len(a) != 7: await reply(u, f"{E.BOT} 參數數量錯誤（需7個）\n{fmt}"); return
    try:
        sym = a[0].upper(); dr = a[1].upper(); lev = int(a[2].replace("x", ""))
        margin = Decimal(a[3]); offset = Decimal(a[4]); tp = Decimal(a[5])
        sl = Decimal(a[6])
    except Exception:
        await reply(u, f"{E.BOT} 參數格式錯誤\n{fmt}"); return
    if dr not in ("L", "S"): await reply(u, f"{E.BOT} 方向須 L 或 S"); return
    k = skey(sym, dr)
    if k in STRATS and STRATS[k].get("alive"):
        await reply(u, f"{E.BOT} {sym} {E.dir_word(dr)} 已在運行"); return
    try: spec = await get_spec(sym)
    except Exception: await reply(u, f"{E.LOSS} 找不到商品 {sym}"); return
    op = await get_last(spec["iid"])
    amb = align(op * (1 - offset / 100) if dr == "L" else op * (1 + offset / 100), spec["tick"], dr)
    size = csize(margin, Decimal(lev), amb, spec["ctval"], spec["lot"])
    if size < spec["minsz"]:
        need = spec["minsz"] * spec["ctval"] * op / Decimal(lev)
        await reply(u, f"{E.BOT} {E.LOSS} 保證金不足：算出 {size} 張 < 最小 {spec['minsz']}\n此槓桿下至少需約 {need:.4f} USDT"); return
    PENDING[u.effective_chat.id] = {"kind": "run", "t": time.time(), "sym": sym, "dir": dr, "tf": ACCOUNT_TF,
        "lev": lev, "margin": margin, "offset": offset, "tp": tp, "sl": sl, "spec": spec}
    await reply(u, f"{E.BOT} OKX原K｜{ACCT}\n事件：交易參數預覽\n━━━━━━━━━━\n"
        f"商　　品：{E.dir_emoji(dr)} {sym} {E.dir_word(dr)} {lev}x\n週　　期：{ACCOUNT_TF}\n"
        f"開盤估價：{op}\n埋伏距離：{offset}%\n埋伏價格：{amb}\n"
        f"止盈 TP：{tp}%\n止損 SL：{sl}%\n保 證 金：{margin} USDT\n下單張數：{size}\n"
        f"━━━━━━━━━━\n⚠ 確認後真實循環交易\n下一步：60秒內 /confirm\n時間：{hhmmss()}")
    asyncio.create_task(_to(c.application, u.effective_chat.id, PENDING[u.effective_chat.id]["t"]))

async def _to(app, chat, stamp):
    await asyncio.sleep(61)
    p = PENDING.get(chat)
    if p and p["t"] == stamp:
        k = p.get("kind", "run")
        del PENDING[chat]
        await notify(app, chat, f"{E.BOT} /{k} 逾時未確認，已取消")

async def cmd_confirm(u, c):
    global CHAT_ID; CHAT_ID = u.effective_chat.id
    p = PENDING.get(u.effective_chat.id)
    if not p: await reply(u, f"{E.BOT} 沒有待確認的指令"); return
    if time.time() - p["t"] > 60:
        del PENDING[u.effective_chat.id]; await reply(u, f"{E.BOT} 確認逾時"); return
    kind = p.get("kind", "run")
    if kind == "stop":
        del PENDING[u.effective_chat.id]
        await do_stop(u, p["key"]); return
    if kind == "stopall":
        del PENDING[u.effective_chat.id]
        await do_stopall(u); return
    del PENDING[u.effective_chat.id]
    k = skey(p["sym"], p["dir"])
    # 先確保同 key 沒有殘存的舊 task 還在跑（幽靈策略防護）
    old_t = TASKS.get(k)
    if old_t and not old_t.done():
        old_s = STRATS.get(k)
        if old_s: old_s["alive"] = False
        old_t.cancel()
        try: await asyncio.wait_for(asyncio.shield(old_t), timeout=5)
        except Exception: pass
    S = {**p, "alive": True, "state": "等下輪", "chat": u.effective_chat.id}
    STRATS[k] = S
    TASKS[k] = asyncio.create_task(loop(c.application, u.effective_chat.id, S))
    save_state()
    cnt = sum(1 for s in STRATS.values() if s.get("alive"))
    await reply(u, f"{E.BOT} ✅ 已確認，{p['sym']} {E.dir_word(p['dir'])} 啟動\n運行中策略：{cnt} 個")

async def cmd_stop(u, c):
    a = c.args
    alive = [k for k, s in STRATS.items() if s.get("alive")]
    if not alive: await reply(u, f"{E.BOT} 目前無運行中策略"); return
    if not a:
        lst = "\n".join(f"・/stop {STRATS[k]['sym']} {STRATS[k]['dir']}" for k in alive)
        await reply(u, f"{E.BOT} 請指定：\n{lst}\n或 /stopall"); return
    sym = a[0].upper()
    tg = [skey(sym, a[1].upper())] if len(a) >= 2 and skey(sym, a[1].upper()) in alive else [k for k in alive if STRATS[k]["sym"] == sym]
    if not tg: await reply(u, f"{E.BOT} 找不到運行中的 {sym}"); return
    if len(tg) > 1: await reply(u, f"{E.BOT} {sym} 有多方向，請指定 /stop {sym} L 或 S"); return
    S = STRATS[tg[0]]
    PENDING[u.effective_chat.id] = {"kind": "stop", "t": time.time(), "key": tg[0]}
    await reply(u, f"{E.BOT} 將停止 {E.dir_emoji(S['dir'])} {S['sym']} {E.dir_word(S['dir'])}\n"
                   f"60秒內 /confirm 確認")
    asyncio.create_task(_to(c.application, u.effective_chat.id, PENDING[u.effective_chat.id]["t"]))

async def do_stop(u, key):
    S = STRATS.get(key)
    if not S or not S.get("alive"):
        await reply(u, f"{E.BOT} 策略已不存在"); return
    d = S["dir"]; iid = S["spec"]["iid"]
    ps = "long" if d == "L" else "short"
    p = await okx_pos(iid, ps)
    S["alive"] = False
    n = await sweep(iid, ps)
    save_state()
    tail = f"\n⚠ 持倉 {p['pos']} 張，請至 OKX 平倉" if p else ""
    await reply(u, f"{E.BOT} 已停止 {E.dir_emoji(d)} {S['sym']} {E.dir_word(d)}｜撤單 {n}{tail}")

async def cmd_stopall(u, c):
    alive = [k for k, s in STRATS.items() if s.get("alive")]
    if not alive:
        await reply(u, f"{E.BOT} 目前無運行中策略"); return
    PENDING[u.effective_chat.id] = {"kind": "stopall", "t": time.time()}
    await reply(u, f"{E.BOT} ⚠ 將停止全部 {len(alive)} 個策略\n60秒內 /confirm 確認")
    asyncio.create_task(_to(c.application, u.effective_chat.id, PENDING[u.effective_chat.id]["t"]))

async def do_stopall(u):
    alive = [k for k, s in STRATS.items() if s.get("alive")]
    held = []; done = []
    for k in list(alive):
        S = STRATS[k]; d = S["dir"]; iid = S["spec"]["iid"]
        ps = "long" if d == "L" else "short"
        p = await okx_pos(iid, ps)
        S["alive"] = False
        await sweep(iid, ps)
        (held if p else done).append(f"{S['sym']} {S['dir']}")
    orphan = 0
    for o in await okx_orders(prefix="n"):
        cr = await api("POST", "/api/v5/trade/cancel-order", {"instId": o["instId"], "ordId": o["ordId"]})
        if cr.get("code") == "0": orphan += 1
    save_state()
    m = f"{E.BOT} 已停止 {len(done)} 個策略｜清殘單 {orphan}"
    if held: m += f"\n⚠ 持倉需手動平倉：" + "、".join(held)
    await reply(u, m)

async def cmd_status(u, c):
    global CHAT_ID; CHAT_ID = u.effective_chat.id
    posr = await api("GET", "/api/v5/account/positions")
    pe = await api("GET", "/api/v5/trade/orders-pending")
    bal = await api("GET", "/api/v5/account/balance")
    eq = av = "?"
    if bal.get("code") == "0":
        x = next((d for d in bal["data"][0].get("details", []) if d["ccy"] == "USDT"), None)
        if x:
            eq = f"{Decimal(x.get('eq','0')):.4f}"
            av = f"{Decimal(x.get('availEq') or x.get('availBal') or '0'):.4f}"
    pl = [p for p in posr.get("data", []) if float(p.get("pos", "0")) != 0] if posr.get("code") == "0" else []
    pdl = pe.get("data", []) if pe.get("code") == "0" else []
    alive = [s for s in STRATS.values() if s.get("alive")]
    okxp = {(p["instId"], p["posSide"]) for p in pl}
    okxo = {(o["instId"], o.get("posSide")) for o in pdl}
    L = [f"{E.BOT} OKX原K｜{ACCT}", "事件：現況（即時查 OKX）", "━━━━━━━━━━",
         f"USDT權益：{eq}", f"可用餘額：{av}", f"帳戶週期：{ACCOUNT_TF}",
         f"運行中策略：{len(alive)} 個"]
    for s in alive:
        k = skey(s["sym"], s["dir"]); placed, entered = get_stat(k)
        key = (s["spec"]["iid"], "long" if s["dir"] == "L" else "short")
        live = "持倉中" if key in okxp else ("委託中" if key in okxo else "等下輪")
        L.append(f"{E.dir_emoji(s['dir'])} {s['sym']}：{live}(掛{placed}/進{entered})")
        L.append(f"策略:{strat_params(s['sym'], s['dir'])}")
    L.append(f"掛單數：{len(pdl)}")
    L.append(f"持倉數：{len(pl)}")
    for p in pl:
        d = "L" if p["posSide"] == "long" else "S"
        L.append(f"{E.dir_emoji(d)} {p['instId'].replace('-USDT-SWAP','USDT')} {d}")
    L += ["━━━━━━━━━━", f"時間：{hhmmss()} UTC+8"]
    await reply(u, "\n".join(L))

# ---------- /summary ----------
def sum_lines(rs_all, placed, entered):
    L = []
    pend = [r for r in rs_all if r.get("src") == "PENDING" or r.get("net") is None]
    rs = [r for r in rs_all if r not in pend]
    m = len(rs)
    hit = (entered / placed * 100) if placed else 0
    amb = ("%d秒" % (sum(int(r.get("ambush_s") or 0) for r in rs_all) / len(rs_all))) if rs_all else "-"
    L.append("次數:%d | %d(%s) | %.2f%%" % (placed, entered, amb, hit))
    L.append("帳已入DB:%d | 帳未入DB:%d" % (m, len(pend)))
    if not m:
        if len(pend):
            L.append("%s 損益待 OKX 回報後補入" % E.LOSS)
        else:
            L.append("本日無進場")
        return L
    NAME = {"Take_Profit": "TP", "Stop_Loss": "SL", "Time_Exit": "TE"}
    for lab, cats in (("獲利", ("Take_Profit", "Time_Exit")), ("虧損", ("Stop_Loss", "Time_Exit"))):
        if lab == "獲利":
            sub = [r for r in rs if Decimal(str(r.get("net") or "0")) > 0]
        else:
            sub = [r for r in rs if Decimal(str(r.get("net") or "0")) < 0]
        ps = []
        for cn in cats:
            gg = [r for r in sub if r.get("reason") == cn]
            if gg:
                sec = "%d秒" % (sum(int(r.get("hold_s") or 0) for r in gg) / len(gg))
            else:
                sec = "0秒"
            ps.append("%s:%d(%s)" % (NAME[cn], len(gg), sec))
        L.append("%s數:%d | %s" % (lab, len(sub), " | ".join(ps)))
    tg = sum((Decimal(str(r.get("gross") or "0")) for r in rs), Decimal(0))
    tf = sum((Decimal(str(r.get("fee") or "0")) for r in rs), Decimal(0))
    tn = sum((Decimal(str(r.get("net") or "0")) for r in rs), Decimal(0))
    nv = sum((Decimal(str(r.get("nv") or "0")) for r in rs), Decimal(0))
    gp = (tg / nv * 100) if nv else Decimal(0)
    fp = (tf / nv * 100) if nv else Decimal(0)
    npc = (tn / nv * 100) if nv else Decimal(0)
    L.append("毛損益:%+.6f (%+.3f%%)" % (tg, gp))
    L.append("手續費:%+.6f (%+.3f%%)" % (tf, fp))
    L.append("淨損益:%+.6f (%+.3f%%) %s" % (tn, npc, E.pnl_emoji(tn)))
    return L

async def cmd_summary(u, c):
    t = today8(); recs = load_trades(t)
    ts = {k: v for k, v in STATS.items() if str(v.get("date")) == str(t)}
    L = [f"{E.BOT} OKX原K｜{ACCT}", f"📊📊📊 Summary {t}"]
    for dr in ("L", "S"):
        rows = [r for r in recs if r["dir"] == dr]
        pa = sum(v["placed"] for k, v in ts.items() if k.endswith("_" + dr))
        en = sum(v["entered"] for k, v in ts.items() if k.endswith("_" + dr))
        L.append(f"{E.dir_emoji(dr)} {E.dir_word(dr)}")
        L += sum_lines(rows, pa, en)
    L.append(f"時間:{hhmmss()}")
    await reply(u, "\n".join(L))
    for sy in sorted({r["sym"] for r in recs}):
        D = [f"\U0001f49a\U0001f499\U0001fa75\U0001f49c {sy} {t}"]
        for dr in ("L", "S"):
            rows = [r for r in recs if r["sym"] == sy and r["dir"] == dr]
            st_ = ts.get(skey(sy, dr)) or {"placed": 0, "entered": 0}
            D.append(f"策略:{E.dir_emoji(dr)} {strat_params(sy, dr)}")
            D += sum_lines(rows, st_["placed"], st_["entered"])
        D.append(f"時間:{hhmmss()}")
        await reply(u, "\n".join(D))

# ---------- /amp 振幅報表（Excel + Email） ----------
AMP_MAX = 2000       # 單次最多抓幾根
AMP_BINS = [Decimal(str(x)) for x in
            ("0.1","0.2","0.3","0.4","0.5","0.6","0.7","0.8","0.9","1.0","1.2","1.5")]

async def get_klines_paged(iid, bar, want):
    """分頁往前抓 K 線（OKX 單次上限 300），回傳舊->新、只含已收線。"""
    out = []
    after = ""
    ep = "candles"          # 近期用 candles，翻不動時自動切 history-candles
    for _ in range(40):
        q = f"/api/v5/market/{ep}?instId={iid}&bar={bar}&limit=300"
        if after:
            q += f"&after={after}"
        r = await pub(q)
        if r.get("code") != "0":
            break
        batch = r.get("data") or []
        if not batch:
            if ep == "candles" and after:
                ep = "history-candles"      # candles 只保留近 ~1440 根，改用歷史端點續抓
                continue
            break
        got = []
        for c in batch:
            try:
                if len(c) >= 9 and str(c[8]) != "1":
                    continue
                got.append({"ts": int(c[0]), "o": Decimal(c[1]), "h": Decimal(c[2]),
                            "l": Decimal(c[3]), "c": Decimal(c[4])})
            except Exception:
                continue
        if not got:
            break
        out.extend(got)                      # OKX 回傳為新->舊
        after = str(min(int(x["ts"]) for x in got))
        if len(out) >= want + 10:
            break
        await asyncio.sleep(0.15)
    out.sort(key=lambda x: x["ts"])          # 轉成舊->新
    seen = set(); uniq = []
    for k in out:
        if k["ts"] in seen:
            continue
        seen.add(k["ts"]); uniq.append(k)
    return uniq[-want:] if want else uniq

async def klines_paged_for_tf(iid, tf, want):
    """依 TF 分頁取 K 線。10m 由兩根 5m 合成。"""
    if tf == "10m":
        return _merge2(await get_klines_paged(iid, "5m", want * 2 + 4))
    bar = NATIVE_BARS.get(tf)
    if not bar:
        return None
    return await get_klines_paged(iid, bar, want)

def build_amp_xlsx(sym, tf, kl, amps, path):
    """產生兩個工作表：明細 + 統計。日期與時間分欄，便於樞紐分析。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    FONT = "PingFang TC"; SIZE = 12
    wb = Workbook()

    ws = wb.active; ws.title = "明細"
    heads = ["幣種", "週期", "日期", "時間", "漲跌", "開", "高", "低", "收", "振幅%", "漲跌幅%"]
    widths = {"A": 3.3, "B": 12, "C": 7, "D": 11, "E": 8.5, "F": 6, "G": 13, "H": 13, "I": 13, "J": 13, "K": 11, "L": 11}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for i, h in enumerate(heads):
        c = ws.cell(row=2, column=2 + i, value=h)
        c.font = Font(name=FONT, size=SIZE, bold=True)
        c.alignment = Alignment(horizontal="center")
    r = 3
    for k, (amp, chg) in zip(kl, amps):
        dt = datetime.fromtimestamp(int(k["ts"]) / 1000, TZ8)
        vals = [sym, tf, dt.date(), dt.strftime("%H:%M:%S"),
                "\U0001f7e9" if k["c"] >= k["o"] else "\U0001f7e5",
                float(k["o"]), float(k["h"]), float(k["l"]), float(k["c"]),
                float(amp) / 100, float(chg) / 100]
        for i, v in enumerate(vals):
            cc = ws.cell(row=r, column=2 + i, value=v)
            cc.font = Font(name=FONT, size=SIZE)
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=4).number_format = "m/d/yy"
        ws.cell(row=r, column=5).number_format = "@"
        ws.cell(row=r, column=6).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=11).number_format = "0.0000%"
        ws.cell(row=r, column=12).number_format = '0.0000%;[Red]-0.0000%'
        r += 1
    ws.freeze_panes = "A3"

    st = wb.create_sheet("統計")
    for col, w in {"A": 3.3, "B": 18, "C": 20, "D": 12}.items():
        st.column_dimensions[col].width = w
    seg = [a for a, _ in amps]
    chgs = [c for _, c in amps]
    ss = sorted(seg); m = len(ss)
    med = ss[m // 2] if m % 2 else (ss[m // 2 - 1] + ss[m // 2]) / 2
    avg = sum(seg, Decimal(0)) / m
    d0 = datetime.fromtimestamp(int(kl[0]["ts"]) / 1000, TZ8)
    d1 = datetime.fromtimestamp(int(kl[-1]["ts"]) / 1000, TZ8)
    up = sum(1 for c in chgs if c > 0); dn = sum(1 for c in chgs if c < 0)
    rows = [("商品", sym, None), ("週期", tf, None), ("根數", m, None),
            ("期間起", d0.strftime("%Y-%m-%d %H:%M"), None),
            ("期間迄", d1.strftime("%Y-%m-%d %H:%M"), None),
            ("漲/跌根數", "%d / %d" % (up, dn), None), ("", "", None),
            ("平均振幅", float(avg) / 100, "pct"),
            ("中位振幅", float(med) / 100, "pct"),
            ("最大振幅", float(max(seg)) / 100, "pct"),
            ("最小振幅", float(min(seg)) / 100, "pct"), ("", "", None),
            ("最大漲幅", float(max(chgs)) / 100, "pct"),
            ("最大跌幅", float(min(chgs)) / 100, "pct"), ("", "", None)]
    rr = 2
    for a, b, kind in rows:
        st.cell(row=rr, column=2, value=a).font = Font(name=FONT, size=SIZE)
        cb = st.cell(row=rr, column=3, value=b); cb.font = Font(name=FONT, size=SIZE)
        if kind == "pct":
            cb.number_format = '0.0000%;[Red]-0.0000%'
        rr += 1
    for i, h in enumerate(["振幅達標門檻", "根數", "佔比"]):
        st.cell(row=rr, column=2 + i, value=h).font = Font(name=FONT, size=SIZE, bold=True)
    rr += 1
    for b in AMP_BINS:
        n = sum(1 for a in seg if a >= b)
        st.cell(row=rr, column=2, value="\u2265 " + str(b) + "%").font = Font(name=FONT, size=SIZE)
        st.cell(row=rr, column=3, value=n).font = Font(name=FONT, size=SIZE)
        cd = st.cell(row=rr, column=4, value=n / m)
        cd.font = Font(name=FONT, size=SIZE); cd.number_format = "0.00%"
        rr += 1
    wb.save(path)

def send_amp_mail(path, name, sym, tf, m, avg, med):
    """寄出振幅報表。回傳 (ok, 訊息)。"""
    import smtplib
    from email.message import EmailMessage
    env = {}
    try:
        for line in open("/srv/1111bot/.env"):
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip().strip('"').strip("'")
    except Exception as e:
        return False, "讀 .env 失敗：%s" % e
    user = env.get("GMAIL_USER")
    pwd = (env.get("GMAIL_APP_PASSWORD") or "").replace(" ", "")
    to = env.get("REPORT_TO") or user
    if not user or not pwd:
        return False, "未設定 GMAIL_USER / GMAIL_APP_PASSWORD"
    msg = EmailMessage()
    msg["Subject"] = "OKX %s 振幅報表 %s %s（%d 根）" % (ACCT, sym, tf, m)
    msg["From"] = user; msg["To"] = to
    msg.set_content("商品 %s\\n週期 %s\\n根數 %d\\n平均振幅 %.4f%%\\n中位振幅 %.4f%%\\n" % (sym, tf, m, avg, med))
    msg.add_attachment(open(path, "rb").read(),
                       maintype="application",
                       subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       filename=name)
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as sv:
        sv.login(user, pwd); sv.send_message(msg)
    return True, to

async def cmd_amp(u, c):
    """原K 振幅報表（Excel 寄信）。用法：/amp SOLUSDT [根數 3~2000]"""
    supported = "3m/5m/10m"
    if not c.args:
        await reply(u, f"{E.BOT} 用法：/amp SOLUSDT 900\n"
                       f"根數 3~{AMP_MAX}（預設 300）\n"
                       f"產生 Excel（明細＋統計）寄到信箱\n"
                       f"支援週期：{supported}\n"
                       f"目前週期：{ACCOUNT_TF}")
        return
    sym = c.args[0].upper()
    n = 300
    if len(c.args) >= 2:
        try: n = max(3, min(AMP_MAX, int(c.args[1])))
        except Exception: pass
    if ACCOUNT_TF != "10m" and ACCOUNT_TF not in NATIVE_BARS:
        await reply(u, f"{E.BOT} 目前週期 {ACCOUNT_TF} 無原生 K 線\n"
                       f"可查週期：{supported}\n請先 /timeframe 切換")
        return
    try:
        spec = await get_spec(sym)
    except Exception:
        await reply(u, f"{E.LOSS} 找不到商品 {sym}"); return
    await reply(u, f"{E.BOT} 產生 {sym} {ACCOUNT_TF} 振幅報表中（{n} 根），請稍候…")
    try:
        kl = await klines_paged_for_tf(spec["iid"], ACCOUNT_TF, n)
    except Exception as e:
        await reply(u, f"{E.LOSS} K 線取得失敗：{type(e).__name__}"); return
    if not kl:
        await reply(u, f"{E.BOT} {sym} K 線取得失敗"); return
    amps = calc_amp(kl)
    m = len(kl)
    seg = [a for a, _ in amps]
    ss = sorted(seg)
    med = ss[m // 2] if m % 2 else (ss[m // 2 - 1] + ss[m // 2]) / 2
    avg = sum(seg, Decimal(0)) / m
    day = now8().strftime("%Y%m%d")
    name = f"OKX_{ACCT}_振幅_{sym}_{ACCOUNT_TF}_{m}根_{day}.xlsx"
    path = f"/srv/1111bot/data/{name}"
    try:
        build_amp_xlsx(sym, ACCOUNT_TF, kl, amps, path)
    except Exception as e:
        await reply(u, f"{E.LOSS} 產生 Excel 失敗：{type(e).__name__}: {e}"); return
    try:
        ok, info = send_amp_mail(path, name, sym, ACCOUNT_TF, m, float(avg), float(med))
    except Exception as e:
        await reply(u, f"{E.LOSS} 寄送失敗：{type(e).__name__}: {e}\n檔案已存於 VPS：{name}"); return
    if not ok:
        await reply(u, f"{E.LOSS} 未寄送：{info}\n檔案已存於 VPS：{name}"); return
    t0 = datetime.fromtimestamp(int(kl[0]["ts"]) / 1000, TZ8).strftime("%m/%d %H:%M")
    t1 = datetime.fromtimestamp(int(kl[-1]["ts"]) / 1000, TZ8).strftime("%m/%d %H:%M")
    await reply(u, f"{E.BOT} ✅ 振幅報表已寄出\n"
                   f"{sym} {ACCOUNT_TF}｜{m} 根\n"
                   f"期間：{t0} ~ {t1}\n"
                   f"平均 {avg:.4f}% | 中位 {med:.4f}%\n"
                   f"最大 {max(seg):.4f}% | 最小 {min(seg):.4f}%\n"
                   f"時間：{hhmmss()}")

async def cmd_audit(u, c):
    """對帳：本地紀錄 vs OKX positions-history。用法：/audit [YYYYMMDD]"""
    day = c.args[0] if c.args else now8().strftime("%Y%m%d")
    t = day[:4] + "-" + day[4:6] + "-" + day[6:8]
    recs = load_trades(t)
    if not recs:
        await reply(u, f"{E.BOT} {t} 無本地交易紀錄"); return
    await reply(u, f"{E.BOT} 對帳中：{t}，本地 {len(recs)} 筆…")
    s8 = now8().replace(hour=0, minute=0, second=0, microsecond=0)
    try:
        s8 = s8.replace(year=int(day[:4]), month=int(day[4:6]), day=int(day[6:8]))
    except Exception:
        pass
    sms = int(s8.timestamp() * 1000)
    ph = []; after = ""; pages = 0
    while pages < 20:
        q = "/api/v5/account/positions-history?instType=SWAP&limit=100"
        if after:
            q += "&after=" + after
        r = await api("GET", q)
        if r.get("code") != "0":
            break
        batch = r.get("data") or []
        if not batch:
            break
        pages += 1
        stop = False
        for p in batch:
            ut = int(p.get("uTime") or 0)
            if ut >= sms and ut < sms + 86400000:
                ph.append(p)
            elif ut < sms:
                stop = True
        if stop or len(batch) < 100:
            break
        after = batch[-1].get("posId") or ""
        if not after:
            break
    ln = sum(Decimal(str(x.get("net") or "0")) for x in recs)
    lf = sum(Decimal(str(x.get("fee") or "0")) for x in recs)
    on = sum(Decimal(p.get("realizedPnl") or "0") for p in ph)
    of = sum(Decimal(p.get("fee") or "0") + Decimal(p.get("fundingFee") or "0") for p in ph)
    from collections import Counter
    srcs = Counter(x.get("src") for x in recs)
    L = [f"{E.BOT} OKX原K｜{ACCT}", f"📋 對帳 {t}", "━" * 10,
         ("%s " % E.LOSS if len(recs) != len(ph) else "") + f"DB筆數：{len(recs)}｜OKX：{len(ph)}",
         "來源分佈：" + "、".join(f"{k}×{v}" for k, v in srcs.items()),
         "━" * 10,
         f"DB 淨損益：{ln:+.6f}",
         f"OKX 淨損益：{on:+.6f}",
         f"差異：{ln - on:+.6f}",
         "━" * 10,
         f"DB 手續費：{lf:+.6f}",
         f"OKX 手續費：{of:+.6f}",
         f"差異：{lf - of:+.6f}",
         "━" * 10]
    if len(recs) == len(ph) and abs(ln - on) < Decimal("0.000001") and abs(lf - of) < Decimal("0.000001"):
        L.append("✅ 完全一致")
    zero = [x for x in recs if Decimal(str(x.get("fee") or "0")) == 0]
    if zero:
        L.append(f"⚠ 手續費為 0 的紀錄：{len(zero)} 筆")
        for x in zero[:5]:
            L.append(f"　{x['ts']} {x['sym']} {x['dir']} src={x.get('src')}")
    L.append(f"時間：{hhmmss()}")
    await reply(u, "\n".join(L))

async def cmd_coins(u, c):
    on = [s["symbol"] for s in SYMS if s["enabled"]]
    L = [f"{E.BOT} OKX原K｜{ACCT}", "事件：幣種清單（即時）", "━━━━━━━━━━"]
    for sym in on:
        try:
            sp = await get_spec(sym); last = await get_last(sp["iid"])
            mm = sp["minsz"] * sp["ctval"] * last
            L.append(f"{sym}｜最小{sp['minsz']}張｜{mm:.4f}U")
        except Exception:
            L.append(f"{sym}｜查詢失敗")
    L += ["━━━━━━━━━━", f"時間：{hhmmss()}"]
    await reply(u, "\n".join(L))

async def cmd_timeframe(u, c):
    global ACCOUNT_TF
    if not c.args:
        await reply(u, f"{E.BOT} 目前週期：{ACCOUNT_TF}\n可選：" + "/".join(TF_SEC.keys()) + "\n變更：/timeframe 10m"); return
    tf = c.args[0]
    if tf not in TF_SEC: await reply(u, f"{E.BOT} 週期須為：" + "/".join(TF_SEC.keys())); return
    ACCOUNT_TF = tf; save_state()
    await reply(u, f"{E.BOT} ✅ 帳戶週期已設為 {tf}\n（僅影響之後新建立的策略）")

async def cmd_menu(u, c):
    await reply(u, f"{E.BOT} OKX原K｜{ACCT}\n使用說明\n━━━━━━━━━━\n"
        "/run 商品 方向 槓桿 保證金 埋伏 TP SL\n"
        f"例：/run ETHUSDT L 1x 3 0.5 0.5 0.5\n　週期依 /timeframe（目前 {ACCOUNT_TF}）\n"
        "/confirm 確認啟動\n/stop 商品 方向\n/stopall 停全部+清殘單\n"
        "/status 所有策略現況\n/summary 當日戰報\n"
        "/amp 商品 根數  振幅報表 Excel 寄信（3~2000根）\n"
        "/audit [YYYYMMDD]  與 OKX 對帳\n"
        "/timeframe 查看/設定週期\n/coins 幣種\n"
        "━━━━━━━━━━\n"
        f"一個 TF 一輪：TF 開始埋伏\n"
        f"未成交且剩餘不足 {ENTRY_CUTOFF}s → 撤單放棄本輪\n"
        f"已進場未觸發 TP/SL → 持倉滿 {HOLD_SEC}s 平倉（TE）\n"
        "⚠ 真實下單，循環交易\n✅ 重啟接管持倉與掛單")

async def cmd_unknown(u, c):
    await reply(u, f"{E.BOT} 指令無法辨識：{u.message.text}\n請用 /menu")

# ---------- 每日自動 summary ----------
class _M:
    def __init__(self, app, chat): self._a = app; self._c = chat
    async def reply_text(self, t): await self._a.bot.send_message(self._c, t)
class _U:
    def __init__(self, app, chat): self.message = _M(app, chat)

async def job_summary(ctx):
    if not CHAT_ID: return
    try: await cmd_summary(_U(ctx.application, CHAT_ID), ctx)
    except Exception as e: print("auto summary fail", e)

# ---------- 啟動 ----------
async def _post_init(app):
    global HTTP
    HTTP = httpx.AsyncClient(timeout=httpx.Timeout(20.0, connect=10.0), limits=httpx.Limits(max_connections=40))
    CMDS = [BotCommand("status", "現況"),
            BotCommand("summary", "當日戰報"),
            BotCommand("coins", "幣種"),
            BotCommand("amp", "振幅報表 Excel"),
            BotCommand("audit", "對帳 OKX"),
            BotCommand("stopall", "停全部"),
            BotCommand("stop", "停指定"),
            BotCommand("run", "建立策略"),
            BotCommand("timeframe", "週期"),
            BotCommand("menu", "說明")]
    # 清除所有 scope 的舊指令（ThisChat/AllPrivateChats 優先權高於 Default，
    # 只刪 Default 會被舊清單蓋住，導致左下 Menu 卡在舊版）
    scopes = [BotCommandScopeDefault(), BotCommandScopeAllPrivateChats()]
    try:
        saved = json.load(open(STATE_FILE)) if os.path.exists(STATE_FILE) else {}
        ch = saved.get("chat")
        if ch: scopes.append(BotCommandScopeChat(ch))
    except Exception:
        pass
    for sc in scopes:
        try: await app.bot.delete_my_commands(scope=sc)
        except Exception as e: print("delete cmds fail", type(sc).__name__, e)
    await app.bot.set_my_commands(CMDS)
    print(f"左下 Menu 已更新（已清除 {len(scopes)} 個 scope 的舊指令）")
    try:
        jq = app.job_queue
        if jq:
            t2359 = datetime.strptime("23:59", "%H:%M").time().replace(tzinfo=TZ8)
            jq.run_daily(job_summary, time=t2359, name="daily_summary")
            print("已排程：每日 23:59 自動 /summary")
    except Exception as e:
        print("schedule fail", e)
    await startup_recover(app)
    _rc = asyncio.create_task(reconcile_watch(app))
    _BG.add(_rc); _rc.add_done_callback(_BG.discard)
    print("補帳背景任務已啟動（每 5 分鐘）")

async def _post_stop(app):
    global SHUTTING_DOWN
    save_state()          # 關閉前最後一次完整存檔（此時 STRATS 仍完整）
    SHUTTING_DOWN = True
    print("關閉中：已保存狀態，停止後續寫檔")

def main():
    os.makedirs(os.path.dirname(STATE_FILE), exist_ok=True)
    print(f"啟動 o3333o 原K B6-1 核心重寫版（token ...{TOKEN[-6:]}）")
    app = (Application.builder().token(TOKEN).post_init(_post_init).post_stop(_post_stop)
           .connect_timeout(30.0).read_timeout(30.0).write_timeout(30.0)
           .pool_timeout(30.0).get_updates_read_timeout(40.0)
           .get_updates_connect_timeout(30.0).build())
    for cmd, fn in [(["menu", "start"], cmd_menu), ("run", cmd_run), ("confirm", cmd_confirm),
                    ("stop", cmd_stop), ("stopall", cmd_stopall), ("status", cmd_status),
                    ("summary", cmd_summary), ("amp", cmd_amp), ("audit", cmd_audit),
                    ("timeframe", cmd_timeframe), ("coins", cmd_coins)]:
        app.add_handler(CommandHandler(cmd, fn))
    app.add_handler(MessageHandler(filters.COMMAND, cmd_unknown))
    app.run_polling()

if __name__ == "__main__":
    main()
